"""Generate FBA invoices in Excel and PDF formats."""
import io
import math
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def num_to_words_indian(num: float) -> str:
    """Convert number to Indian currency words."""
    if num == 0:
        return "Zero"

    ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine",
            "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen",
            "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]

    def two_digits(n):
        if n < 20:
            return ones[n]
        return tens[n // 10] + (" " + ones[n % 10] if n % 10 else "")

    def three_digits(n):
        if n >= 100:
            return ones[n // 100] + " Hundred" + (" and " + two_digits(n % 100) if n % 100 else "")
        return two_digits(n)

    rupees = int(num)
    paise = round((num - rupees) * 100)

    if rupees == 0:
        result = ""
    else:
        parts = []
        if rupees >= 10000000:
            parts.append(two_digits(rupees // 10000000) + " Crore")
            rupees %= 10000000
        if rupees >= 100000:
            parts.append(two_digits(rupees // 100000) + " Lakh")
            rupees %= 100000
        if rupees >= 1000:
            parts.append(two_digits(rupees // 1000) + " Thousand")
            rupees %= 1000
        if rupees > 0:
            parts.append(three_digits(rupees))
        result = " ".join(parts)

    if paise:
        return f"Rupees {result} and {two_digits(paise)} Paise Only"
    return f"Rupees {result} Only"


def generate_excel_invoice(invoice_data: dict) -> io.BytesIO:
    """
    Generate Excel invoice matching the GST stock transfer format.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "GST Invoice"

    # Styling
    bold = Font(bold=True)
    bold_large = Font(bold=True, size=14)
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    supplier = invoice_data.get("supplier", {})
    recipient = invoice_data.get("recipient", {})
    items = invoice_data.get("items", [])
    details = invoice_data.get("details", {})

    # Title
    ws.merge_cells("A1:L1")
    ws["A1"] = "TAX INVOICE / STOCK TRANSFER"
    ws["A1"].font = bold_large
    ws["A1"].alignment = center

    # Supplier / Recipient
    ws["A3"] = "Supplier / From"
    ws["A3"].font = bold
    ws["G3"] = "Recipient / Ship To"
    ws["G3"].font = bold

    ws["A4"] = "Name"
    ws["B4"] = supplier.get("name", "")
    ws["A5"] = "Address"
    ws["B5"] = supplier.get("address", "")
    ws["A6"] = "GST No."
    ws["B6"] = supplier.get("gstin", "")

    ws["G4"] = "Name"
    ws["H4"] = recipient.get("name", "")
    ws["G5"] = "Address"
    ws["H5"] = recipient.get("address", "")
    ws["G6"] = "GST No."
    ws["H6"] = recipient.get("gstin", "")

    # Invoice Details
    ws["A8"] = "Invoice Details"
    ws["A8"].font = bold
    ws["A9"] = "Invoice No."
    ws["B9"] = details.get("invoice_no", "")
    ws["C9"] = "Date"
    ws["D9"] = details.get("date", datetime.now().strftime("%d/%m/%Y"))
    ws["E9"] = "Shipment ID"
    ws["F9"] = details.get("shipment_id", "")
    ws["G9"] = "Place of Supply"
    ws["H9"] = details.get("place_of_supply", "")

    # Transport Details
    ws["A11"] = "Transport Details"
    ws["A11"].font = bold
    ws["A12"] = "Transporter"
    ws["B12"] = details.get("transporter", "")
    ws["E12"] = "E-Way Bill No."
    ws["F12"] = details.get("eway_bill", "")
    ws["G12"] = "Boxes"
    ws["H12"] = details.get("boxes", "")
    ws["I12"] = "Weight"
    ws["J12"] = details.get("weight", "")

    # Table Header
    row = 14
    headers = ["S.No", "Merchant SKU", "Title", "HSN/SAC", "GST Rate", "ASIN", "FNSKU",
               "Rate", "Qty", "Taxable Value", "IGST", "IGST Value", "Total"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # Data rows
    grand_taxable = 0
    grand_igst = 0
    grand_total = 0
    total_qty = 0

    for i, item in enumerate(items, 1):
        row += 1
        qty = item["quantity"]
        rate = item.get("rate", 0)
        gst_rate = item.get("gst_rate", 5)
        taxable = round(qty * rate, 2)
        igst = round(taxable * gst_rate / 100, 2)
        total = round(taxable + igst, 2)

        grand_taxable += taxable
        grand_igst += igst
        grand_total += total
        total_qty += qty

        values = [
            i, item["sku"], item.get("short_title", item["title"]),
            item["hsn_code"], f"{gst_rate}%", item.get("asin", ""),
            item.get("fnsku", ""), rate, qty, taxable,
            f"{gst_rate}%", igst, total
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center if col != 3 else left_wrap

    # Totals row
    row += 1
    ws.cell(row=row, column=8, value="Total").font = bold
    ws.cell(row=row, column=9, value=total_qty).font = bold
    ws.cell(row=row, column=10, value=round(grand_taxable, 2)).font = bold
    ws.cell(row=row, column=12, value=round(grand_igst, 2)).font = bold
    ws.cell(row=row, column=13, value=round(grand_total, 2)).font = bold
    for col in range(1, 14):
        ws.cell(row=row, column=col).border = thin_border

    # Amount in words
    row += 2
    ws.cell(row=row, column=1, value="Amount in Words:")
    ws.cell(row=row, column=2, value=num_to_words_indian(grand_total))
    ws.cell(row=row, column=2).font = bold

    # Footer
    row += 2
    ws.cell(row=row, column=1, value="We declare that the above mentioned goods are being moved for branch/stock transfer.")
    row += 1
    ws.cell(row=row, column=1, value="The above mentioned details are true and correct to the best of our knowledge.")

    row += 2
    ws.cell(row=row, column=1, value="Checked / Verified")
    ws.cell(row=row, column=5, value="Approved By")
    ws.cell(row=row, column=9, value=f"For {supplier.get('name', '')}")
    row += 2
    ws.cell(row=row, column=1, value="Signature")
    ws.cell(row=row, column=5, value="Signature")
    ws.cell(row=row, column=9, value="Authorised Signatory")

    # Column widths
    col_widths = [5, 18, 40, 8, 8, 12, 12, 8, 5, 12, 6, 10, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_pdf_invoice(invoice_data: dict) -> io.BytesIO:
    """
    Generate PDF invoice using the detailed format (matching the reference PDF).
    Uses reportlab if available, otherwise generates from HTML.
    """
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm, cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        return _generate_pdf_reportlab(invoice_data)
    except ImportError:
        return _generate_pdf_simple(invoice_data)


def _generate_pdf_reportlab(invoice_data: dict) -> io.BytesIO:
    """Generate PDF using reportlab."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=8*mm, rightMargin=8*mm,
                            topMargin=8*mm, bottomMargin=8*mm)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("Title", parent=styles["Heading1"], fontSize=13, alignment=TA_CENTER)
    small = ParagraphStyle("Small", parent=styles["Normal"], fontSize=6, leading=7.5)
    tiny = ParagraphStyle("Tiny", parent=styles["Normal"], fontSize=5.5, leading=7)

    supplier = invoice_data.get("supplier", {})
    recipient = invoice_data.get("recipient", {})
    items = invoice_data.get("items", [])
    details = invoice_data.get("details", {})

    # Extract just the state name for Place of Supply
    place_of_supply = details.get("place_of_supply", "")
    # If it contains a full address, try to extract just the state
    for state in ["Maharashtra", "Karnataka", "Tamil Nadu", "Telangana", "Delhi",
                  "Haryana", "Gujarat", "Rajasthan", "Uttar Pradesh", "West Bengal",
                  "Bihar", "Jharkhand", "Odisha", "Assam", "Punjab", "Kerala", "Goa",
                  "Madhya Pradesh", "Chhattisgarh", "Andhra Pradesh"]:
        if state.lower() in place_of_supply.lower():
            place_of_supply = state
            break

    # Title
    elements.append(Paragraph("TAX INVOICE / STOCK TRANSFER", title_style))
    elements.append(Spacer(1, 2*mm))

    # Get the full FC address for Shipped To
    shipped_to_address = recipient.get("address", place_of_supply)

    # Header info table — compact with boxes/weight at top
    header_data = [
        ["Shipment ID", details.get("shipment_id", ""), "Invoice No.", details.get("invoice_no", ""),
         "Boxes", details.get("boxes", ""), "Weight", details.get("weight", "")],
        ["GSTIN (From)", supplier.get("gstin", ""), "GSTIN (To)", recipient.get("gstin", ""),
         "Date", details.get("date", ""), "Place of Supply", place_of_supply],
        ["Shipped From", Paragraph(supplier.get("name", "") + ", " + supplier.get("address", ""), tiny),
         "Shipped To", Paragraph(shipped_to_address, tiny),
         "Transporter", details.get("transporter", "") or "-", "E-Way Bill", details.get("eway_bill", "") or "-"],
    ]
    ht = Table(header_data, colWidths=[22*mm, 52*mm, 20*mm, 52*mm, 18*mm, 30*mm, 22*mm, 42*mm])
    ht.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTNAME", (4, 0), (4, -1), "Helvetica-Bold"),
        ("FONTNAME", (6, 0), (6, -1), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(ht)
    elements.append(Spacer(1, 3*mm))

    # Items table — with GST Rate column
    table_header = ["#", "SKU", "Title", "HSN", "GST%", "ASIN", "FNSKU",
                    "Rate", "Qty", "Taxable", "IGST", "Total"]

    table_data = [table_header]
    grand_taxable = 0
    grand_igst = 0
    grand_total = 0
    total_qty = 0

    for i, item in enumerate(items, 1):
        qty = item["quantity"]
        rate = item.get("rate", 0)
        gst_rate = item.get("gst_rate", 5)
        taxable = round(qty * rate, 2)
        igst = round(taxable * gst_rate / 100, 2)
        total = round(taxable + igst, 2)
        grand_taxable += taxable
        grand_igst += igst
        grand_total += total
        total_qty += qty

        # Title: 12 words for better readability
        short_title = " ".join(item["title"].split()[:12])

        table_data.append([
            str(i),
            Paragraph(item["sku"], tiny),
            Paragraph(short_title, small),
            item["hsn_code"],
            f"{gst_rate}%",
            item.get("asin", ""),
            item.get("fnsku", ""),
            str(rate), str(qty), str(taxable),
            str(igst), str(total)
        ])

    # Totals row
    table_data.append(["", "", "", "", "", "", "Total",
                       "", str(total_qty), str(round(grand_taxable, 2)),
                       str(round(grand_igst, 2)), str(round(grand_total, 2))])

    # Column widths with GST% column added
    col_widths = [7*mm, 22*mm, 62*mm, 10*mm, 9*mm, 17*mm, 17*mm,
                  12*mm, 8*mm, 17*mm, 14*mm, 17*mm]

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 6.5),
        ("FONTSIZE", (0, 1), (-1, -1), 6),
        ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.2, 0.3, 0.3)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (1, 1), (1, -1), "LEFT"),  # SKU left
        ("ALIGN", (2, 1), (2, -1), "LEFT"),  # Title left
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.Color(0.95, 0.95, 0.95)),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 3*mm))

    # Footer
    gst_pct = "5%"  # All products at 5%
    footer_data = [
        [f"Amount in Words: {num_to_words_indian(grand_total)}", "", ""],
        ["", "", ""],
        ["We declare that the above goods are being moved for branch/stock transfer.", "", f"For {supplier.get('name', '')}"],
        ["The details are true and correct to the best of our knowledge.", "", ""],
        ["", "", "Authorised Signatory"],
    ]
    ft = Table(footer_data, colWidths=[120*mm, 60*mm, 80*mm])
    ft.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("FONTNAME", (0, 0), (0, 0), "Helvetica-Bold"),
        ("ALIGN", (2, 2), (2, -1), "RIGHT"),
    ]))
    elements.append(ft)

    doc.build(elements)
    buffer.seek(0)
    return buffer


def _generate_pdf_simple(invoice_data: dict) -> io.BytesIO:
    """Fallback: generate a simple text-based representation if reportlab not available."""
    buffer = io.BytesIO()
    buffer.write(b"PDF generation requires reportlab. Install with: pip install reportlab\n")
    buffer.seek(0)
    return buffer
