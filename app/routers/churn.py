"""Product Churn Analysis — upload Business Report CSV, score every ASIN."""
import io
import re
import logging
from datetime import datetime, timedelta

import pandas as pd
from fastapi import APIRouter, Request, Depends, UploadFile, File, Form
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, desc

from app.database import get_db
from app.models import (
    Product, BSRHistory, RatingHistory, ChurnReport, ChurnScore
)
from app.routers.auth import require_auth

router = APIRouter(prefix="/churn")
logger = logging.getLogger(__name__)


# ─── CSV Parsing ─────────────────────────────────────────────────────────────

def _clean_number(val: str) -> float:
    """Strip ₹, commas, % and return float. Returns 0 on failure."""
    if not val or str(val).strip() in ("", "-"):
        return 0.0
    cleaned = re.sub(r"[₹,%\s]", "", str(val))
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def parse_business_report(content: bytes) -> dict[str, dict]:
    """
    Parse Amazon Business Report CSV.
    Returns {asin: {sessions, units_ordered, revenue, conversion_rate, buy_box_pct, title}}
    Deduplicates ASINs: sums units/revenue/sessions, takes max conversion/buybox.
    """
    try:
        df = pd.read_csv(io.BytesIO(content))
    except Exception:
        df = pd.read_csv(io.BytesIO(content), encoding="utf-8-sig")

    asin_col = "(Child) ASIN"
    if asin_col not in df.columns:
        raise ValueError("Could not find '(Child) ASIN' column in CSV")

    result: dict[str, dict] = {}

    for _, row in df.iterrows():
        asin = str(row.get(asin_col, "")).strip()
        if not asin or len(asin) < 10:
            continue

        sessions = _clean_number(row.get("Sessions - Total", 0))
        units = _clean_number(row.get("Units Ordered", 0))
        revenue = _clean_number(row.get("Ordered Product Sales", 0))
        conv = _clean_number(row.get("Unit Session Percentage", 0))
        buybox = _clean_number(row.get("Featured Offer Percentage", 0))
        title = str(row.get("Title", "")).strip()

        if asin in result:
            result[asin]["sessions"] += sessions
            result[asin]["units_ordered"] += units
            result[asin]["revenue"] += revenue
            result[asin]["conversion_rate"] = max(result[asin]["conversion_rate"], conv)
            result[asin]["buy_box_pct"] = max(result[asin]["buy_box_pct"], buybox)
        else:
            result[asin] = {
                "sessions": sessions,
                "units_ordered": int(units),
                "revenue": revenue,
                "conversion_rate": conv,
                "buy_box_pct": buybox,
                "title": title,
            }

    return result


# ─── Scoring ──────────────────────────────────────────────────────────────────

def _revenue_score(revenue: float, all_revenues: list[float]) -> int:
    """0-20: rank within all products by revenue."""
    if not all_revenues or revenue == 0:
        return 0
    pct = sum(1 for r in all_revenues if r <= revenue) / len(all_revenues)
    return int(pct * 20)


def _conversion_score(conv: float) -> int:
    if conv >= 15:
        return 20
    if conv >= 10:
        return 15
    if conv >= 5:
        return 10
    if conv > 0:
        return 5
    return 0


def _rating_score(rating: float | None) -> int:
    if rating is None:
        return 10  # neutral
    if rating >= 4.3:
        return 20
    if rating >= 4.0:
        return 15
    if rating >= 3.5:
        return 10
    return 0


def _bsr_trend_score(bsr_now: int | None, bsr_15d: int | None) -> tuple[int, str]:
    """Returns (score 0-20, trend label)."""
    if bsr_now is None:
        return 10, "unknown"
    if bsr_15d is None:
        return 10, "unknown"
    change_pct = (bsr_now - bsr_15d) / bsr_15d * 100
    if change_pct <= -10:   # BSR went down (improved)
        return 20, "improving"
    if change_pct <= 10:    # Within ±10% (stable)
        return 12, "stable"
    return 0, "declining"   # BSR went up (worse)


def _review_score(count: int | None) -> int:
    if count is None:
        return 5
    if count >= 50:
        return 20
    if count >= 20:
        return 15
    if count >= 5:
        return 10
    return 5  # new product — penalise less


def _classify(score: int, listing_age_days: int | None) -> str:
    """Determine status, applying age override."""
    age = listing_age_days or 999
    if score >= 70:
        return "keep"
    if score >= 40:
        return "monitor"
    # Churn candidate — but override if listing is < 60 days old
    if age < 60:
        return "monitor"
    return "churn"


def _build_reason(
    units: int, conv: float, rating: float | None,
    bsr_trend: str, review_count: int | None, listing_age: int | None,
    status: str,
) -> str:
    parts = []
    age = listing_age or 0

    if status == "no_data":
        return "Not found in uploaded Business Report (zero or no sales recorded)"

    if units == 0:
        parts.append("0 units sold in 30 days")
    elif units < 10:
        parts.append(f"Very low sales ({units} units)")

    if conv > 0 and conv < 5:
        parts.append(f"Low conversion ({conv:.1f}%)")
    elif conv >= 15:
        parts.append(f"Strong conversion ({conv:.1f}%)")

    if rating is not None:
        if rating < 3.5:
            parts.append(f"Poor rating ({rating}★)")
        elif rating >= 4.3:
            parts.append(f"Excellent rating ({rating}★)")

    if bsr_trend == "declining":
        parts.append("BSR worsening")
    elif bsr_trend == "improving":
        parts.append("BSR improving")

    if review_count is not None and review_count < 5:
        parts.append(f"Only {review_count} reviews")

    if age > 0 and age < 60:
        parts.append(f"New listing ({age}d old) — give it time")

    if not parts:
        if status == "keep":
            return "Strong performer across all metrics"
        return "Average performance — monitor closely"

    return "; ".join(parts)


# ─── DB Helpers ───────────────────────────────────────────────────────────────

async def _get_latest_rating(db: AsyncSession, product_id: int) -> RatingHistory | None:
    return (await db.execute(
        select(RatingHistory)
        .where(RatingHistory.product_id == product_id)
        .order_by(desc(RatingHistory.scraped_at))
        .limit(1)
    )).scalar_one_or_none()


async def _get_bsr_now_and_15d(
    db: AsyncSession, product_id: int
) -> tuple[int | None, int | None]:
    """Returns (current_bsr_rank, bsr_rank_~15d_ago)."""
    now = (await db.execute(
        select(BSRHistory)
        .where(BSRHistory.product_id == product_id)
        .order_by(desc(BSRHistory.scraped_at))
        .limit(1)
    )).scalar_one_or_none()

    cutoff = datetime.utcnow() - timedelta(days=12)
    old = (await db.execute(
        select(BSRHistory)
        .where(
            BSRHistory.product_id == product_id,
            BSRHistory.scraped_at <= cutoff,
        )
        .order_by(desc(BSRHistory.scraped_at))
        .limit(1)
    )).scalar_one_or_none()

    return (
        now.bsr_rank if now else None,
        old.bsr_rank if old else None,
    )


# ─── Endpoints ────────────────────────────────────────────────────────────────

@router.post("/upload")
async def upload_report(
    request: Request,
    file: UploadFile = File(...),
    period_label: str = Form(""),
    _=Depends(require_auth),
    db: AsyncSession = Depends(get_db),
):
    """Upload Amazon Business Report CSV → run churn analysis → save to DB."""
    content = await file.read()
    try:
        csv_data = parse_business_report(content)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=400)

    # Use submitted period_label, fallback to filename-derived label
    if not period_label:
        fname = file.filename or ""
        period_label = fname.replace("BusinessReport-", "").replace(".csv", "") or "Unknown Period"

    # Fetch all active products from DB
    all_products = (await db.execute(
        select(Product).where(Product.is_active == True)
    )).scalars().all()

    # Build lookup: asin → Product
    db_products: dict[str, Product] = {p.asin: p for p in all_products}

    # All ASINs = CSV ASINs + DB-only ASINs not in CSV
    all_asins = set(csv_data.keys()) | set(db_products.keys())

    # Collect all revenues to compute relative ranking
    all_revenues = [v["revenue"] for v in csv_data.values()]

    scores: list[ChurnScore] = []

    for asin in all_asins:
        csv_row = csv_data.get(asin)
        product = db_products.get(asin)

        # --- Sales data from CSV ---
        units = int(csv_row["units_ordered"]) if csv_row else 0
        revenue = csv_row["revenue"] if csv_row else 0.0
        conv = csv_row["conversion_rate"] if csv_row else 0.0
        sessions = int(csv_row["sessions"]) if csv_row else 0
        buybox = csv_row["buy_box_pct"] if csv_row else 0.0
        title = (csv_row["title"] if csv_row else None) or (product.title if product else asin)

        # --- Scraped data from DB ---
        rating_val = None
        review_count = None
        bsr_now = None
        bsr_15d = None
        listing_age = None

        if product:
            latest_rating = await _get_latest_rating(db, product.id)
            if latest_rating:
                rating_val = float(latest_rating.rating) if latest_rating.rating else None
                review_count = latest_rating.rating_count

            bsr_now, bsr_15d = await _get_bsr_now_and_15d(db, product.id)

            if product.first_seen:
                listing_age = (datetime.utcnow() - product.first_seen).days

        # --- No sales data case ---
        if csv_row is None:
            score = ChurnScore(
                asin=asin,
                title=title,
                score=0,
                status="no_data",
                units_ordered=0,
                revenue=0,
                conversion_rate=0,
                sessions=0,
                buy_box_pct=0,
                rating=rating_val,
                review_count=review_count,
                bsr_current=bsr_now,
                bsr_trend="unknown",
                listing_age_days=listing_age,
                reason="Not found in uploaded Business Report (zero or no sales recorded)",
            )
            scores.append(score)
            continue

        # --- Score calculation ---
        s_revenue = _revenue_score(revenue, all_revenues)
        s_conv = _conversion_score(conv)
        s_rating = _rating_score(rating_val)
        s_bsr, bsr_trend = _bsr_trend_score(bsr_now, bsr_15d)
        s_review = _review_score(review_count)

        total_score = s_revenue + s_conv + s_rating + s_bsr + s_review
        status = _classify(total_score, listing_age)
        reason = _build_reason(units, conv, rating_val, bsr_trend, review_count, listing_age, status)

        score = ChurnScore(
            asin=asin,
            title=title,
            score=total_score,
            status=status,
            units_ordered=units,
            revenue=revenue,
            conversion_rate=conv,
            sessions=sessions,
            buy_box_pct=buybox,
            rating=rating_val,
            review_count=review_count,
            bsr_current=bsr_now,
            bsr_trend=bsr_trend,
            listing_age_days=listing_age,
            reason=reason,
        )
        scores.append(score)

    # --- Save ChurnReport + ChurnScores ---
    keep_n = sum(1 for s in scores if s.status == "keep")
    mon_n = sum(1 for s in scores if s.status == "monitor")
    churn_n = sum(1 for s in scores if s.status == "churn")
    nd_n = sum(1 for s in scores if s.status == "no_data")

    report = ChurnReport(
        period_label=period_label,
        total_asins=len(scores),
        keep_count=keep_n,
        monitor_count=mon_n,
        churn_count=churn_n,
        no_data_count=nd_n,
    )
    db.add(report)
    await db.flush()  # Get report.id

    for s in scores:
        s.report_id = report.id
        db.add(s)

    await db.commit()

    # Sort by score ascending (worst first)
    scores.sort(key=lambda s: (s.score or 0))

    return JSONResponse({
        "report_id": report.id,
        "period_label": period_label,
        "total": len(scores),
        "keep": keep_n,
        "monitor": mon_n,
        "churn": churn_n,
        "no_data": nd_n,
        "scores": [_score_to_dict(s) for s in scores],
    })


@router.get("/reports")
async def list_reports(
    request: Request,
    _=Depends(require_auth),
    db: AsyncSession = Depends(get_db),
):
    """List all past churn reports."""
    reports = (await db.execute(
        select(ChurnReport).order_by(desc(ChurnReport.report_date))
    )).scalars().all()

    return JSONResponse([
        {
            "id": r.id,
            "period_label": r.period_label,
            "report_date": r.report_date.isoformat() if r.report_date else "",
            "total_asins": r.total_asins,
            "keep": r.keep_count,
            "monitor": r.monitor_count,
            "churn": r.churn_count,
            "no_data": r.no_data_count,
        }
        for r in reports
    ])


@router.get("/reports/{report_id}")
async def get_report(
    report_id: int,
    request: Request,
    _=Depends(require_auth),
    db: AsyncSession = Depends(get_db),
):
    """Get full scores for a specific churn report."""
    report = (await db.execute(
        select(ChurnReport).where(ChurnReport.id == report_id)
    )).scalar_one_or_none()

    if not report:
        return JSONResponse({"error": "Report not found"}, status_code=404)

    scores = (await db.execute(
        select(ChurnScore)
        .where(ChurnScore.report_id == report_id)
        .order_by(ChurnScore.score)
    )).scalars().all()

    return JSONResponse({
        "id": report.id,
        "period_label": report.period_label,
        "report_date": report.report_date.isoformat() if report.report_date else "",
        "total": report.total_asins,
        "keep": report.keep_count,
        "monitor": report.monitor_count,
        "churn": report.churn_count,
        "no_data": report.no_data_count,
        "scores": [_score_to_dict(s) for s in scores],
    })


@router.get("/reports/{report_id}/download")
async def download_report(
    report_id: int,
    request: Request,
    _=Depends(require_auth),
    db: AsyncSession = Depends(get_db),
):
    """Download churn report as Excel."""
    report = (await db.execute(
        select(ChurnReport).where(ChurnReport.id == report_id)
    )).scalar_one_or_none()

    if not report:
        return JSONResponse({"error": "Report not found"}, status_code=404)

    scores = (await db.execute(
        select(ChurnScore)
        .where(ChurnScore.report_id == report_id)
        .order_by(ChurnScore.score)
    )).scalars().all()

    rows = [
        {
            "Status": s.status.upper() if s.status else "",
            "Score": s.score,
            "ASIN": s.asin,
            "Title": s.title or "",
            "Units (30d)": s.units_ordered or 0,
            "Revenue (₹)": float(s.revenue or 0),
            "Conversion %": float(s.conversion_rate or 0),
            "Sessions": s.sessions or 0,
            "Buy Box %": float(s.buy_box_pct or 0),
            "Rating": float(s.rating) if s.rating else "",
            "Reviews": s.review_count or 0,
            "BSR": s.bsr_current or "",
            "BSR Trend": s.bsr_trend or "",
            "Listing Age (days)": s.listing_age_days or "",
            "Reason": s.reason or "",
        }
        for s in scores
    ]

    df = pd.DataFrame(rows)
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Churn Analysis")

        # Color-code the Status column
        ws = writer.sheets["Churn Analysis"]
        from openpyxl.styles import PatternFill, Font
        colors = {
            "KEEP": "C6EFCE",
            "MONITOR": "FFEB9C",
            "CHURN": "FFC7CE",
            "NO_DATA": "D9D9D9",
        }
        for row_idx, score in enumerate(scores, start=2):
            status_key = (score.status or "").upper()
            fill_color = colors.get(status_key, "FFFFFF")
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            for col_idx in range(1, len(rows[0]) + 2):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    buffer.seek(0)
    label = report.period_label or str(report.id)
    filename = f"ChurnAnalysis_{label.replace(' ', '_')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/compare")
async def compare_reports(
    report1: int,
    report2: int,
    request: Request,
    _=Depends(require_auth),
    db: AsyncSession = Depends(get_db),
):
    """Compare two reports — show ASINs that changed status."""
    scores1 = {s.asin: s.status for s in (await db.execute(
        select(ChurnScore).where(ChurnScore.report_id == report1)
    )).scalars().all()}

    scores2_raw = (await db.execute(
        select(ChurnScore).where(ChurnScore.report_id == report2)
    )).scalars().all()

    changes = []
    for s in scores2_raw:
        old_status = scores1.get(s.asin)
        if old_status and old_status != s.status:
            changes.append({
                "asin": s.asin,
                "title": s.title,
                "old_status": old_status,
                "new_status": s.status,
                "score": s.score,
                "reason": s.reason,
            })

    return JSONResponse({"changes": changes, "total_changes": len(changes)})


def _score_to_dict(s: ChurnScore) -> dict:
    return {
        "asin": s.asin,
        "title": s.title or "",
        "score": s.score or 0,
        "status": s.status or "no_data",
        "units_ordered": s.units_ordered or 0,
        "revenue": float(s.revenue or 0),
        "conversion_rate": float(s.conversion_rate or 0),
        "sessions": s.sessions or 0,
        "buy_box_pct": float(s.buy_box_pct or 0),
        "rating": float(s.rating) if s.rating else None,
        "review_count": s.review_count,
        "bsr_current": s.bsr_current,
        "bsr_trend": s.bsr_trend or "unknown",
        "listing_age_days": s.listing_age_days,
        "reason": s.reason or "",
    }
