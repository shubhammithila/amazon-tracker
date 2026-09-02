"""Projections — forecast next month's sales from live data, and reorder alerts.

**Parents come from the MRP sheet's own active-and-name grouping, never from
`app/invoice/product_families.json`.** That static file is what left Triphala Sattu invisible
here in the first place — active in the sheet, in two pack sizes, never added to the file, and
the file was the source of every row this screen showed. `app.shipment.catalogue.load_catalogue()`
is the same source and fallback chain (sheet -> cached copy -> static file) the Shipment tab
already relies on for exactly this reason.

**Sales come from `economics_snapshot`, already stored by the Portfolio tab's nightly refresh —
no new Amazon integration.** The manual Business Report CSV upload stays as an explicit override:
any edit through `/calculate` or `/upload-csv` marks that parent's row `sales_source="manual"`,
and a manual row is skipped by the weekly recompute (`app.projections.refresh.run`) until the
owner explicitly resets it through `/reset-row`.
"""
import io
import json
import logging
from pathlib import Path

import pandas as pd
from fastapi import APIRouter, Request, Depends, UploadFile, File
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.projections import logic, repository
from app.routers.auth import require_auth
from app.shipment import catalogue

router = APIRouter(prefix="/projections")
logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).parent.parent
DEFAULTS_FILE = BASE_DIR / "invoice" / "projection_defaults.json"
FAMILIES_FILE = BASE_DIR / "invoice" / "product_families.json"


def load_defaults() -> dict:
    if DEFAULTS_FILE.exists():
        with open(DEFAULTS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


DEFAULTS = load_defaults()

#: The config a parent gets when nothing in `projection_defaults.json` matches its name. Global
#: Defaults, editable on screen, unchanged from the pre-existing behaviour for an unmatched row.
GLOBAL_DEFAULTS = {
    "growth_rate": 0.3, "seasonal_impact": 1.0, "supplier_to_wh": 5, "packing": 2,
    "wh_to_ixd": 10, "ixd_to_fba": 5, "wh_buffer_days": 10.0,
}


async def build_current_rows(db: AsyncSession) -> tuple[list[dict], dict]:
    """Every currently-active parent, merged with its stored row (sales, any manual edit) or a
    freshly-built one. Returns `(rows, catalogue_report)`.

    **A brand-new parent (no stored row yet) is written to the database here**, with
    `sales_source="sheet"` and zero sales, so it exists for the weekly refresh to update and is
    never re-synthesised on every page load. A parent hidden this load (no longer active) is left
    in the database untouched — its row is not deleted, only excluded from what is returned, so a
    reactivated product keeps its history rather than starting over.
    """
    sheet_products, sheet_warning, sheet_source = await catalogue.load_catalogue()
    live_groups = logic.group_active_by_name(sheet_products)

    stored = {r["parent_product"]: r for r in await repository.load_rows(db)}
    hidden = logic.hidden_parent_names(set(stored), live_groups)

    rows: list[dict] = []
    for name, group in live_groups.items():
        if name in stored:
            rows.append(stored[name])
            continue
        config = logic.build_parent_config(name, group, DEFAULTS, GLOBAL_DEFAULTS)
        created = await repository.save_row(
            db, name,
            {**config, "last_month_sale": 0, "seven_day_rate": None, "thirty_day_rate": None,
             "daily_rate": 0, "diverged": False, "current_fba_stock": 0, "current_wh_stock": 0},
            source="sheet",
        )
        rows.append(created)

    report = {
        "source": sheet_source,
        "active_parents": len(live_groups),
        "hidden_count": len(hidden),
        "hidden_names": hidden[:8],
        "warning": sheet_warning,
    }
    return rows, report


@router.get("/last")
async def get_current(request: Request, _=Depends(require_auth), db: AsyncSession = Depends(get_db)):
    """The live table: every active parent, its purchasing config, and its sales rate.

    Replaces the old file-backed `/last`+`/init` pair. There is no "never initialised" state any
    more — the sheet always has active parents, so this always has rows to show.
    """
    rows, report = await build_current_rows(db)
    return JSONResponse({
        "products": logic.calculate_projections(rows),
        "catalogue": report,
        "blend": await repository.load_blend_settings(db),
        "last_refresh": await repository.last_refresh(db),
    })


@router.post("/calculate")
async def calculate(
    request: Request, _=Depends(require_auth), db: AsyncSession = Depends(get_db),
):
    """Persist the owner's edited values for every product in the body, recompute, and return.

    **Every product in the body is saved as `sales_source="manual"`** — the table has no way to
    tell "the owner retyped this number" from "this is what the sheet already said", so treating
    every save through this route as an edit is the safe direction. A parent whose numbers
    genuinely came from the sheet keeps reading that way until the next weekly refresh recomputes
    it; this route is only reached when the owner pressed Recalculate.
    """
    body = await request.json()
    products = body.get("products", [])

    saved = []
    for p in products:
        name = p.get("product") or p.get("parent_product")
        if not name:
            continue
        row = await repository.save_row(db, name, {
            "purchase_rate": p.get("purchase_rate", 0), "supplier_to_wh": p.get("supplier_to_wh", 5),
            "packing": p.get("packing", 2), "wh_to_ixd": p.get("wh_to_ixd", 10),
            "ixd_to_fba": p.get("ixd_to_fba", 5), "wh_buffer_days": p.get("wh_buffer_days", 10),
            "seasonal_impact": p.get("seasonal_impact", 1.0), "growth_rate": p.get("growth_rate", 0.3),
            "last_month_sale": p.get("last_month_sale", 0),
            "current_fba_stock": p.get("current_fba_stock", 0),
            "current_wh_stock": p.get("current_wh_stock", 0),
        }, source="manual")
        saved.append(row)

    products = logic.calculate_projections(saved)
    total_forecast = sum(p["monthly_forecast"] for p in products)
    total_ideal_value = sum(p["ideal_stock_value"] for p in products)
    total_current_value = sum(p["current_stock_value"] for p in products)
    ship_alerts = sum(1 for p in products if p["shipment_alert"] > 0)
    reorder_alerts = sum(1 for p in products if p["reorder_alert"] > 0)
    critical_alerts = sum(1 for p in products if p.get("inventory_days", 999) < 7)

    return JSONResponse({
        "products": products,
        "summary": {
            "total_products": len(products),
            "total_forecast_kg": round(total_forecast, 0),
            "total_ideal_value": round(total_ideal_value, 0),
            "total_current_value": round(total_current_value, 0),
            "shipment_alerts": ship_alerts,
            "reorder_alerts": reorder_alerts,
            "critical_alerts": critical_alerts,
        },
    })


@router.post("/reset-row")
async def reset_row(request: Request, _=Depends(require_auth), db: AsyncSession = Depends(get_db)):
    """Clear one parent's manual override, body: `{"parent_product": "..."}`. The next weekly
    refresh (or a manual one) will then update it again."""
    body = await request.json()
    name = (body.get("parent_product") or "").strip()
    if not name:
        return JSONResponse({"error": "parent_product is required."}, status_code=400)
    result = await repository.reset_to_sheet(db, name)
    if result is None:
        return JSONResponse({"error": f"No row found for {name!r}."}, status_code=404)
    return JSONResponse({"row": result})


@router.get("/blend-settings")
async def get_blend_settings(_=Depends(require_auth), db: AsyncSession = Depends(get_db)):
    """The blend weight and divergence threshold, with their bounds — for the settings panel."""
    return {
        "blend": await repository.load_blend_settings(db),
        "defaults": dict(logic.DEFAULT_BLEND),
        "ranges": {k: list(v) for k, v in logic.BLEND_RANGES.items()},
        "help": {
            "seven_day_weight": "How much the last 7 days counts against the last 30 when "
                                 "forecasting. Higher reacts faster to a real spike or drop; "
                                 "lower is steadier against a noisy week.",
            "divergence_pct": "When the 7-day and 30-day rates disagree by more than this, the "
                               "row is flagged so you can see why its forecast moved.",
        },
    }


@router.post("/blend-settings")
async def set_blend_settings(
    request: Request, _=Depends(require_auth), db: AsyncSession = Depends(get_db),
):
    """Edit the blend weight/threshold, or reset to the measured defaults."""
    body = await request.json()
    if body.get("reset"):
        return {"blend": await repository.reset_blend_settings(db), "status": "reset"}
    try:
        saved = await repository.save_blend_settings(db, body.get("blend") or {})
    except ValueError as exc:
        return JSONResponse({"error": str(exc)}, status_code=400)
    return {"blend": saved, "status": "saved"}


@router.post("/refresh-now")
async def refresh_now(_=Depends(require_auth), db: AsyncSession = Depends(get_db)):
    """Run the weekly 7d/30d recompute immediately — "I want it now", the same button every
    other refreshable tab in this app offers."""
    from app.projections import refresh

    result = await refresh.run(db)
    return JSONResponse(result)


@router.delete("/clear")
async def clear_all_overrides(_=Depends(require_auth), db: AsyncSession = Depends(get_db)):
    """Reset EVERY parent's manual override at once. The pre-existing "Clear" button's meaning
    changes with this feature: there is no longer a file to delete, and the equivalent action —
    discarding every hand-typed correction so the next refresh recomputes everything from the
    sheet — is this.
    """
    rows = await repository.load_rows(db)
    for row in rows:
        if row["sales_source"] == "manual":
            await repository.reset_to_sheet(db, row["parent_product"])
    return JSONResponse({"status": "cleared", "reset_count": sum(
        1 for r in rows if r["sales_source"] == "manual"
    )})


# ─── CSV upload: an explicit manual override, not a live source ────────────────

def _clean_number(val) -> float:
    if val is None or str(val).strip() in ("", "-", "nan"):
        return 0.0
    import re
    cleaned = re.sub(r"[₹,%\s]", "", str(val))
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def parse_business_report_for_projections(content: bytes, groups: dict[str, dict]) -> dict[str, float]:
    """Business Report CSV -> `{parent_name: total_kg_sold}`, using the LIVE sheet's own
    ASIN->parent grouping (`groups`, from `logic.group_active_by_name`) — never
    `product_families.json`. An ASIN outside every active group (discontinued, or unknown to the
    sheet) is skipped, the same rule `sales_kg_by_parent` follows for economics rows.
    """
    try:
        df = pd.read_csv(io.BytesIO(content))
    except Exception:
        df = pd.read_csv(io.BytesIO(content), encoding="utf-8-sig")

    child_col = "(Child) ASIN"
    if child_col not in df.columns:
        raise ValueError("Could not find '(Child) ASIN' column. Upload Amazon Business Report (By ASIN).")

    asin_to_parent: dict[str, tuple[str, float]] = {}
    for parent, group in groups.items():
        for asin in group["asins"]:
            asin_to_parent[asin] = (parent, group["weights"].get(asin) or 0)

    product_kg: dict[str, float] = {}
    for _, row in df.iterrows():
        asin = str(row.get(child_col, "")).strip()
        if not asin or len(asin) < 10:
            continue
        mapping = asin_to_parent.get(asin.upper())
        if not mapping:
            continue
        parent, weight = mapping
        units = _clean_number(row.get("Units Ordered", 0))
        product_kg[parent] = product_kg.get(parent, 0) + units * weight

    return product_kg


@router.post("/upload-csv")
async def upload_csv(
    request: Request, file: UploadFile = File(...), _=Depends(require_auth),
    db: AsyncSession = Depends(get_db),
):
    """Upload a Business Report CSV -> fill `last_month_sale` (kg) per active parent, marked
    as a manual override — see the module docstring."""
    content = await file.read()

    sheet_products, _warning, _source = await catalogue.load_catalogue()
    groups = logic.group_active_by_name(sheet_products)

    try:
        product_kg = parse_business_report_for_projections(content, groups)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=400)

    rows, _report = await build_current_rows(db)
    saved = []
    for row in rows:
        name = row["parent_product"]
        kg = round(product_kg.get(name, 0), 2)
        updated = await repository.save_row(db, name, {**row, "last_month_sale": kg}, source="manual")
        saved.append(updated)

    products = logic.calculate_projections(saved)
    products.sort(key=lambda x: x.get("monthly_forecast", 0), reverse=True)
    filled = sum(1 for p in products if p["last_month_sale"] > 0)

    return JSONResponse({
        "products": products,
        "total_products": len(products),
        "filled_from_csv": filled,
        "total_kg_from_csv": round(sum(product_kg.values()), 1),
        "total_forecast_kg": round(sum(p["monthly_forecast"] for p in products), 0),
    })


@router.get("/defaults")
async def get_defaults(request: Request, _=Depends(require_auth)):
    """Every entry in `projection_defaults.json`, unchanged from before this feature — kept for
    reference; not read by the current template."""
    return JSONResponse(DEFAULTS)


@router.get("/download")
async def download_projection(request: Request, _=Depends(require_auth), db: AsyncSession = Depends(get_db)):
    """Download the current table as Excel."""
    rows, _report = await build_current_rows(db)
    products = logic.calculate_projections(rows)
    if not products:
        return JSONResponse({"error": "No projection data"}, status_code=404)

    out_rows = [{
        "Product": p["parent_product"],
        "Brand": p.get("brand", ""),
        "Last Month Sale (kg)": p.get("last_month_sale", 0),
        "Seasonal Impact": p.get("seasonal_impact", 1),
        "Growth Rate": p.get("growth_rate", 0.3),
        "Monthly Forecast (kg)": p.get("monthly_forecast", 0),
        "Daily Rate (kg)": p.get("daily_rate", 0),
        "Lead Time (days)": p.get("total_lead_time", 0),
        "Ideal FBA Stock (kg)": p.get("ideal_fba_stock", 0),
        "Current FBA Stock (kg)": p.get("current_fba_stock", 0),
        "Shipment Alert (kg)": p.get("shipment_alert", 0),
        "WH Buffer Days": p.get("wh_buffer_days", 0),
        "Ideal WH Stock (kg)": p.get("ideal_wh_stock", 0),
        "Current WH Stock (kg)": p.get("current_wh_stock", 0),
        "Reorder Alert (kg)": p.get("reorder_alert", 0),
        "Purchase Rate (Rs/kg)": p.get("purchase_rate", 0),
        "Ideal Stock Value (Rs)": p.get("ideal_stock_value", 0),
        "Inventory Days": p.get("inventory_days", 0),
        "Sales Source": p.get("sales_source", "sheet"),
        "Needs Review": "yes" if p.get("needs_review") else "",
    } for p in products]

    df = pd.DataFrame(out_rows)
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Projections")
        ws = writer.sheets["Projections"]
        from openpyxl.styles import PatternFill
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        for row_idx, p in enumerate(products, start=2):
            if p.get("shipment_alert", 0) > 0:
                ws.cell(row=row_idx, column=11).fill = red_fill
            if p.get("reorder_alert", 0) > 0:
                ws.cell(row=row_idx, column=15).fill = red_fill
            if p.get("inventory_days", 999) < 7:
                ws.cell(row=row_idx, column=18).fill = red_fill
            elif p.get("inventory_days", 999) < 14:
                ws.cell(row=row_idx, column=18).fill = yellow_fill

    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Projections.xlsx"},
    )
