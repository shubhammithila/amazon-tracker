"""The shipment downloads. Each returns an in-memory io.BytesIO.

Four working documents, three of which share one column layout:

  plan       xlsx + pdf   what to pack this week, for the owner
  packed     xlsx + pdf   what was boxed, over a date or a range
  remaining  xlsx + pdf   the morning clipboard sheet, for the packer
  shipment file  xlsx     the merchant-SKU + quantity upload for Amazon

The first three go through ``build_simple_xlsx`` / ``build_simple_pdf``, differing
only in the quantity column.

> Four earlier builders lived here — ``build_packing_plan_xlsx``,
> ``build_packing_plan_pdf``, ``build_packed_xlsx`` and ``build_remaining_pdf`` —
> each with its own column list. The shared layout replaced them on every route,
> and then their own tests kept them alive for a whole step. That is worse than
> plain dead code: twelve tests went on asserting column content for sheets nobody
> could download, so this file read as though it described what the owner gets.
> Deleted, and the tests with them.

**Every builder takes rows that are already sorted and renders them in the order
given.** None of them sorts. Row order comes from
``repository.load_plan_items()`` and nowhere else, which is what stops a download
from disagreeing with the screen the owner was just looking at. There is a test
that reads a generated xlsx back and asserts its row order equals
``logic.sort_items``; if you add a ``sorted()`` call in here, that is what fails.

Nothing here touches the database or FastAPI — the router loads and orders the
rows, these functions only format. That keeps them directly unit-testable and
means an ordering bug can only have one home.
"""
import io
import logging
from datetime import date

from app.shipment import logic

logger = logging.getLogger(__name__)

#: Matches the invoice PDFs (app/invoice/generator.py) so the shipment downloads
#: do not look like they came from a different product.
HEADER_RGB = (0.2, 0.3, 0.3)

#: openpyxl wants "RRGGBB"; reportlab wants floats. Derived from HEADER_RGB so
#: the Excel and PDF documents cannot drift apart in appearance.
HEADER_HEX = "".join(f"{int(round(c * 255)):02X}" for c in HEADER_RGB)


def _weight_label(weight) -> str:
    """Pack size as the warehouse says it — see logic.weight_label.

    Kept as a thin alias because this module calls it in several places and the
    name reads better in a row-building expression. The RULE lives in logic.py with
    the other shared rules: grams under 1 kg, kilos above, and one implementation
    so the printed sheet cannot disagree with the screen.
    """
    return logic.weight_label(weight)


# ─── Excel ───────────────────────────────────────────────────────────────────

def _autosize(worksheet, widths: list[int]) -> None:
    from openpyxl.utils import get_column_letter

    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width


def _write_sheet(worksheet, headers: list[str], rows: list[list], widths: list[int]) -> None:
    """Header row + data rows, with the header frozen and styled.

    Freezing matters in practice: the packed sheet grows a column per packing day,
    so by the end of a week the owner is scrolling and would otherwise lose track
    of which column is which date.
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    worksheet.append(headers)
    fill = PatternFill("solid", fgColor=HEADER_HEX)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        worksheet.append(row)

    worksheet.freeze_panes = "A2"
    _autosize(worksheet, widths)


def build_shipment_file_xlsx(
    items: list[dict],
    mode: str = "remaining",
    days: list[dict] | None = None,
    fc_code: str = "",
) -> io.BytesIO:
    """The Amazon upload sheet: merchant SKU + quantity, and the destination FC.

    ``mode`` picks the quantity column:

      remaining  what is still to be packed (planning a future shipment)
      all        the full planned quantity
      verified   only units on days the owner has verified

    ``fc_code`` is the destination the owner chose (ISK3, DED3, BLR4 …), written onto
    **every** row rather than stated once at the top. The sheet is read by a machine
    and gets sorted and filtered by hand, so a value that lives on the row cannot be
    detached from it — and this file decides where real boxes are sent. The column is
    omitted entirely when no FC is given, rather than appearing blank, because an
    empty destination column invites someone to fill it in later.

    Rows whose quantity is 0 are omitted — Amazon has nothing to do with them and
    they only make the file harder to check.

    A row with no merchant SKU is written with a blank SKU rather than silently
    falling back to the ASIN. Amazon's upload keys on the merchant SKU, so an
    ASIN there is rejected on their side; a visible blank is a problem the owner
    can fix before uploading, whereas a plausible-looking ASIN is one they find
    out about from Amazon.
    """
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Shipment"

    verified = logic.verified_units_by_asin(days or [])

    rows = []
    missing_sku = 0
    for item in items:
        asin = item.get("asin", "")
        planned = int(item.get("shipment_plan") or 0)
        if mode == "all":
            quantity = planned
        elif mode == "verified":
            quantity = int(verified.get(asin, 0))
        else:
            quantity = int(item.get("remaining") or 0)

        if quantity <= 0:
            continue
        sku = item.get("fba_sku") or ""
        if not sku:
            missing_sku += 1
        row = [sku, asin, item.get("item", ""), _weight_label(item.get("weight")), quantity]
        if fc_code:
            row.append(fc_code)
        rows.append(row)

    if missing_sku:
        # Surfaced rather than swallowed: this used to fail invisibly.
        logger.warning(
            "shipment file: %d row(s) have no merchant SKU and will be rejected by Amazon",
            missing_sku,
        )

    headers = ["Merchant SKU", "ASIN", "Product", "Weight", "Quantity"]
    widths = [22, 14, 30, 9, 11]
    if fc_code:
        headers.append("Ship to FC")
        widths.append(12)

    _write_sheet(sheet, headers, rows, widths)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


# ─── The three working documents ─────────────────────────────────────────────
#
# One column layout, requested explicitly:
#
#     S · M · B · Brand · ASIN · SKU · Product · Size · <quantity>
#
# The quantity column is the only thing that differs, because each sheet answers a
# different question: the plan and the remaining sheet say what is still TO PACK,
# while the packed sheet says what WAS packed.
#
# All three drop rows with nothing to do. That is the point of them: a 205-row
# sheet where 88 rows read 0 is a sheet nobody checks, and the owner asked for
# "only the final rows".
#
# Both formats are generated from the SAME rows list by the same two functions
# below, so the Excel and the PDF of one document cannot disagree.
#
# **Visual hierarchy is a correctness feature on these, not decoration.** They are
# read on a clipboard, at arm's length, by someone holding a carton. Three cells
# decide what he does — product, pack size, quantity — and the other five only
# matter when something has gone wrong and a line has to be identified. So those
# three are set large and bold and the identifiers are set small and grey. Printing
# all eight at one weight is what produced "which of these numbers am I packing".

#: The shared header, minus the quantity column.
#:
#: ``Size`` is its own column rather than being glued onto the product name.
#: They were combined to save width on a portrait page, and it was wrong: the eye
#: cannot scan a column of sizes that are buried at the end of names of different
#: lengths, which is exactly the scan the packer makes when he has all sizes of one
#: product in front of him. Separated, the sizes line up.
#:
#: The first seven are unchanged and in the order the owner asked for; a test pins
#: that prefix, so ``Size`` is appended rather than inserted.
IDENTITY_HEADERS = ["S", "M", "B", "Brand", "ASIN", "Merchant SKU", "Product", "Size"]
IDENTITY_WIDTHS = [4, 4, 4, 7, 15, 24, 32, 11]

#: Column headings whose cells are identifiers: needed to resolve a query, never
#: read while packing. Rendered small and grey in both formats.
QUIET_HEADERS = frozenset({"ASIN", "Merchant SKU"})

#: The headings that carry the actual instruction. Rendered large and bold.
LOUD_HEADERS = frozenset({"Product", "Size"})


def _identity_cells(item: dict) -> list:
    """The eight columns every one of the three documents starts with.

    S/M/B are rendered as a tick or blank rather than TRUE/FALSE: they are carton
    sizes being read off a page, and "S Y" scans in a way "S TRUE" does not.
    """
    return [
        "Y" if item.get("s") else "",
        "Y" if item.get("m") else "",
        "Y" if item.get("b") else "",
        item.get("brand", ""),
        item.get("asin", ""),
        item.get("fba_sku", ""),
        item.get("item", ""),
        _weight_label(item.get("weight")),
    ]


def _rows_with_quantity(items: list[dict], quantity_key: str) -> list[list]:
    """Identity columns + one quantity, keeping only rows with something to do."""
    rows = []
    for item in items:
        quantity = int(item.get(quantity_key) or 0)
        if quantity <= 0:
            continue
        rows.append(_identity_cells(item) + [quantity])
    return rows


#: Where the "TOTAL · N rows" label is written on the totals line. The Product
#: column, because it is the widest and the one the eye is already running down.
_TOTALS_LABEL_COLUMN = IDENTITY_HEADERS.index("Product")


def _totals_row(headers: list[str], rows: list[list]) -> list:
    """A totals line under the data, summing every quantity column present.

    Driven by the header count rather than a fixed position, so the packed sheet's
    two quantity columns (units and cartons) total correctly without a second
    implementation. The label position is derived from IDENTITY_HEADERS too — it
    was a hand-counted run of blanks, which silently shifted when Size was
    added and put the label under a quantity heading.
    """
    totals: list = [""] * len(IDENTITY_HEADERS)
    totals[_TOTALS_LABEL_COLUMN] = f"TOTAL · {len(rows)} rows"
    for offset in range(len(headers) - len(IDENTITY_HEADERS)):
        column = len(IDENTITY_HEADERS) + offset
        totals.append(sum(int(row[column] or 0) for row in rows))
    return totals


def build_simple_xlsx(
    title: str, subtitle: str, headers: list[str], rows: list[list], widths: list[int]
) -> io.BytesIO:
    """A one-sheet workbook in the shared layout, with a totals row.

    The same emphasis as the PDF: quantity and product bold, ASIN and SKU small and
    grey. Deliberately mirrored rather than left plain, because the owner reads the
    Excel of the same document the packer holds on paper, and two versions that look
    unrelated get treated as two different reports.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31]  # Excel's limit; a longer name makes the file invalid.

    body = list(rows) + ([_totals_row(headers, rows)] if rows else [])
    _write_sheet(sheet, headers, body, widths)

    quiet = Font(size=9, color="FF6B7280")
    loud = Font(size=12, bold=True)
    quantity_font = Font(size=12, bold=True)
    for index, heading in enumerate(headers, start=1):
        for row in sheet.iter_rows(min_row=2, min_col=index, max_col=index):
            cell = row[0]
            if heading in QUIET_HEADERS:
                cell.font = quiet
            elif heading in LOUD_HEADERS:
                cell.font = loud
            elif index > len(IDENTITY_HEADERS):
                cell.font = quantity_font
                cell.alignment = Alignment(horizontal="right")

    # The totals line, bold across the whole width so it cannot be misread as data.
    if rows:
        for cell in sheet[sheet.max_row]:
            cell.font = Font(size=11, bold=True)

    sheet.row_dimensions[1].height = 22
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_portfolio_xlsx(
    title: str, subtitle: str, headers: list[str], rows: list[list], widths: list[int]
) -> io.BytesIO:
    """The portfolio review as a workbook: parents with their sizes indented beneath.

    **A sibling of ``build_simple_xlsx`` rather than a parameter on it, and the reason is
    ``_totals_row``.** That helper sums every column past the identity ones with
    ``int(row[column] or 0)``, which is correct for the picking documents it was written for —
    every trailing column there is a quantity. This sheet's trailing columns include percentages
    rendered as text ("+43.6%"), an em dash where there is no denominator, a star rating and a
    prose reason, so that sum raises ``ValueError: invalid literal for int()``. Caught by a test
    rather than in production, but only because the export had one.

    Widening ``_totals_row`` to skip non-numeric cells would silently change the four documents
    that depend on its current behaviour, so this builds its own sheet instead. No totals row at
    all: a "TOTAL" line under a mixture of parent rows and size rows would double-count, because
    each parent already contains its sizes. The KPI figures live in the subtitle.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31]  # Excel's limit; a longer name makes the file invalid.

    _write_sheet(sheet, headers, list(rows), widths)

    # **The subtitle is written into the sheet, unlike `build_simple_xlsx` where it is accepted
    # and silently discarded.** It carries the date window and the PRE-COGS caveat, and a
    # workbook leaves the app to be read without the screen's banner beside it — so a file
    # showing "+8.8% net" with no caveat gets forwarded to someone who reads it as profit.
    # Inserted above the header row so it is the first thing read.
    sheet.insert_rows(1)
    sheet["A1"] = subtitle
    sheet["A1"].font = Font(size=9, italic=True, color="FF6B7280")
    sheet["A1"].alignment = Alignment(vertical="center")

    quiet = Font(size=9, color="FF6B7280")
    for index, heading in enumerate(headers, start=1):
        for row in sheet.iter_rows(min_row=2, min_col=index, max_col=index):
            cell = row[0]
            if heading in ("ASIN", "Why"):
                cell.font = quiet
            elif heading in ("Sales", "Ad spend", "Net", "Units", "Net %", "TACOS"):
                cell.alignment = Alignment(horizontal="right")
    # The reason column is prose and needs to wrap rather than run under its neighbours.
    if "Why" in headers:
        column = headers.index("Why") + 1
        for row in sheet.iter_rows(min_row=2, min_col=column, max_col=column):
            row[0].alignment = Alignment(vertical="top", wrap_text=True)

    sheet.row_dimensions[1].height = 22
    sheet.freeze_panes = "A3"       # subtitle + headers stay visible while scrolling 90 products
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_simple_pdf(
    title: str, subtitle: str, headers: list[str], rows: list[list]
) -> io.BytesIO:
    """The same rows as a portrait A4 PDF — a clipboard page, not a wide report.

    **Every cell is a Paragraph, not a string, and that is the fix for a real bug
    rather than a styling choice.** reportlab does not wrap a plain string in a
    table cell: it draws it at full width and lets it run straight over the
    gridline into the next column. Merchant SKUs here are up to 24 characters
    ("Beetroot_Sattu_500g FBA"), so the SKU column was printing on top of the
    product column on most rows of a real 117-row plan. A Paragraph wraps, and the
    row grows to fit instead.

    The typographic hierarchy is set here rather than in the table style because it
    is per-column and reportlab's TableStyle can only set a font over a rectangle —
    which is fine, but the wrapping already requires per-cell objects, so the
    emphasis rides along for free and stays in one place.
    """
    from reportlab.lib.units import mm
    from reportlab.platypus import Table

    buffer = io.BytesIO()
    doc, elements = _pdf_document(buffer, title, subtitle, landscape_mode=False)

    body = list(rows) + ([_totals_row(headers, rows)] if rows else [])
    if not body:
        body = [["—"] * len(headers)]
    totals_index = len(body) - 1 if rows else None

    data = [[_head_cell(h) for h in headers]]
    for position, row in enumerate(body):
        emphasise = position == totals_index
        data.append([
            _body_cell(row[column], headers[column], column, len(headers), emphasise)
            for column in range(len(headers))
        ])

    table = Table(data, colWidths=_pdf_column_widths(headers, body), repeatRows=1)
    table.setStyle(
        _pdf_table_style(
            len(headers),
            # +1 for the header row, which is data[0].
            totals_row=totals_index + 1 if totals_index is not None else None,
        )
    )
    elements.append(table)
    doc.build(elements, onFirstPage=_page_footer, onLaterPages=_page_footer)
    buffer.seek(0)
    return buffer


#: A4 portrait (210 mm) minus the 10 mm margins each side.
_PAGE_WIDTH_MM = 190

#: LEFTPADDING + RIGHTPADDING from the table style, in mm. A column must be this
#: much wider than its text or the text wraps — the trap that wrapped every ASIN
#: when the widths were first measured.
_PADDING_MM = 8 / 72 * 25.4  # 4pt + 4pt ≈ 2.8 mm

#: Slack on top of the measured content width. An EXACT fit wraps: reportlab breaks
#: when the text is >= the available width, and float arithmetic makes "equal" a
#: coin toss. Observed as "4560" wrapping to "456" / "0" in a totals row whose
#: column had been sized to precisely 8.6303 mm for 8.6303 mm of digits.
_SLACK_MM = 0.8


#: Columns that must never wrap, with the ceiling each may claim (mm). An ASIN or a
#: pack size broken across two lines is unreadable rather than merely ugly — "B0GW3
#: 88QP6" cannot be typed into a search box, and "1.75" over "kg" reads as two
#: facts. Every one of these has bounded content, so a ceiling is safe.
_NO_WRAP_CEILING_MM = {"S": 8, "M": 8, "B": 8, "Brand": 14, "ASIN": 24, "Size": 20}

#: Columns allowed to wrap when their content is genuinely long, and the width past
#: which they will. The Merchant SKU's longest real value is 68 mm — giving it that
#: would take a third of the page for a string nobody reads while packing.
_WRAP_CEILING_MM = {"Merchant SKU": 42}

#: Quantity columns. Bounded by construction (an integer), so a ceiling is safe.
_QUANTITY_CEILING_MM = 22

#: Product never gets less than this, however many quantity columns are added.
_PRODUCT_FLOOR_MM = 34


def _pdf_column_widths(headers: list[str], rows: list[list]) -> list:
    """Column widths in points, summing to exactly the printable width.

    **Measured from the rows being rendered, not from constants.** Two earlier passes
    hardcoded millimetres — the first by eye, which gave Product 62 mm when three
    quarters of the names need 32, and the second from percentiles of the catalogue,
    which still wrapped every ASIN because it forgot that padding eats 2.8 mm of a
    column. Both were the same mistake: a number written down here has to be right
    about data it cannot see, and it silently degrades when the data changes.

    So each column asks for what its own longest value actually needs at its own font
    size, capped so one freak value cannot take the page, and **Product absorbs the
    remainder**. Product is the flexible one because its content has no natural bound
    and because it is the least harmful to wrap: a long name over two lines is still
    completely readable, where a split ASIN is not.

    The consequence worth knowing: the packed sheet's extra quantity column narrows
    Product rather than squeezing the digits or clipping the SKU.
    """
    from reportlab.lib.units import mm
    from reportlab.pdfbase.pdfmetrics import stringWidth

    def content_mm(column: int, heading: str) -> float:
        """The widest cell in this column, plus its header, at the right font."""
        style = _body_cell_style_name(heading, column)
        font, size = _STYLE_FONTS[style]
        widest = max(
            (stringWidth(str(row[column] or ""), font, size) for row in rows),
            default=0.0,
        )
        # The header is bold 8.5pt and can easily be the widest thing in a narrow
        # column ("Cartons" beats any 3-digit carton count).
        widest = max(widest, stringWidth(heading, "Helvetica-Bold", 8.5))
        return widest / mm + _PADDING_MM + _SLACK_MM

    widths_mm: list[float | None] = []
    for column, heading in enumerate(headers):
        if heading == "Product":
            widths_mm.append(None)
            continue
        needed = content_mm(column, heading)
        if heading in _NO_WRAP_CEILING_MM:
            widths_mm.append(min(needed, _NO_WRAP_CEILING_MM[heading]))
        elif heading in _WRAP_CEILING_MM:
            widths_mm.append(min(needed, _WRAP_CEILING_MM[heading]))
        else:
            widths_mm.append(min(needed, _QUANTITY_CEILING_MM))

    spare = _PAGE_WIDTH_MM - sum(w for w in widths_mm if w is not None)
    return [
        (max(_PRODUCT_FLOOR_MM, spare) if w is None else w) * mm for w in widths_mm
    ]


def _paragraph_styles():
    """Cached paragraph styles for the table cells.

    Built once per process rather than per cell: a 205-row plan with nine columns is
    1,845 cells, and constructing a ParagraphStyle for each one is measurable.
    """
    global _STYLES
    if _STYLES is not None:
        return _STYLES

    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.lib.styles import ParagraphStyle

    base = ParagraphStyle(
        "Cell", fontName="Helvetica", fontSize=9, leading=11, textColor=colors.black
    )
    _STYLES = {
        # Identifiers: present for lookups, deliberately recessive. 7.5pt grey is
        # still legible on paper at reading distance — this is de-emphasis, not
        # fine print.
        "quiet": ParagraphStyle(
            "Quiet", parent=base, fontSize=7.5, leading=9,
            textColor=colors.Color(0.42, 0.45, 0.50),
        ),
        # What the packer is actually reading.
        "loud": ParagraphStyle(
            "Loud", parent=base, fontName="Helvetica-Bold", fontSize=10.5, leading=12.5
        ),
        "quantity": ParagraphStyle(
            "Qty", parent=base, fontName="Helvetica-Bold", fontSize=11, leading=13,
            alignment=TA_RIGHT,
        ),
        "flag": ParagraphStyle("Flag", parent=base, alignment=TA_CENTER),
        "plain": base,
        "totals": ParagraphStyle(
            "Totals", parent=base, fontName="Helvetica-Bold", fontSize=10, leading=12
        ),
        "totals_qty": ParagraphStyle(
            "TotalsQty", parent=base, fontName="Helvetica-Bold", fontSize=11,
            leading=13, alignment=TA_RIGHT,
        ),
        "head": ParagraphStyle(
            "Head", parent=base, fontName="Helvetica-Bold", fontSize=8.5, leading=10,
            textColor=colors.white,
        ),
    }
    return _STYLES


_STYLES = None


def _escape(value) -> str:
    """Text for a Paragraph. Product names come from an uploaded CSV.

    Paragraph parses its input as mini-HTML, so an ampersand in a product name
    ("Salt & Pepper") raises a parse error and takes the whole download down with a
    500. Found the same way the template escaping was.
    """
    return (
        str("" if value is None else value)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def _head_cell(heading: str):
    from reportlab.platypus import Paragraph

    return Paragraph(_escape(heading), _paragraph_styles()["head"])


#: Font and size per style name, so _pdf_column_widths can measure a cell without
#: instantiating its Paragraph. Kept beside _body_cell_style_name because the two
#: must agree: a style added to one and not the other measures at the wrong size and
#: the column silently wraps.
_STYLE_FONTS = {
    "quiet": ("Helvetica", 7.5),
    "loud": ("Helvetica-Bold", 10.5),
    "quantity": ("Helvetica-Bold", 11),
    "flag": ("Helvetica", 9),
    "plain": ("Helvetica", 9),
    "totals": ("Helvetica-Bold", 10),
    "totals_qty": ("Helvetica-Bold", 11),
}


def _body_cell_style_name(heading: str, column: int, totals: bool = False) -> str:
    """Which style a cell gets. The single rule, read by both renderer and measurer.

    Order matters: the quantity test comes before the heading tests so a quantity
    column named "Units" is right-aligned and bold regardless of its name.
    """
    is_quantity = column >= len(IDENTITY_HEADERS)
    if totals:
        return "totals_qty" if is_quantity else "totals"
    if is_quantity:
        return "quantity"
    if heading in QUIET_HEADERS:
        return "quiet"
    if heading in LOUD_HEADERS:
        return "loud"
    if heading in ("S", "M", "B"):
        return "flag"
    return "plain"


def _body_cell(value, heading: str, column: int, column_count: int, totals: bool):
    """One table cell as a wrapping, styled Paragraph."""
    from reportlab.platypus import Paragraph

    style = _paragraph_styles()[_body_cell_style_name(heading, column, totals)]
    return Paragraph(_escape(value), style)


# ─── PDF ─────────────────────────────────────────────────────────────────────

def _pdf_table_style(column_count: int, totals_row: int | None = None):
    """The shared table style. Fonts and colours per cell live in the Paragraphs.

    What is set here is only what a TableStyle can express better than a cell can:
    the header band, the zebra striping, the padding, and the rules.

    The grid is deliberately gone. A full box around every cell of a 117-row table
    is a lot of ink competing with the numbers; horizontal rules alone are enough to
    keep your place on a line, which is the actual job — and the striping already
    does most of that work. Vertical separation comes from the padding.
    """
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle

    rule = colors.Color(0.85, 0.87, 0.90)
    commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.Color(*HEADER_RGB)),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        # Room to breathe. The old 2pt default with a 9pt font made the rows read as
        # one grey mass at arm's length.
        ("TOPPADDING", (0, 0), (-1, -1), 4.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4.5),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        # Zebra striping: this is read on a clipboard in a warehouse, where losing
        # your place mid-row is the realistic failure.
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.Color(0.97, 0.975, 0.98)]),
        ("LINEBELOW", (0, 1), (-1, -1), 0.3, rule),
        ("LINEBELOW", (0, 0), (-1, 0), 0.8, colors.Color(*HEADER_RGB)),
    ]
    if totals_row is not None:
        # A rule above the totals, so it reads as a summary and not as one more SKU.
        commands += [
            ("LINEABOVE", (0, totals_row), (-1, totals_row), 0.9, colors.Color(0.3, 0.34, 0.4)),
            ("BACKGROUND", (0, totals_row), (-1, totals_row), colors.Color(0.93, 0.94, 0.96)),
        ]
    return TableStyle(commands)


def _pdf_document(buffer, title: str, subtitle: str, landscape_mode: bool = False):
    """A document plus its heading block.

    The heading is left-aligned rather than centred, and the page number is in the
    footer. Both are for the same reason: these get printed, stapled and put on a
    clipboard, and a centred title with no page number on sheet 3 of 4 is a document
    nobody can tell is incomplete.
    """
    from reportlab.lib import colors, pagesizes
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

    page = pagesizes.landscape(pagesizes.A4) if landscape_mode else pagesizes.A4
    doc = SimpleDocTemplate(
        buffer, pagesize=page,
        leftMargin=10 * mm, rightMargin=10 * mm, topMargin=10 * mm, bottomMargin=14 * mm,
        title=title,
    )
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(
            _escape(title),
            ParagraphStyle(
                "DocTitle", parent=styles["Heading1"], fontSize=15, leading=18,
                spaceAfter=1, textColor=colors.Color(*HEADER_RGB),
            ),
        ),
        Paragraph(
            _escape(subtitle),
            ParagraphStyle(
                "DocSub", parent=styles["Normal"], fontSize=8.5,
                textColor=colors.Color(0.42, 0.45, 0.50),
            ),
        ),
        Spacer(1, 4 * mm),
    ]
    return doc, elements


def _page_footer(canvas, doc):
    """"Page 1 of 3" plus when it was printed, bottom-right on every page.

    Registered as onPage rather than appended as flowables, because the total page
    count is only known once the document has been laid out — and a footer that says
    "page 1" with no total is exactly as useless as no footer when four sheets get
    separated on a warehouse bench.
    """
    from reportlab.lib import colors
    from reportlab.lib.units import mm

    canvas.saveState()
    canvas.setFont("Helvetica", 7.5)
    canvas.setFillColor(colors.Color(0.5, 0.53, 0.58))
    canvas.drawRightString(
        doc.pagesize[0] - 10 * mm, 7 * mm,
        f"Page {canvas.getPageNumber()} · printed {date.today().isoformat()}",
    )
    canvas.restoreState()


# ─── The dispatch sheet: today's parcels, for the warehouse floor ─────────────

#: Millimetres of indent on a size row, so the nesting reads without a tree glyph.
_NEST_INDENT_MM = 6


def _purchase_table(purchasing: dict, styles: dict):
    """The weight-and-purchase table: ordered against raw stock, and what to buy.

    A separate builder because it is the only section that can appear alone (`tab=weight`) and
    the only one whose numbers are kilograms throughout.
    """
    from reportlab.platypus import Paragraph, Table

    rows = []
    for row in purchasing.get("rows") or []:
        rows.append([
            Paragraph(_escape(row["product"]), styles["loud"]),
            Paragraph(_escape(row["brand"]), styles["quiet"]),
            Paragraph(f"{float(row['ordered_kg']):.2f}", styles["plain"]),
            Paragraph(f"{float(row['raw_kg']):.2f}", styles["plain"]),
            # An em dash, not 0.00: a zero in a purchasing column reads as a measurement
            # rather than as "nothing to do".
            Paragraph(
                "—" if row["covered"] else f"{float(row['to_buy_kg']):.2f}",
                styles["quantity"],
            ),
        ])
    if not rows:
        rows = [[Paragraph("Nothing due today.", styles["plain"])]
                + [Paragraph("", styles["plain"])] * 4]

    totals = purchasing.get("totals") or {}
    rows.append([
        Paragraph("TOTAL", styles["totals"]),
        Paragraph("", styles["totals"]),
        Paragraph(f"{float(totals.get('ordered_kg') or 0):.2f}", styles["totals_qty"]),
        Paragraph(f"{float(totals.get('raw_kg') or 0):.2f}", styles["totals_qty"]),
        # Sum of the CLAMPED rows, never total_ordered - total_raw.
        Paragraph(f"{float(totals.get('to_buy_kg') or 0):.2f}", styles["totals_qty"]),
    ])

    table = Table(
        [[_head_cell(head) for head in
          ["Product", "Brand", "Ordered kg", "Raw kg", "To buy kg"]]] + rows,
        colWidths=_dispatch_widths([70, 22, 30, 30, 32]),
        repeatRows=1,
    )
    table.setStyle(_dispatch_table_style(totals_row=len(rows)))
    return table


def build_dispatch_pdf(
    sheet: dict, subtitle: str, tab: str = "all", purchasing: dict | None = None
) -> io.BytesIO:
    """Today's dispatch as up to three tables on one portrait page.

    Asked for as *"Each parent item total weight orders … uske niche 500g, 1kg - kitne kitne
    units. sort it total weight wise"*, plus a list of *"all orders with order id, name,
    product and qty"*.

    **Sections in one document, not separate downloads.** They are read together: purchasing
    says what to bring to the bench, the summary says how much of each product, the order list
    is what gets checked off against the parcels. Separate files get separated on a warehouse
    bench — the same reason `_page_footer` prints "page 1 of 3".

    `tab` narrows the document to one section, and `purchasing` supplies the weight-and-purchase
    table that tab 1 shows. Both are defaulted so the existing two-argument calls keep working.

    **A PDF has no dropdown, so nesting is indentation plus weight.** A parent row carries the
    product name and its total kilograms in bold; its pack sizes sit indented beneath in the
    recessive style, heaviest first. That is the paper equivalent of a collapsible row, and it
    keeps one product's sizes contiguous so a packer visits one shelf once.

    **A separate function rather than a parameter on `build_simple_pdf`.** That builds one flat
    table with one totals row and cannot express a parent/child relationship; bending it would
    make both callers harder to read than having two.

    Every cell is a `Paragraph`. reportlab draws a bare string at full width straight over the
    next column's gridline — the bug that printed merchant SKUs on top of product names on a
    real 117-row plan, with the whole suite green because the bytes were still a valid PDF of
    a plausible size.
    """
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, Spacer, Table

    buffer = io.BytesIO()
    doc, elements = _pdf_document(buffer, "Dispatch sheet", subtitle, landscape_mode=False)
    styles = _paragraph_styles()

    # ── Section 0: weight and purchasing, when asked for ──
    if tab in ("all", "weight") and purchasing is not None:
        elements.append(Paragraph("Weight &amp; purchase", styles["loud"]))
        elements.append(Spacer(1, 2 * mm))
        elements.append(_purchase_table(purchasing, styles))
        if tab == "all":
            elements.append(Spacer(1, 7 * mm))

    if tab not in ("all", "sku", "orders", "weight"):
        # Unreachable through the routes, which validate first — but a direct caller passing a
        # typo would otherwise get a document with no tables, and reportlab raises on that.
        raise ValueError(f"unknown dispatch tab {tab!r}")

    # ── Section 1: parent products, heaviest first, sizes nested ──
    summary_headers = ["Product", "Size", "Units", "Orders", "Net kg", "Packed"]
    summary_rows: list[list] = []
    #: Table row indices (header included) that are PARENT rows, so the style can rule above
    #: each group. Collected while building, since only this loop knows which is which.
    parent_row_indices: list[int] = []

    for parent in sheet.get("parents") or []:
        parent_row_indices.append(len(summary_rows) + 1)      # +1 for the header row
        summary_rows.append([
            Paragraph(_escape(parent["product"]), styles["loud"]),
            Paragraph(_escape(parent.get("brand") or ""), styles["quiet"]),
            Paragraph(str(int(parent["units"])), styles["quantity"]),
            Paragraph(str(int(parent["orders"])), styles["quantity"]),
            Paragraph(f"{float(parent['kg'] or 0):.2f}", styles["quantity"]),
            Paragraph(str(int(parent.get("packed") or 0)), styles["quantity"]),
        ])
        for size in parent.get("sizes") or []:
            # An unweighed size shows blank, never "0.00": a zero reads as a measurement,
            # and this number is handed to a courier.
            kg = "" if size["kg"] is None else f"{float(size['kg']):.2f}"
            indent = f'<para leftIndent="{_NEST_INDENT_MM * mm:.1f}">'
            label = _escape(size.get("seller_sku") or size.get("asin") or "")
            summary_rows.append([
                Paragraph(f"{indent}{label}</para>", styles["quiet"]),
                Paragraph(_escape(size["weight_label"] or "—"), styles["plain"]),
                Paragraph(str(int(size["units"])), styles["plain"]),
                Paragraph(str(int(size["orders"])), styles["plain"]),
                Paragraph(kg, styles["plain"]),
                Paragraph(str(int(size.get("packed") or 0)), styles["plain"]),
            ])

    blank = len(summary_headers) - 1
    if not summary_rows:
        summary_rows = [
            [Paragraph("Nothing to dispatch today.", styles["plain"])]
            + [Paragraph("", styles["plain"])] * blank
        ]

    totals = sheet.get("totals") or {}
    summary_rows.append([
        Paragraph("TOTAL", styles["totals"]),
        Paragraph("", styles["totals"]),
        Paragraph(str(int(totals.get("units") or 0)), styles["totals_qty"]),
        Paragraph(str(int(totals.get("orders") or 0)), styles["totals_qty"]),
        Paragraph(f"{float(totals.get('kg') or 0):.2f}", styles["totals_qty"]),
        Paragraph(str(int(totals.get("packed") or 0)), styles["totals_qty"]),
    ])

    if tab in ("all", "sku"):
        summary = Table(
            [[_head_cell(head) for head in summary_headers]] + summary_rows,
            colWidths=_dispatch_widths([72, 24, 20, 20, 26, 22]),
            repeatRows=1,
        )
        summary.setStyle(
            _dispatch_table_style(
                totals_row=len(summary_rows),          # header offset already counted
                parent_rows=parent_row_indices,
            )
        )
        if tab == "all":
            elements.append(Paragraph("By product", styles["loud"]))
            elements.append(Spacer(1, 2 * mm))
        elements.append(summary)

    if tab not in ("all", "orders"):
        doc.build(elements, onFirstPage=_page_footer, onLaterPages=_page_footer)
        buffer.seek(0)
        return buffer

    # ── Section 2: every order, in the same parent order as the summary ──
    if tab == "all":
        elements.append(Spacer(1, 7 * mm))
    elements.append(Paragraph("Orders", styles["loud"]))
    elements.append(Spacer(1, 2 * mm))

    order_headers = ["Order", "Name", "Product (SKU)", "Qty", "Destination"]
    order_rows = []
    for row in sheet.get("orders") or []:
        product = " · ".join(
            part for part in (row.get("weight_label"), row.get("seller_sku")) if part
        )
        destination = ", ".join(
            part for part in (row.get("city"), row.get("state")) if part
        )
        order_rows.append([
            Paragraph(_escape(row["amazon_order_id"]), styles["quiet"]),
            Paragraph(_escape(row["parent"]), styles["plain"]),
            Paragraph(_escape(product), styles["quiet"]),
            Paragraph(str(int(row["quantity"])), styles["quantity"]),
            Paragraph(_escape(destination), styles["plain"]),
        ])
    if not order_rows:
        order_rows = [
            [Paragraph("No orders.", styles["plain"])]
            + [Paragraph("", styles["plain"])] * (len(order_headers) - 1)
        ]

    order_table = Table(
        [[_head_cell(head) for head in order_headers]] + order_rows,
        colWidths=_dispatch_widths([34, 44, 52, 14, 40]),
        repeatRows=1,
    )
    order_table.setStyle(_dispatch_table_style())
    elements.append(order_table)

    doc.build(elements, onFirstPage=_page_footer, onLaterPages=_page_footer)
    buffer.seek(0)
    return buffer


def _dispatch_widths(widths: list[float]) -> list:
    """A millimetre layout scaled to exactly the printable width.

    Fixed proportions rather than `_pdf_column_widths`, which measures the widest cell per
    column. That is right for the plan documents, where a merchant SKU can genuinely need
    68 mm — but here the first column holds BOTH a product name and an indented SKU, so a
    measured width would size it for the longer of two different kinds of content and starve
    the number columns. Those are bounded by construction (integers, two-decimal kilograms),
    so fixing them is safe.

    Scaled rather than hardcoded to 190 mm, so one column can be re-tuned by eye without
    every other number having to be recomputed to stop the row overflowing the page.
    """
    from reportlab.lib.units import mm

    total = sum(widths)
    if total <= 0:
        return [w * mm for w in widths]
    return [w * _PAGE_WIDTH_MM / total * mm for w in widths]


def _dispatch_table_style(
    totals_row: int | None = None, parent_rows: list[int] | None = None
):
    """`_pdf_table_style`'s look, plus a rule above each parent, minus the zebra striping.

    Striping fights the nesting: alternating bands make a parent and its first size read as
    two unrelated rows, when the entire point of the layout is that they belong together. A
    hairline above each parent group keeps your place on the page the way striping does,
    while also showing the structure.
    """
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle

    rule = colors.Color(0.85, 0.87, 0.90)
    commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.Color(*HEADER_RGB)),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("LINEBELOW", (0, 1), (-1, -1), 0.25, rule),
        ("LINEBELOW", (0, 0), (-1, 0), 0.8, colors.Color(*HEADER_RGB)),
    ]
    for index in parent_rows or []:
        commands.append(
            ("LINEABOVE", (0, index), (-1, index), 0.7, colors.Color(0.55, 0.6, 0.66))
        )
    if totals_row is not None:
        commands += [
            ("LINEABOVE", (0, totals_row), (-1, totals_row), 0.9,
             colors.Color(0.3, 0.34, 0.4)),
            ("BACKGROUND", (0, totals_row), (-1, totals_row),
             colors.Color(0.93, 0.94, 0.96)),
        ]
    return TableStyle(commands)


# ─── The dispatch workbook: one worksheet per screen tab ──────────────────────

#: Column widths for the three dispatch worksheets, in Excel character units.
_XLSX_PURCHASE_WIDTHS = [34, 8, 14, 14, 14]
_XLSX_SKU_WIDTHS = [30, 10, 22, 12, 14, 10]
_XLSX_ORDER_WIDTHS = [24, 22, 8, 30, 12, 28, 16]

#: Which worksheet each `tab=` value keeps. `all` keeps every one.
_XLSX_TAB_SHEETS = {
    "weight": "Weight & purchase",
    "sku": "By SKU",
    "orders": "Orders",
}

#: The purchasing table's headers, shared by the workbook and the to-buy list.
_PURCHASE_HEADERS = ["Product", "Brand", "Ordered kg", "Raw stock kg", "To buy kg"]


def _purchase_rows(purchasing: dict) -> list[list]:
    """The purchasing rows, shared by the workbook and the to-buy list.

    One function so the two files cannot disagree about a weight — the same reason all five
    downloads come from one aggregation.
    """
    return [
        [
            row["product"],
            row["brand"],
            float(row["ordered_kg"]),
            float(row["raw_kg"]),
            # A covered product shows blank rather than 0.00: a zero reads as a measurement,
            # and this column is a purchasing instruction.
            "" if row["covered"] else float(row["to_buy_kg"]),
        ]
        for row in purchasing.get("rows") or []
    ]


def build_dispatch_xlsx(
    sheet: dict, purchasing: dict, subtitle: str, tab: str = "all"
) -> io.BytesIO:
    """Today's dispatch as one workbook with a worksheet per screen tab.

    **One file, not three.** The tabs are read together — purchasing says what to bring to the
    bench, the SKU sheet is what gets counted, the order sheet is what gets checked off. Three
    separate files get separated on a warehouse bench, the same reason `_page_footer` prints
    "page 1 of 3".

    `tab` drops the worksheets that were not asked for, so a single-tab download is this same
    builder rather than a second code path that could disagree about a quantity.

    Worksheet names match the tab labels exactly, so nobody has to work out which is which.
    """
    from openpyxl import Workbook

    book = Workbook()

    # ── Sheet 1: weight and purchasing ──
    purchase = book.active
    purchase.title = "Weight & purchase"
    ptotals = purchasing.get("totals") or {}
    purchase_rows = _purchase_rows(purchasing) + [[
        "TOTAL", "",
        float(ptotals.get("ordered_kg") or 0),
        float(ptotals.get("raw_kg") or 0),
        # Sum of the CLAMPED rows, never total_ordered - total_raw: a surplus of one product
        # cannot cover a shortfall of another.
        float(ptotals.get("to_buy_kg") or 0),
    ]]
    _write_sheet(purchase, _PURCHASE_HEADERS, purchase_rows, _XLSX_PURCHASE_WIDTHS)

    # ── Sheet 2: per SKU, flat, in the sheet's existing order ──
    by_sku = book.create_sheet("By SKU")
    sku_rows = [
        [
            parent["product"],
            size.get("weight_label") or "",
            size.get("seller_sku") or size.get("asin") or "",
            int(size["units"]),
            int(size.get("packed") or 0),
            int(size.get("remaining") or 0),
        ]
        for parent in sheet.get("parents") or []
        for size in parent.get("sizes") or []
    ]
    stotals = sheet.get("totals") or {}
    sku_rows.append([
        "TOTAL", "", "",
        int(stotals.get("units") or 0),
        int(stotals.get("packed") or 0),
        int(stotals.get("remaining") or 0),
    ])
    _write_sheet(
        by_sku, ["Product", "Size", "SKU", "Ordered", "Packed today", "Left"],
        sku_rows, _XLSX_SKU_WIDTHS,
    )

    # ── Sheet 3: every order line ──
    orders_sheet = book.create_sheet("Orders")
    order_rows = [
        [
            row["amazon_order_id"],
            row.get("seller_sku") or row.get("asin") or "",
            int(row["quantity"]),
            row["parent"],
            row.get("weight_label") or "",
            ", ".join(part for part in (row.get("city"), row.get("state")) if part),
            row.get("easyship_status") or "",
        ]
        for row in sheet.get("orders") or []
    ]
    _write_sheet(
        orders_sheet,
        ["Order", "SKU", "Qty", "Item", "Weight", "Destination", "Amazon status"],
        order_rows or [["No orders", "", 0, "", "", "", ""]],
        _XLSX_ORDER_WIDTHS,
    )

    wanted = _XLSX_TAB_SHEETS.get(tab)
    if wanted is not None:
        for name in list(book.sheetnames):
            if name != wanted:
                del book[name]

    # Provenance goes in the workbook's metadata rather than a spare row, so it cannot be
    # sorted away from the data it describes.
    book.properties.title = "Dispatch sheet"
    book.properties.description = subtitle

    buffer = io.BytesIO()
    book.save(buffer)
    buffer.seek(0)
    return buffer


def build_tobuy_xlsx(purchasing: dict, subtitle: str) -> io.BytesIO:
    """The purchasing shortfall only — the products where `to_buy > 0`.

    **Filtered, not sorted.** A covered product is ABSENT rather than shown with a zero: this
    file gets pasted into a supplier email, and a row reading "ABC Sattu … 0" invites someone to
    order zero of it.

    With nothing short it says so in words rather than rendering an empty table, because an
    empty grid reads as a failed download on the one day the news is good.
    """
    from openpyxl import Workbook

    short = [row for row in (purchasing.get("rows") or []) if not row["covered"]]

    book = Workbook()
    worksheet = book.active
    worksheet.title = "To buy"

    if not short:
        rows = [["Nothing to buy — every product is covered by raw stock.", "", "", "", ""]]
    else:
        rows = [
            [row["product"], row["brand"], float(row["ordered_kg"]),
             float(row["raw_kg"]), float(row["to_buy_kg"])]
            for row in short
        ]
        rows.append([
            "TOTAL", "",
            round(sum(row["ordered_kg"] for row in short), 2),
            round(sum(row["raw_kg"] for row in short), 2),
            round(sum(row["to_buy_kg"] for row in short), 2),
        ])

    _write_sheet(worksheet, _PURCHASE_HEADERS, rows, _XLSX_PURCHASE_WIDTHS)
    book.properties.title = "To buy"
    book.properties.description = subtitle

    buffer = io.BytesIO()
    book.save(buffer)
    buffer.seek(0)
    return buffer
