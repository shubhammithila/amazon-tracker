"""Building the invoice DOCUMENTS: an incomplete line must not 500 the download.

Distinct from tests/test_invoice_save.py, which guards the GST *number* series.
Nothing here allocates a number — ``/invoice/generate-excel`` and
``/invoice/generate-pdf`` only render bytes — and one test asserts exactly that,
because it is the property that makes these routes safe to retry.

Found by QA. ``generator.py`` read ``item["sku"]``, ``item["title"]`` and
``item["hsn_code"]`` with hard brackets while every other field on the same row used
``.get()``. A line missing any of the three raised KeyError, which surfaced as
``500 Internal Server Error``.

**Why that mattered more than it looks.** The owner is at the last step of raising a
real invoice when he clicks Download. A 500 is indistinguishable from the app being
broken, tells him nothing about which line is wrong, and the natural response is to
retry — so the fix is a visibly blank cell he can spot on the document before sending
it. HSN is the exception: it falls back to the module default, because a GST invoice
with an empty HSN column is rejected outright.

The route is reachable with partial lines in practice. ``/invoice/parse-shipment``
builds items from an Amazon TSV whose columns vary, the shipment bridge composes its
own line dicts, and the invoice screen lets rows be hand-edited before download.
"""
import io

import pytest

pytestmark = pytest.mark.regression

XLSX_MAGIC = b"PK\x03\x04"
PDF_MAGIC = b"%PDF"

ROUTES = ["/invoice/generate-excel", "/invoice/generate-pdf"]


def _payload(items):
    """A complete invoice apart from the item lines under test."""
    return {
        "details": {
            "shipment_id": "FBA15TEST001",
            "date": "2026-08-10",
            "fc_code": "ISK3",
            "place_of_supply": "Maharashtra",
            "transporter": "VRL Logistics",
            "boxes": "40",
            "weight": "120",
        },
        "supplier": {"gstin": "20AAFCF9848M1Z7"},
        "recipient": {"gstin": "27AAFCF9848M1ZT"},
        "items": items,
    }


COMPLETE_ITEM = {
    "sku": "MF-CH-1KG",
    "title": "Mithila Foods Chana Sattu 1 kg",
    "short_title": "Chana Sattu 1 kg",
    "asin": "B0AAA00001",
    "fnsku": "X001ABCDEF",
    "quantity": 100,
    "hsn_code": "1106",
    "gst_rate": 5,
    "rate": 95.0,
    "unit": "Pcs",
}


def _magic(route: str) -> bytes:
    return XLSX_MAGIC if route.endswith("excel") else PDF_MAGIC


# ─── The happy path ──────────────────────────────────────────────────────────

@pytest.mark.parametrize("route", ROUTES)
async def test_a_complete_invoice_renders(auth_client, route):
    r = await auth_client.post(route, json=_payload([COMPLETE_ITEM]))
    assert r.status_code == 200, r.text[:300]
    assert r.content.startswith(_magic(route))
    assert len(r.content) > 1000, "suspiciously small document"


# ─── Incomplete lines: a blank cell, never a 500 ─────────────────────────────

#: Each case drops one field the generator used to require. The names are the fields
#: that were hard lookups; `description only` is the shape a caller writes when they
#: assume this route takes the same keys as /invoice/save.
INCOMPLETE_ITEMS = {
    "no sku": {k: v for k, v in COMPLETE_ITEM.items() if k != "sku"},
    "no title": {k: v for k, v in COMPLETE_ITEM.items() if k != "title"},
    "no hsn_code": {k: v for k, v in COMPLETE_ITEM.items() if k != "hsn_code"},
    "no short_title": {k: v for k, v in COMPLETE_ITEM.items() if k != "short_title"},
    "blank sku": {**COMPLETE_ITEM, "sku": ""},
    "blank title": {**COMPLETE_ITEM, "title": ""},
    "null title": {**COMPLETE_ITEM, "title": None},
    "quantity and rate only": {"quantity": 10, "rate": 9.0},
    "description only": {"description": "Chana Sattu 1kg", "quantity": 10, "rate": 9.0},
    "empty item": {},
}


@pytest.mark.parametrize("label", sorted(INCOMPLETE_ITEMS))
@pytest.mark.parametrize("route", ROUTES)
async def test_an_incomplete_line_still_produces_a_document(auth_client, route, label):
    """20 combinations, and every one used to be a coin toss on KeyError."""
    r = await auth_client.post(route, json=_payload([INCOMPLETE_ITEMS[label]]))
    assert r.status_code == 200, (
        f"{label} on {route} -> {r.status_code}: {r.text[:200]}. A missing field must "
        "leave a blank cell on the document, not fail the download — the owner cannot "
        "tell a 500 apart from the app being down."
    )
    assert r.content.startswith(_magic(route))


@pytest.mark.parametrize("route", ROUTES)
async def test_a_missing_hsn_falls_back_rather_than_blanking(auth_client, route):
    """The one field that must NOT be blank.

    A GST invoice with an empty HSN column is rejected outright, so an absent code
    falls back to the module default (1106 at 5%, which is every F2D food product)
    rather than leaving a hole the owner might not notice.
    """
    from app.invoice.hsn_codes import DEFAULT_HSN

    item = {k: v for k, v in COMPLETE_ITEM.items() if k != "hsn_code"}
    r = await auth_client.post(route, json=_payload([item]))
    assert r.status_code == 200, r.text[:200]

    if route.endswith("excel"):
        from openpyxl import load_workbook

        sheet = load_workbook(io.BytesIO(r.content)).active
        cells = {
            str(c.value)
            for row in sheet.iter_rows()
            for c in row
            if c.value is not None
        }
        assert DEFAULT_HSN in cells, (
            f"no HSN on the document; a GST invoice without one is rejected: "
            f"{sorted(cells)[:25]}"
        )


async def test_a_blank_title_leaves_the_cell_empty_not_the_word_none(auth_client):
    """`str(None)` renders the literal text "None" onto a tax document.

    Worse than blank, because it looks like a product name and would be sent.
    """
    from openpyxl import load_workbook

    item = {**COMPLETE_ITEM, "title": None, "short_title": None}
    r = await auth_client.post("/invoice/generate-excel", json=_payload([item]))
    assert r.status_code == 200, r.text[:200]

    sheet = load_workbook(io.BytesIO(r.content)).active
    values = [
        str(c.value)
        for row in sheet.iter_rows()
        for c in row
        if c.value is not None
    ]
    assert "None" not in values, f'the literal string "None" is on the invoice: {values[:25]}'


# ─── Many lines, and the totals ──────────────────────────────────────────────

@pytest.mark.parametrize("route", ROUTES)
async def test_several_lines_render(auth_client, route):
    """Multi-line is the normal case — a shipment is many SKUs."""
    items = [
        {**COMPLETE_ITEM, "sku": f"MF-{n}", "title": f"Product {n}", "quantity": n * 10}
        for n in range(1, 9)
    ]
    r = await auth_client.post(route, json=_payload(items))
    assert r.status_code == 200, r.text[:200]
    assert r.content.startswith(_magic(route))


async def test_the_excel_totals_are_arithmetically_right(auth_client):
    """100 x 95 + 60 x 58 = 12,980 taxable; 5% IGST = 649; total 13,629.

    Hand-checked, and asserted on the rendered document rather than on a helper — the
    numbers on the sheet are what gets filed.
    """
    from openpyxl import load_workbook

    items = [
        {**COMPLETE_ITEM, "quantity": 100, "rate": 95.0},
        {**COMPLETE_ITEM, "sku": "MF-JAU-500G", "title": "Jau Sattu 500 g",
         "quantity": 60, "rate": 58.0},
    ]
    r = await auth_client.post("/invoice/generate-excel", json=_payload(items))
    assert r.status_code == 200, r.text[:200]

    sheet = load_workbook(io.BytesIO(r.content)).active
    numbers = {
        round(float(c.value), 2)
        for row in sheet.iter_rows()
        for c in row
        if isinstance(c.value, (int, float))
    }
    assert 12980.0 in numbers, f"taxable total missing: {sorted(numbers)}"
    assert 649.0 in numbers, f"IGST missing: {sorted(numbers)}"
    assert 13629.0 in numbers, f"grand total missing: {sorted(numbers)}"


# ─── These routes must not touch the GST series ──────────────────────────────

async def test_generating_a_document_allocates_no_invoice_number(
    auth_client, count_rows
):
    """The property that makes these routes safe to retry.

    ``POST /invoice/save`` is the only writer of the legally-sequential series. If
    rendering a document ever consumed a number, a failed download would burn one —
    and a gap in the sequence is a question answered during an audit.
    """
    from app.models import Invoice

    before = await count_rows(Invoice)
    for route in ROUTES:
        for _ in range(2):
            r = await auth_client.post(route, json=_payload([COMPLETE_ITEM]))
            assert r.status_code == 200, r.text[:200]
    assert await count_rows(Invoice) == before, (
        "generating a document created an invoice row, which burns a GST number"
    )


# ─── Auth and hostile bodies ─────────────────────────────────────────────────

@pytest.mark.parametrize("route", ROUTES)
async def test_generating_requires_auth(client, route):
    r = await client.post(route, json=_payload([COMPLETE_ITEM]))
    assert r.status_code in (303, 401, 403), r.status_code
    assert not r.content.startswith(XLSX_MAGIC)
    assert not r.content.startswith(PDF_MAGIC)


@pytest.mark.parametrize("body", ["not json", "", "[1,2,3]", '"a string"', "42", "null"])
@pytest.mark.parametrize("route", ROUTES)
async def test_hostile_bodies_do_not_500(auth_client, route, body):
    """Same class of bug as the shipment router's, so the same standard applies."""
    r = await auth_client.post(
        route, content=body, headers={"Content-Type": "application/json"}
    )
    assert r.status_code < 500, f"body {body!r} on {route} -> {r.status_code}"


@pytest.mark.parametrize("route", ROUTES)
async def test_no_items_does_not_500(auth_client, route):
    """An empty shipment is a mistake, not a crash. It may legitimately render an
    empty document — what it must not do is fail opaquely."""
    r = await auth_client.post(route, json=_payload([]))
    assert r.status_code < 500, f"{route} -> {r.status_code}: {r.text[:200]}"
