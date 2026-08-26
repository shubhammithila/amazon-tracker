"""The Orders routes: the picking sheet, the order list, the export, and access.

Every route here reads LOCAL rows. `getOrders` allows one request every 22 seconds, so a
route that called Amazon would hang the page and 429 the moment two people opened the tab —
which is why `POST /orders/refresh` starts a background job and returns immediately.
"""
from datetime import datetime, timedelta

import pytest

from app.orders import logic, refresh, repository

pytestmark = pytest.mark.regression

#: An ASIN this file's own catalogue stub knows.
KNOWN_ASIN = "B0CWGXYLT6"


@pytest.fixture(autouse=True)
def stub_catalogue(monkeypatch):
    """A two-product MRP catalogue for these tests.

    The suite's autouse `no_live_product_sheet` returns an EMPTY catalogue so nothing hits
    the real Google Sheet — right for the shipment tests, but it means no ASIN resolves to
    a product here, and a picking sheet with no product names proves nothing. So the
    products these tests need are supplied explicitly rather than depending on whatever
    the live sheet holds today.
    """
    async def _catalogue():
        return (
            {
                KNOWN_ASIN: {"asin": KNOWN_ASIN, "name": "Chana Sattu", "weight": 0.5,
                             "brand": "Mithila Foods", "active": True},
                "B0RAGI1KG0": {"asin": "B0RAGI1KG0", "name": "Ragi Atta", "weight": 1.0,
                               "brand": "Mithila Foods", "active": True},
            },
            None,
            "sheet",
        )

    monkeypatch.setattr("app.shipment.catalogue.load_catalogue", _catalogue, raising=True)
    return _catalogue


def _ship_by_for(ist_date):
    """The UTC instant Amazon would send for "end of `ist_date` in India".

    23:59 IST is 18:29Z on the SAME calendar day, because the offset is +5:30 and 23:59
    minus 5:30 is 18:29 — still the same date. Written as a helper so no test builds it
    from the local clock by accident.
    """
    return datetime(ist_date.year, ist_date.month, ist_date.day, 18, 29)


def _row(order_id, **overrides):
    row = {
        "amazon_order_id": order_id,
        "purchase_date_utc": datetime.utcnow() - timedelta(hours=6),
        # 18:29Z on the UTC day whose IST evening is TODAY. Derived from the IST date,
        # not from utcnow(): between 18:30Z and midnight UTC the IST date is already
        # tomorrow, so building this from the UTC clock puts the deadline a day out — the
        # exact confusion this feature exists to prevent, and it failed here first.
        "latest_ship_date_utc": _ship_by_for(datetime.now(logic.IST).date()),
        "status": "Unshipped",
        "easyship_status": "PendingSchedule",
        "ship_service_level": "Std IN EZ National COD",
        "order_total": 319.0,
        "currency": "INR",
        "items_ordered": 1,
        "items_shipped": 0,
        "is_prime": False,
        "is_cod": True,
        "city": "NAVSARI",
        "state": "GUJARAT",
        "postal_code": "396445",
    }
    row.update(overrides)
    return row


async def _seed(db, orders=None, items=None):
    await repository.upsert_orders(db, orders or [_row("403-1")])
    for order_id, rows in (items or {"403-1": [
        {"asin": KNOWN_ASIN, "seller_sku": "cs-500", "title": "Chana Sattu",
         "quantity_ordered": 2}
    ]}).items():
        await repository.replace_items(db, order_id, rows)


async def test_the_orders_route_returns_the_sheet_and_the_list(auth_client, db):
    """One payload feeds both views, so they cannot disagree about a quantity."""
    await _seed(db)
    body = (await auth_client.get("/orders")).json()

    assert body["total_orders"] == 1
    assert "sheet" in body and "sections" in body["sheet"]
    # All three sections are present even when empty: a section that vanished would read
    # as a bug rather than an empty queue.
    for bucket in (logic.BUCKET_TODAY, logic.BUCKET_PICKUP, logic.BUCKET_LATER):
        assert bucket in body["sheet"]["sections"], f"{bucket} missing from the sheet"
    assert body["section_labels"][logic.BUCKET_TODAY]
    assert body["last_refreshed_at"], "the banner has no timestamp to show"


async def test_the_sheet_aggregates_the_seeded_order(auth_client, db):
    """The end-to-end shape the warehouse reads: product, size, units, orders, kg."""
    await _seed(db)
    body = (await auth_client.get("/orders")).json()
    today = body["sheet"]["sections"][logic.BUCKET_TODAY]

    assert today["totals"]["orders"] == 1
    assert today["totals"]["quantity"] == 2
    line = today["lines"][0]
    assert line["quantity"] == 2
    assert line["weight_label"], "the pack size did not resolve from the catalogue"
    assert line["kg"] and line["kg"] > 0, "no net weight on a known product"


async def test_a_ship_by_deadline_is_rendered_in_ist_not_utc(auth_client, db):
    """18:29Z must reach the client as 23:59 IST.

    The single most consequential rendering decision on this screen: UTC would show every
    deadline 5.5 hours early on the page whose only job is "what must go out today".
    """
    await _seed(db)
    body = (await auth_client.get("/orders")).json()
    order = body["orders"][0]
    assert order["ship_by_ist"], "no ship-by date reached the client"
    # ship_by_ist is an IST calendar DATE; the same instant in UTC could be the day before.
    assert order["ship_by_ist"] == datetime.now(logic.IST).date().isoformat()


async def test_the_1995_sentinel_is_not_sent_as_a_date(auth_client, db):
    """Amazon's "no deadline" value must not reach the client as a real date.

    Rendered, it reads as 31 years overdue and sorts to the top of the packer's list.
    """
    await _seed(db, orders=[_row("403-odd",
                                 latest_ship_date_utc=datetime(1995, 1, 1),
                                 ship_service_level="Std IN EZ National COD")],
                items={"403-odd": [{"asin": KNOWN_ASIN, "quantity_ordered": 1}]})
    body = (await auth_client.get("/orders")).json()
    order = next(o for o in body["orders"] if o["amazon_order_id"] == "403-odd")
    assert order["ship_by_ist"] is None, "the 1995 sentinel was sent as a date"
    assert order["overdue"] is False, "the sentinel was treated as overdue"


async def test_the_picking_sheet_downloads_as_excel(auth_client, db):
    """The printed sheet comes from the same aggregation as the screen."""
    await _seed(db)
    r = await auth_client.get("/orders/download/picking-sheet.xlsx?bucket=to_pack")
    assert r.status_code == 200, r.text
    assert "spreadsheet" in r.headers["content-type"]
    assert len(r.content) > 1000, "the workbook is suspiciously small"


async def test_an_unknown_section_is_refused(auth_client, db):
    """A typo'd bucket must not silently export the wrong section."""
    await _seed(db)
    r = await auth_client.get("/orders/download/picking-sheet.xlsx?bucket=nonsense")
    assert r.status_code == 400
    assert "nonsense" in r.json()["error"]


async def test_a_second_refresh_is_refused_over_http(auth_client, db, monkeypatch):
    """409 rather than a second job: two refreshes would 429 each other.

    The state is set directly here because this asserts the ROUTE's guard; the job's own
    guard is covered in tests/test_orders_refresh.py.
    """
    refresh.reset_state()
    refresh.STATE["running"] = True
    try:
        r = await auth_client.post("/orders/refresh")
        assert r.status_code == 409
        assert "already running" in r.json()["error"].lower()
    finally:
        refresh.reset_state()


async def test_refresh_status_is_readable_without_starting_one(auth_client):
    """The banner polls this every few seconds, so it must be cheap and side-effect free."""
    refresh.reset_state()
    body = (await auth_client.get("/orders/refresh-status")).json()
    assert body["running"] is False
    assert body["phase"] == "idle"


async def test_the_orders_area_is_required(ops_client):
    """Deny by default. An ops user without the `orders` area sees none of it.

    The area exists so the WAREHOUSE can be granted the picking sheet deliberately —
    the default must still be no access, or granting means nothing.
    """
    for method, path in (
        ("get", "/orders"),
        ("get", "/orders/refresh-status"),
        ("post", "/orders/refresh"),
        ("get", "/orders/download/picking-sheet.xlsx"),
        ("get", "/orders-page"),
    ):
        r = await getattr(ops_client, method)(path)
        assert r.status_code in (302, 303, 401, 403), f"{path} -> {r.status_code}"


def test_the_orders_area_is_grantable_and_denied_by_default():
    """It must be a real area, or the Users screen cannot offer it."""
    from app import permissions

    assert permissions.ORDERS in permissions.AREA_KEYS
    assert any(key == permissions.ORDERS for key, _l, _h in permissions.AREAS), (
        "the area is not in AREAS, so it cannot be ticked on the Users screen"
    )
    # Deny by default: an empty grant must not include it.
    assert not permissions.has("", permissions.ORDERS)
    assert permissions.has(permissions.ORDERS, permissions.ORDERS)


# ─── The screen ──────────────────────────────────────────────────────────────

def _orders_source() -> str:
    from pathlib import Path

    return (
        Path(__file__).resolve().parent.parent / "templates" / "orders.html"
    ).read_text(encoding="utf-8")


def test_no_calendar_date_is_rendered_through_a_time_formatter():
    """A date put through a time formatter lands on the WRONG DAY.

    An IST calendar date ("2026-08-25") is a DAY, and `new Date("2026-08-25")` is UTC
    midnight by specification — so a time formatter printed "26 Aug, 05:30", five and a half
    hours into the following day, in the column the warehouse plans against. Found in a
    browser, not by a test: the server was correct and the formatter was correct, and only
    their combination was wrong.

    Asserted as a PROHIBITION on the pairing rather than on one call site. The original
    version pinned `esc(dateIST(o.ship_by_ist))`, which broke the moment the screen became a
    dispatch sheet — every order there is due today, so the column went away and the test
    failed while the bug it guards was still absent. A rule about which formatter may touch a
    date survives the screen being redesigned; a rule about one expression does not.
    """
    import re

    source = _orders_source()
    # Any *_ist / *_date field fed to the time formatter is the bug.
    offenders = re.findall(r"timeIST\(\s*[A-Za-z_.]*(?:pack_date|ship_by|_date_ist)\b", source)
    assert not offenders, (
        f"a calendar date is being formatted as a time, which shifts it a day: {offenders}"
    )
    # And the date helper must still exist to be used at all.
    assert "function dateIST(" in source, "dateIST was removed"
    assert "dateIST(" in source.replace("function dateIST(", ""), (
        "dateIST is defined but never called, so a date is being rendered some other way"
    )


def test_date_ist_does_not_construct_a_date_object():
    """Splitting the string is the fix; parsing it is the bug.

    `new Date("2026-08-25")` is UTC midnight by specification, so ANY timezone conversion
    applied afterwards moves the day. dateIST must therefore not build a Date at all.
    """
    source = _orders_source()
    start = source.index("function dateIST(")
    body = source[start:source.index("\n}", start)]
    assert "new Date(" not in body, (
        "dateIST constructs a Date, which reintroduces the UTC-midnight shift"
    )
    assert ".split(" in body, "dateIST should split the ISO date rather than parse it"


async def test_the_refresh_status_is_json_serialisable_with_real_timestamps(
    auth_client, db
):
    """The production 500: `TypeError: Object of type datetime is not JSON serializable`.

    `STATE` holds real `datetime` objects while a refresh runs, and `JSONResponse` cannot
    serialise one. Every route that returns the status was affected, but it surfaced on the
    409 "already running" path — the one request the owner makes precisely when he wants to
    know what is happening.

    The earlier 409 test set `running` by hand while `started_at` was still None, so the
    offending value was never present. This one sets a REAL timestamp, which is what
    production had.
    """
    import json

    refresh.reset_state()
    refresh.STATE.update({
        "running": True,
        "started_at": datetime.utcnow(),
        "finished_at": datetime.utcnow(),
        "phase": "orders",
        "orders_seen": 113,
    })
    try:
        # The status endpoint itself.
        r = await auth_client.get("/orders/refresh-status")
        assert r.status_code == 200, r.text
        assert isinstance(r.json()["started_at"], str), "started_at is not a string"

        # And the 409 path, which is where it actually broke.
        conflict = await auth_client.post("/orders/refresh")
        assert conflict.status_code == 409, conflict.text
        json.dumps(conflict.json())          # would raise if a datetime leaked through
        assert isinstance(conflict.json()["refresh"]["started_at"], str)
    finally:
        refresh.reset_state()


async def test_the_orders_payload_survives_a_running_refresh(auth_client, db):
    """GET /orders embeds the refresh status, so it shared the same defect.

    The screen calls this on load; a 500 here is the "Could not reach the server" banner
    the owner actually saw, with the picking sheet never rendering at all.
    """
    import json

    await _seed(db)
    refresh.reset_state()
    refresh.STATE.update({"running": True, "started_at": datetime.utcnow()})
    try:
        r = await auth_client.get("/orders")
        assert r.status_code == 200, r.text
        json.dumps(r.json())
    finally:
        refresh.reset_state()


# ─── Today's dispatch: the warehouse's screen ────────────────────────────────


def _dispatched(order_id, **overrides):
    """An order that IS today's dispatch: due today, and labelled by Amazon."""
    return _row(order_id, status="Shipped", easyship_status="PendingPickUp", **overrides)


async def test_the_dispatch_route_returns_parents_sizes_and_packed(auth_client, db):
    """One payload behind the screen and the PDF, so paper and monitor cannot disagree."""
    await repository.upsert_orders(db, [_dispatched("403-1")])
    await repository.replace_items(db, "403-1", [
        {"asin": KNOWN_ASIN, "seller_sku": "cs-500", "title": "Chana Sattu",
         "quantity_ordered": 4}
    ])

    body = (await auth_client.get("/orders/dispatch")).json()

    assert body["pack_date"] == body["today_ist"], "the save date is not today in IST"
    parents = body["sheet"]["parents"]
    assert len(parents) == 1
    assert parents[0]["product"] == "Chana Sattu"
    assert parents[0]["units"] == 4
    assert parents[0]["kg"] == pytest.approx(2.0)          # 4 x 0.5 kg
    assert parents[0]["sizes"][0]["packed"] == 0, "nothing has been packed yet"
    assert body["sheet"]["orders"][0]["amazon_order_id"] == "403-1"


async def test_the_dispatch_route_excludes_the_ecom_teams_orders(auth_client, db):
    """An order still in `Unshipped` is theirs to ship on Seller Central.

    `_row`'s default IS that state, which is why this seeds it plainly: the sheet must come
    back empty rather than showing the floor work it cannot do.
    """
    await _seed(db)                                        # default: Unshipped/PendingSchedule
    body = (await auth_client.get("/orders/dispatch")).json()
    assert body["sheet"]["parents"] == [], "an unshipped order reached the dispatch sheet"
    assert body["sheet"]["totals"]["orders"] == 0


async def test_packed_counts_round_trip_through_the_route(auth_client, db):
    """Type, save, reload — the number has to still be there."""
    await repository.upsert_orders(db, [_dispatched("403-1")])
    await repository.replace_items(db, "403-1", [
        {"asin": KNOWN_ASIN, "quantity_ordered": 10}
    ])
    today = datetime.now(logic.IST).date().isoformat()

    r = await auth_client.post(
        f"/orders/packed/{today}", json={"entries": [{"asin": KNOWN_ASIN, "units": 6}]}
    )
    assert r.status_code == 200, r.text
    assert r.json()["packed"][KNOWN_ASIN] == 6

    body = (await auth_client.get("/orders/dispatch")).json()
    size = body["sheet"]["parents"][0]["sizes"][0]
    assert size["packed"] == 6
    assert size["remaining"] == 4


async def test_saving_against_another_day_is_refused(auth_client, db):
    """The SERVER decides what today is, not the browser.

    A laptop set to another timezone — or a page left open past midnight IST — would
    otherwise file this morning's count against yesterday, a silent off-by-one-day nobody
    notices until the numbers are being reconciled. Refused with the real date, so the screen
    can say "reload" instead of quietly moving the work.
    """
    await _seed(db)
    r = await auth_client.post(
        "/orders/packed/2026-01-01", json={"entries": [{"asin": KNOWN_ASIN, "units": 3}]}
    )
    assert r.status_code == 409, r.text
    assert r.json()["pack_date"] == datetime.now(logic.IST).date().isoformat()


async def test_a_malformed_packed_body_is_refused(auth_client, db):
    """`entries` must be a list. A bare object would silently save nothing."""
    await _seed(db)
    today = datetime.now(logic.IST).date().isoformat()
    r = await auth_client.post(f"/orders/packed/{today}", json={"entries": {"asin": "x"}})
    assert r.status_code == 400, r.text


async def test_the_dispatch_sheet_downloads_as_a_pdf(auth_client, db):
    """The floor's copy. A PDF because it is read at a bench and ticked with a pen."""
    await repository.upsert_orders(db, [_dispatched("403-1")])
    await repository.replace_items(db, "403-1", [
        {"asin": KNOWN_ASIN, "seller_sku": "cs-500", "quantity_ordered": 3}
    ])

    r = await auth_client.get("/orders/download/dispatch.pdf")
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == "application/pdf"
    assert r.content.startswith(b"%PDF-"), "not a PDF"
    assert len(r.content) > 1000, "the PDF is suspiciously small"


async def test_an_empty_dispatch_day_still_produces_a_readable_pdf(auth_client, db):
    """Nothing due is a normal morning, not an error.

    Asserted because an empty table is exactly where a document builder raises — and the
    download failing would read as a broken app on the quietest day of the week.
    """
    await _seed(db)                                        # nothing dispatched
    r = await auth_client.get("/orders/download/dispatch.pdf")
    assert r.status_code == 200, r.text
    assert r.content.startswith(b"%PDF-")


async def test_ops_may_reach_the_dispatch_screen_but_is_not_told_it_is_admin(
    auth_client, db
):
    """`is_admin` gates the owner's three extra groups on the screen.

    The flag travels in the payload only so the page can decide what to draw; the guard that
    matters is on the routes. An admin session must report True, or the owner loses the
    picking sheet he still needs.
    """
    await _seed(db)
    body = (await auth_client.get("/orders/dispatch")).json()
    assert body["is_admin"] is True, "an admin session was not told so"


# ─── Purchasing: raw stock and what to buy ───────────────────────────────────


async def test_the_dispatch_payload_carries_the_purchasing_view(auth_client, db):
    """Tab 1 reads the same payload as tabs 2 and 3, so nothing can disagree."""
    await repository.upsert_orders(db, [_dispatched("403-1")])
    await repository.replace_items(db, "403-1", [
        {"asin": KNOWN_ASIN, "seller_sku": "cs-500", "quantity_ordered": 10}
    ])

    body = (await auth_client.get("/orders/dispatch")).json()

    assert "purchasing" in body, "the purchasing view is missing from the payload"
    row = next(row for row in body["purchasing"]["rows"]
               if row["product"] == "Chana Sattu")
    assert row["ordered_kg"] == pytest.approx(5.0)        # 10 x 0.5 kg
    assert row["raw_kg"] == 0.0
    assert row["to_buy_kg"] == pytest.approx(5.0)
    assert row["covered"] is False


async def test_raw_stock_saves_and_reaches_the_purchasing_view(auth_client, db):
    """Type a stock figure, reload, and `to_buy` reflects it."""
    await repository.upsert_orders(db, [_dispatched("403-1")])
    await repository.replace_items(db, "403-1", [
        {"asin": KNOWN_ASIN, "quantity_ordered": 10}
    ])

    r = await auth_client.post(
        "/orders/raw-stock", json={"entries": [{"product": "Chana Sattu", "raw_kg": 3.5}]}
    )
    assert r.status_code == 200, r.text
    assert r.json()["raw_stock"]["Chana Sattu"] == pytest.approx(3.5)

    body = (await auth_client.get("/orders/dispatch")).json()
    row = next(row for row in body["purchasing"]["rows"]
               if row["product"] == "Chana Sattu")
    assert row["raw_kg"] == pytest.approx(3.5)
    assert row["to_buy_kg"] == pytest.approx(1.5)         # 5.0 ordered - 3.5 on hand


async def test_the_raw_stock_response_is_json_safe(auth_client, db):
    """`Numeric` returns `Decimal`, which JSONResponse cannot serialise.

    Asserted over HTTP rather than on the repository, because that is where the failure
    surfaced last time: a 500 and a "Could not reach the server" banner, found in a browser.
    """
    import json

    await _seed(db)
    r = await auth_client.post(
        "/orders/raw-stock", json={"entries": [{"product": "Chana Sattu", "raw_kg": 12.25}]}
    )
    assert r.status_code == 200, r.text
    json.dumps(r.json())


async def test_a_malformed_raw_stock_body_is_refused(auth_client, db):
    """`entries` must be a list, or the save silently stores nothing."""
    await _seed(db)
    r = await auth_client.post("/orders/raw-stock", json={"entries": {"product": "x"}})
    assert r.status_code == 400, r.text


async def test_the_raw_stock_route_takes_no_date(auth_client, db):
    """Unlike `/packed/{pack_date}`, and deliberately.

    A packed count belongs to a day, so that route refuses any date but today — a laptop in
    another timezone would otherwise file this morning's count against yesterday. Raw stock is
    standing, so a date here would make the number unreachable tomorrow.
    """
    # A DISPATCHED order, not the `_seed` default: that one is Unshipped/PendingSchedule, which
    # is correctly excluded from the dispatch sheet — so the purchasing view would hold no rows
    # at all and this test would prove nothing about the route.
    await repository.upsert_orders(db, [_dispatched("403-1")])
    await repository.replace_items(db, "403-1", [
        {"asin": KNOWN_ASIN, "quantity_ordered": 4}
    ])
    response = await auth_client.post(
        "/orders/raw-stock", json={"entries": [{"product": "Chana Sattu", "raw_kg": 8}]}
    )
    assert response.status_code == 200, response.text
    body = (await auth_client.get("/orders/dispatch")).json()
    row = next(row for row in body["purchasing"]["rows"]
               if row["product"] == "Chana Sattu")
    assert row["raw_kg"] == pytest.approx(8.0)
    # 4 x 0.5 kg ordered against 8 kg on hand, so nothing to buy.
    assert row["to_buy_kg"] == 0.0
    assert row["covered"] is True


async def test_every_download_variant_returns_a_file(auth_client, db):
    """Five files, two formats, one aggregation behind all of them."""
    await repository.upsert_orders(db, [_dispatched("403-1")])
    await repository.replace_items(db, "403-1", [
        {"asin": KNOWN_ASIN, "seller_sku": "cs-500", "quantity_ordered": 4}
    ])

    for tab in ("all", "weight", "sku", "orders"):
        r = await auth_client.get(f"/orders/download/dispatch.pdf?tab={tab}")
        assert r.status_code == 200, f"pdf tab={tab}: {r.text[:200]}"
        assert r.content.startswith(b"%PDF-"), f"pdf tab={tab} is not a PDF"

    for tab in ("all", "weight", "sku", "orders", "tobuy"):
        r = await auth_client.get(f"/orders/download/dispatch.xlsx?tab={tab}")
        assert r.status_code == 200, f"xlsx tab={tab}: {r.text[:200]}"
        # Every .xlsx is a zip archive, so it starts with PK.
        assert r.content[:2] == b"PK", f"xlsx tab={tab} is not a workbook"


async def test_an_unknown_download_tab_is_refused(auth_client, db):
    """A typo must not silently export the wrong section."""
    await _seed(db)
    r = await auth_client.get("/orders/download/dispatch.xlsx?tab=nonsense")
    assert r.status_code == 400, r.text
    assert "nonsense" in r.json()["error"]


async def test_tobuy_is_excel_only(auth_client, db):
    """It is pasted into a supplier email, not read at a bench.

    Refused rather than silently served as the combined PDF, which is not what a caller asking
    for `tab=tobuy` would expect to receive.
    """
    await _seed(db)
    r = await auth_client.get("/orders/download/dispatch.pdf?tab=tobuy")
    assert r.status_code == 400, r.text


async def test_the_downloads_agree_with_the_screen_about_the_totals(auth_client, db):
    """The property the single aggregation exists to guarantee.

    A download that aggregated separately is how a printed sheet and a monitor start disagreeing
    about a quantity — the failure `_document_rows` prevents on the shipment side.
    """
    import io as _io

    from openpyxl import load_workbook

    await repository.upsert_orders(db, [_dispatched("403-1"), _dispatched("403-2")])
    for order_id in ("403-1", "403-2"):
        await repository.replace_items(db, order_id, [
            {"asin": KNOWN_ASIN, "seller_sku": "cs-500", "quantity_ordered": 3}
        ])

    screen = (await auth_client.get("/orders/dispatch")).json()
    expected_units = screen["sheet"]["totals"]["units"]

    r = await auth_client.get("/orders/download/dispatch.xlsx?tab=sku")
    book = load_workbook(_io.BytesIO(r.content))
    values = [[cell.value for cell in row] for row in book.active.iter_rows()]
    total_row = next(row for row in values if row and row[0] == "TOTAL")
    assert total_row[3] == expected_units, (
        f"the workbook says {total_row[3]} units, the screen says {expected_units}"
    )


# ─── The three-tab template ──────────────────────────────────────────────────


def test_the_template_has_three_tabs_each_with_its_own_search():
    """Three tabs, three search boxes. Asserted on the markup because the ids are contracts.

    `tests/test_template_render_targets.py` already fails any getElementById that is written to
    without a matching element — this asserts the other direction, that the tabs the design calls
    for actually exist.
    """
    source = _orders_source()
    for tab in ("tab-weight", "tab-sku", "tab-orders"):
        assert f'id="{tab}"' in source, f"{tab} panel is missing"
    for box in ("search-weight", "search-sku", "search-orders"):
        assert f'id="{box}"' in source, f"{box} is missing, so that tab cannot be filtered"


def test_the_poll_never_re_renders_a_tab_with_inputs():
    """A 60-second poll that redrew the SKU tab would eat the packer's keystrokes mid-number.

    The poll exists so tab 3's Amazon statuses update themselves. Tabs 1 and 2 hold number
    boxes, so the poll must touch neither — asserted on the source because the failure is
    invisible until someone is typing.
    """
    source = _orders_source()
    assert "function pollOrders(" in source, "the orders poll is missing"
    start = source.index("function pollOrders(")
    body = source[start:start + 1200]
    for forbidden in ("renderWeight(", "renderSku("):
        assert forbidden not in body, (
            f"pollOrders calls {forbidden} — it would redraw a tab containing inputs and "
            "discard whatever the packer was typing"
        )


def test_the_packed_and_raw_stock_saves_post_to_their_own_routes():
    """Two facts, two routes, and only one of them carries a date."""
    source = _orders_source()
    assert "/orders/packed/${" in source, "the packed save does not send a date"
    assert '"/orders/raw-stock"' in source, "the raw stock save is missing"
    assert "/orders/raw-stock/${" not in source, (
        "the raw stock route must not take a date: raw stock is standing, and a date would "
        "make the number unreachable tomorrow"
    )


def test_a_covered_product_renders_a_dash_not_a_zero():
    """A zero in a purchasing column reads as a measurement, not as "nothing to do".

    Asserted on the template because it is a rendering decision: the number is already 0.0 in
    the payload, and printing it literally is what makes the tab ambiguous.

    The condition is the LIVE shortfall rather than the server's `covered` flag, deliberately:
    the dash has to appear the moment enough raw stock is typed, before anything is saved. A
    flag-driven version would keep showing a number until the page reloaded.
    """
    source = _orders_source()
    dash_rule = "buy <= 0 ? '<span class=\"tag ok\">—</span>'"
    assert dash_rule in source, (
        "a covered product must render an em dash from the live shortfall; 0.00 reads as a "
        "weight someone measured"
    )
    # Both the initial render and the as-you-type update must apply it, or the dash appears on
    # load and turns back into a number the moment the box is touched.
    assert source.count(dash_rule) == 2, (
        "the dash rule is not applied in both the table render and the live update"
    )
