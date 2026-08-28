"""The Portfolio routes, the stored snapshot, and the decision record.

Covers the three defects this app has already shipped once each and must not ship again:
Decimal reaching `JSONResponse`, a datetime reaching `JSONResponse`, and an N+1 query pattern
inside a page render.
"""
import json
from datetime import datetime, timedelta
from pathlib import Path

import pytest

from app.models import EconomicsSnapshot, Product, ProductDecision, RatingHistory
from app.portfolio import repository

pytestmark = pytest.mark.regression

FIXTURES = Path(__file__).parent / "fixtures"
WINDOW = ("2026-07-28", "2026-08-26")


def _rows():
    return json.loads((FIXTURES / "economics_rows.json").read_text(encoding="utf-8"))


async def _seed_snapshot(db, rows=None):
    stored = await repository.save_snapshot(db, WINDOW[0], WINDOW[1], rows or _rows())
    return stored


async def _seed_rating(db, asin, rating, count, *, days_ago=0):
    product = Product(asin=asin, title=f"title {asin}", first_seen=datetime.utcnow())
    db.add(product)
    await db.flush()
    db.add(RatingHistory(
        product_id=product.id, rating=rating, rating_count=count,
        scraped_at=datetime.utcnow() - timedelta(days=days_ago),
    ))
    await db.commit()
    return product


# ─── The snapshot round trip ─────────────────────────────────────────────────


async def test_the_snapshot_round_trips_through_the_database(db):
    """Stored rows must come back in Amazon's own shape.

    **One input format for `logic.portfolio`, whether the rows arrived from the API a second ago
    or from the database.** A second code path for stored rows is how a cached dashboard starts
    disagreeing with a freshly refreshed one.
    """
    rows = _rows()
    stored = await _seed_snapshot(db, rows)
    assert stored == len(rows)

    back = await repository.load_snapshot(db)
    assert len(back) == len(rows)

    # The nested shape logic.size_row expects.
    sample = back[0]
    assert "childAsin" in sample and "parentAsin" in sample
    assert "orderedProductSales" in sample["sales"]
    assert isinstance(sample["fees"], list)
    assert isinstance(sample["ads"], list)
    assert "total" in sample["netProceeds"]


async def test_the_same_window_upserts_rather_than_doubling_the_portfolio(db):
    """Pressing Refresh twice in a morning must correct the figures, not duplicate them."""
    from sqlalchemy import func, select

    rows = _rows()
    for _ in range(3):
        await _seed_snapshot(db, rows)

    count = (await db.execute(select(func.count()).select_from(EconomicsSnapshot))).scalar()
    assert count == len({r["childAsin"] for r in rows}), (
        f"{count} rows stored for {len(rows)} products — the upsert is inserting"
    )


async def test_no_decimal_survives_the_load(db):
    """**A Decimal reaching JSONResponse is a 500 and a blank screen.**

    This app has shipped that defect twice — datetimes on the orders payload, then `raw_kg` on
    the purchasing view — and both were found in a browser on production. `load_snapshot` casts
    once so every route inherits the fix rather than remembering it.
    """
    await _seed_snapshot(db)
    back = await repository.load_snapshot(db)
    json.dumps(back)            # would raise on a Decimal


async def test_the_latest_window_is_the_newest_END_date(db):
    """Ordered by window_end, not by fetched_at.

    Re-running an OLDER window — a deliberate look at last month, say — would otherwise become
    "the latest" and shift the whole dashboard backwards in time without saying so.
    """
    rows = _rows()[:2]
    await repository.save_snapshot(db, "2026-06-01", "2026-06-30", rows)
    await repository.save_snapshot(db, "2026-07-28", "2026-08-26", rows)
    # Stored last, but covers an older period.
    await repository.save_snapshot(db, "2026-05-01", "2026-05-31", rows)

    assert await repository.latest_window(db) == ("2026-07-28", "2026-08-26")


# ─── Ratings: one query, not one per product ─────────────────────────────────


async def test_the_latest_rating_per_product_comes_back(db):
    """Only the newest scrape per product, not every row in the history."""
    product = await _seed_rating(db, "B0RATE0001", 3.8, 100, days_ago=30)
    db.add(RatingHistory(
        product_id=product.id, rating=4.3, rating_count=180, scraped_at=datetime.utcnow()
    ))
    await db.commit()

    ratings = await repository.load_ratings(db)
    assert ratings["B0RATE0001"]["rating"] == 4.3
    assert ratings["B0RATE0001"]["rating_count"] == 180


async def test_ratings_load_in_ONE_query_regardless_of_catalogue_size(db):
    """**The N+1 the old tab had.**

    It ran two queries per ASIN inside a loop: at 262 products that is 524 round trips to render
    one page, and it grew with the catalogue. Asserted by COUNTING queries rather than by timing,
    so it fails deterministically if someone reintroduces the loop.
    """
    for index in range(12):
        await _seed_rating(db, f"B0RATE{index:04d}", 4.0 + index / 100, 50 + index)

    # Counted by wrapping the session's own execute, which works regardless of how the engine
    # is wired — the previous version reached for `engine.sync_engine` and broke on the test
    # engine, testing the harness rather than the code.
    seen = []
    original = db.execute

    async def counting_execute(statement, *args, **kwargs):
        seen.append(str(statement))
        return await original(statement, *args, **kwargs)

    db.execute = counting_execute
    try:
        ratings = await repository.load_ratings(db)
    finally:
        db.execute = original

    assert len(ratings) == 12
    assert len(seen) == 1, (
        f"{len(seen)} queries to load 12 ratings — the per-ASIN loop is back:\n"
        + "\n".join(q[:120] for q in seen[:4])
    )


async def test_a_rating_is_a_float_not_a_decimal(db):
    """`Numeric(2, 1)` returns Decimal, which JSONResponse cannot serialise."""
    await _seed_rating(db, "B0RATE0001", 4.2, 477)
    ratings = await repository.load_ratings(db)
    assert isinstance(ratings["B0RATE0001"]["rating"], float)
    json.dumps(ratings)


# ─── Decisions ───────────────────────────────────────────────────────────────


async def test_a_decision_is_stored_and_clearing_it_deletes_the_row(db):
    """Absence is the single representation of "not decided".

    A stored empty decision would be a second way to say the same thing, and no stale note may
    sit behind a cleared flag — the same rule the per-order packed tick follows.
    """
    from sqlalchemy import func, select

    await repository.save_decision(
        db, "B0PARENT01", "kill", note="ads off, sell through", decided_by="owner"
    )
    stored = await repository.load_decisions(db)
    assert stored["B0PARENT01"]["decision"] == "kill"
    assert stored["B0PARENT01"]["note"] == "ads off, sell through"
    assert stored["B0PARENT01"]["decided_at"], "no timestamp, so it cannot be reviewed later"

    await repository.save_decision(db, "B0PARENT01", "")
    assert await repository.load_decisions(db) == {}
    count = (await db.execute(select(func.count()).select_from(ProductDecision))).scalar()
    assert count == 0


async def test_re_deciding_replaces_rather_than_appending(db):
    """One standing decision per product, so the dashboard never guesses which row is current."""
    from sqlalchemy import func, select

    for decision in ("watch", "keep", "kill"):
        await repository.save_decision(db, "B0PARENT01", decision, note=decision)

    count = (await db.execute(select(func.count()).select_from(ProductDecision))).scalar()
    assert count == 1
    assert (await repository.load_decisions(db))["B0PARENT01"]["decision"] == "kill"


# ─── The routes ──────────────────────────────────────────────────────────────


async def test_the_portfolio_route_returns_products_and_totals(auth_client, db):
    """One payload behind the screen and the export, so they cannot disagree."""
    await _seed_snapshot(db)
    response = await auth_client.get("/portfolio")
    assert response.status_code == 200, response.text
    body = response.json()

    assert body["parents"], "no products in the payload"
    assert body["totals"]["parents"] == len(body["parents"])
    assert "verdicts" in body["totals"]
    assert body["pre_cogs"] is True, (
        "the payload does not flag that margins exclude manufacturing cost, so an export "
        "could present the number without the caveat"
    )
    for parent in body["parents"]:
        assert parent["verdict"]
        assert parent["verdict_reason"]


async def test_the_whole_payload_is_json_safe(auth_client, db):
    """Decimal AND datetime, at the HTTP boundary rather than only in the repository."""
    await _seed_snapshot(db)
    await _seed_rating(db, "B0CWGXYLT6", 4.0, 355)
    await repository.save_decision(db, "B0DR322QYG", "watch", note="seasonal")

    response = await auth_client.get("/portfolio")
    assert response.status_code == 200, response.text
    json.dumps(response.json())


async def test_an_empty_database_renders_rather_than_erroring(auth_client, db):
    """Before the first refresh there is nothing stored, and that must not 500."""
    response = await auth_client.get("/portfolio")
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["parents"] == []
    assert body["totals"]["parents"] == 0


async def test_the_decision_route_records_and_clears(auth_client, db):
    """The owner's judgement, saved and revocable."""
    await _seed_snapshot(db)
    response = await auth_client.post(
        "/portfolio/decision",
        json={"parent_asin": "B0DR322QYG", "decision": "kill", "note": "ads off"},
    )
    assert response.status_code == 200, response.text
    assert response.json()["decisions"]["B0DR322QYG"]["decision"] == "kill"

    cleared = await auth_client.post(
        "/portfolio/decision", json={"parent_asin": "B0DR322QYG", "decision": ""}
    )
    assert cleared.status_code == 200, cleared.text
    assert "B0DR322QYG" not in cleared.json()["decisions"]


async def test_a_decision_stores_the_figures_it_was_taken_on(auth_client, db):
    """**Why decisions are kept at all.**

    Revisiting a kill in three months otherwise means trusting memory about what the margin was.
    Being able to ask "I marked this KILL at -56.8%; what is it now?" is the whole reason this
    replaced a bare name in a JSON set.
    """
    await _seed_snapshot(db)
    await auth_client.post(
        "/portfolio/decision",
        json={"parent_asin": "B0DR322QYG", "decision": "kill", "note": "why"},
    )
    stored = await repository.load_decisions(db)
    snapshot = stored["B0DR322QYG"]["snapshot"]
    assert snapshot, "no figures were recorded with the decision"
    assert "verdict" in snapshot and "net_pct" in snapshot and "tacos" in snapshot


async def test_an_unknown_decision_is_refused(auth_client, db):
    """A typo must not create a fourth category the dashboard can neither filter nor count."""
    response = await auth_client.post(
        "/portfolio/decision", json={"parent_asin": "B0DR322QYG", "decision": "delete"}
    )
    assert response.status_code == 400, response.text
    assert "kill" in response.json()["error"]


async def test_a_decision_without_a_product_is_refused(auth_client, db):
    response = await auth_client.post("/portfolio/decision", json={"decision": "kill"})
    assert response.status_code == 400, response.text


async def test_a_malformed_decision_body_is_refused(auth_client, db):
    response = await auth_client.post(
        "/portfolio/decision", content=b"not json",
        headers={"Content-Type": "application/json"},
    )
    assert response.status_code == 400, response.text


async def test_the_refresh_status_route_is_json_safe(auth_client, db):
    """The status carries datetimes, and this is the route polled while a refresh runs.

    The orders feature shipped exactly this bug on exactly this kind of path — the one that fires
    when someone is trying to find out what is happening.
    """
    from app.portfolio import refresh

    refresh.reset_state()
    refresh.STATE.update({"running": True, "started_at": datetime.utcnow()})
    try:
        response = await auth_client.get("/portfolio/refresh-status")
        assert response.status_code == 200, response.text
        json.dumps(response.json())
        assert isinstance(response.json()["started_at"], str)
    finally:
        refresh.reset_state()


async def test_a_second_refresh_is_refused_while_one_runs(auth_client, db):
    """Two concurrent Data Kiosk queries would spend two minutes to store the same rows."""
    from app.portfolio import refresh

    refresh.reset_state()
    refresh.STATE.update({"running": True, "started_at": datetime.utcnow()})
    try:
        response = await auth_client.post("/portfolio/refresh")
        assert response.status_code == 409, response.text
        json.dumps(response.json())
    finally:
        refresh.reset_state()


async def test_the_portfolio_downloads_as_a_workbook(auth_client, db):
    """Built through the same aggregation as the screen, so the file cannot disagree."""
    await _seed_snapshot(db)
    response = await auth_client.get("/portfolio/download.xlsx")
    assert response.status_code == 200, response.text
    assert response.content[:2] == b"PK", "not a workbook"
    assert len(response.content) > 1000


async def test_the_export_states_that_margins_are_pre_cogs(auth_client, db):
    """A file leaves the app and gets read without the screen's banner beside it.

    So the caveat travels IN the document — otherwise a spreadsheet showing "+8.8% net" gets
    forwarded to someone who reads it as profit.
    """
    import io

    from openpyxl import load_workbook

    await _seed_snapshot(db)
    response = await auth_client.get("/portfolio/download.xlsx")
    book = load_workbook(io.BytesIO(response.content))
    text = " ".join(
        str(cell.value or "")
        for row in book.active.iter_rows(max_row=6)
        for cell in row
    )
    assert "PRE-COGS" in text.upper(), text[:300]


async def test_the_portfolio_pages_are_gated_on_the_area(client, db):
    """Unauthenticated requests must not reach the figures."""
    for path in ("/portfolio", "/portfolio/refresh-status", "/portfolio/download.xlsx"):
        response = await client.get(path)
        assert response.status_code in (303, 401, 403), f"{path} -> {response.status_code}"


# ─── The refresh job ─────────────────────────────────────────────────────────


async def test_the_refresh_stores_what_it_fetched(monkeypatch, db_schema):
    """The job's whole point: fetch, store, and record that it ran."""
    from app.database import async_session
    from app.portfolio import economics, refresh

    rows = _rows()

    async def fake_fetch(**kwargs):
        on_progress = kwargs.get("on_progress")
        if on_progress:
            on_progress("econ_poll", 1, 2)
        # Four values now: the ASIN grain, the per-SKU grain, and the window.
        return rows, [], WINDOW[0], WINDOW[1]

    monkeypatch.setattr(economics, "fetch_economics", fake_fetch)
    refresh.reset_state()
    result = await refresh.run(db_factory=async_session)

    assert result["phase"] == "done", result
    assert result["rows"] == len(rows)
    assert result["percent"] == 100
    assert result["running"] is False
    # No ads credentials in the test environment, so ACOS is skipped — and that must be reported
    # as its own note rather than failing the refresh, because the margins are what matter.
    assert result["ads_error"], "a skipped ACOS phase said nothing at all"
    assert "not configured" in result["ads_error"]

    async with async_session() as session:
        assert len(await repository.load_snapshot(session)) == len(rows)
        last = await repository.last_refresh(session)
        assert last and last["rows_stored"] == len(rows)


async def test_a_failed_refresh_is_recorded_and_does_not_wedge_the_flag(monkeypatch, db_schema):
    """**A crash must not leave `running` True for the life of the process.**

    It would refuse every later refresh, including the nightly one, and the only symptom would be
    a dashboard that quietly stopped updating.
    """
    from app.database import async_session
    from app.portfolio import economics, refresh
    from app.shipment.spapi import SpApiError

    async def fake_fetch(**kwargs):
        raise SpApiError("Amazon said no")

    monkeypatch.setattr(economics, "fetch_economics", fake_fetch)
    refresh.reset_state()
    result = await refresh.run(db_factory=async_session)

    assert result["phase"] == "failed"
    assert "Amazon said no" in result["error"]
    assert result["running"] is False, "the running flag stayed set after a failure"

    async with async_session() as session:
        last = await repository.last_refresh(session)
        assert last and "Amazon said no" in last["error"], (
            "a failed refresh left no record, so a stale dashboard cannot say why"
        )


async def test_an_unexpected_crash_also_clears_the_flag(monkeypatch, db_schema):
    """Not just SpApiError — any exception."""
    from app.database import async_session
    from app.portfolio import economics, refresh

    async def fake_fetch(**kwargs):
        raise ValueError("something nobody predicted")

    monkeypatch.setattr(economics, "fetch_economics", fake_fetch)
    refresh.reset_state()
    result = await refresh.run(db_factory=async_session)

    assert result["running"] is False
    assert result["phase"] == "failed"


async def test_a_cancelled_refresh_also_clears_the_flag(monkeypatch, db_schema):
    """**The hole a mutation found: `CancelledError` is a BaseException.**

    `except Exception` does not catch it, so only the `finally` clears the flag — and this job
    runs as a fire-and-forget `asyncio.create_task`, which is precisely the kind of thing that
    gets cancelled when the event loop shuts down or the task is torn down mid-flight.

    Without `finally`, `running` would stay True for the life of the process and EVERY later
    refresh — including the nightly one — would be silently refused. The only symptom would be a
    dashboard that quietly stopped updating, which is the hardest class of bug to notice here.

    The three exception tests above all passed with `finally` removed, so this is the one that
    actually pins it.
    """
    import asyncio

    from app.database import async_session
    from app.portfolio import economics, refresh

    async def fake_fetch(**kwargs):
        raise asyncio.CancelledError()

    monkeypatch.setattr(economics, "fetch_economics", fake_fetch)
    refresh.reset_state()
    try:
        await refresh.run(db_factory=async_session)
    except asyncio.CancelledError:
        pass        # re-raised on purpose: cancellation must not be swallowed

    assert refresh.STATE["running"] is False, (
        "a cancelled refresh left `running` set, so every later refresh would be refused "
        "and the dashboard would silently stop updating"
    )
    refresh.reset_state()


async def test_a_concurrent_refresh_is_refused_rather_than_raising(db_schema):
    """The nightly job overlapping a manual one is a no-op, not an error in the log."""
    from app.database import async_session
    from app.portfolio import refresh

    refresh.reset_state()
    refresh.STATE["running"] = True
    try:
        result = await refresh.run(db_factory=async_session)
        assert result["refused"] is True
    finally:
        refresh.reset_state()


def test_the_progress_bar_never_steps_backwards():
    """A bar that jumps back reads as a fault.

    It matters here because the poll phase is a fraction of a CEILING: a query finishing on
    attempt 3 of 40 would otherwise report 15% after the download phase had already reported 85%.
    """
    from app.portfolio import refresh

    refresh.reset_state()
    refresh._progress("poll", 30, 40)
    high = refresh.STATE["percent"]
    refresh._progress("poll", 1, 40)
    assert refresh.STATE["percent"] == high, "the bar went backwards"
    refresh.reset_state()


# ─── The per-SKU rows must never reach the totals ────────────────────────────


async def test_sku_rows_live_beside_the_asin_rows_without_doubling_the_totals(db):
    """**The invariant that keeps every figure on the dashboard honest.**

    Both grains share one table: ASIN-level rows carry the totals (`seller_sku IS NULL`), and
    per-SKU rows carry the merchant/FBA split. A product's two channel rows SUM to its ASIN row, so
    if `load_snapshot` returned both, every number on screen would roughly double.

    Asserted by storing both and checking the totals are unchanged — the failure would otherwise be
    a plausible-looking dashboard reporting twice the real revenue.
    """
    asin_rows = _rows()
    before = await repository.load_snapshot(db)
    await _seed_snapshot(db, asin_rows)
    after_asin_only = await repository.load_snapshot(db)

    sku_rows = [
        {"childAsin": r["childAsin"], "parentAsin": r.get("parentAsin"),
         "msku": f"sku-{r['childAsin']}",
         "sales": {"orderedProductSales": {"amount": 100.0}, "netUnitsSold": 1,
                   "unitsOrdered": 1, "unitsRefunded": 0,
                   "refundedProductSales": {"amount": 0.0}},
         "netProceeds": {"total": {"amount": 10.0}}, "fees": [], "ads": []}
        for r in asin_rows
    ]
    stored_skus = await repository.save_sku_snapshot(db, WINDOW[0], WINDOW[1], sku_rows)
    assert stored_skus == len(sku_rows), "the per-SKU rows were not stored"

    after_both = await repository.load_snapshot(db)
    assert len(after_both) == len(after_asin_only), (
        f"load_snapshot returned {len(after_both)} rows after adding per-SKU rows, up from "
        f"{len(after_asin_only)} — the SKU rows are leaking into the totals and every figure "
        "on the dashboard would be inflated"
    )
    # And they ARE retrievable by their own loader.
    assert len(await repository.load_sku_snapshot(db)) == stored_skus


async def test_re_saving_the_asin_rows_does_not_overwrite_a_sku_row(db):
    """A refresh must not clobber the split it stored moments earlier.

    Without the `seller_sku IS NULL` filter in `save_snapshot`, the SELECT could match a per-SKU
    row and overwrite it with an ASIN total — corrupting the split rather than failing loudly.
    """
    asin_rows = _rows()[:2]
    await _seed_snapshot(db, asin_rows)
    sku_rows = [{
        "childAsin": asin_rows[0]["childAsin"], "msku": "one-sku",
        "sales": {"orderedProductSales": {"amount": 55.0}, "netUnitsSold": 5,
                  "unitsOrdered": 5, "unitsRefunded": 0,
                  "refundedProductSales": {"amount": 0.0}},
        "netProceeds": {"total": {"amount": 5.0}}, "fees": [], "ads": [],
    }]
    await repository.save_sku_snapshot(db, WINDOW[0], WINDOW[1], sku_rows)

    await _seed_snapshot(db, asin_rows)          # the refresh runs again
    kept = await repository.load_sku_snapshot(db)
    assert len(kept) == 1, "the per-SKU row was destroyed by a re-save of the ASIN rows"
    assert kept[0]["sales"]["orderedProductSales"]["amount"] == 55.0


# ─── Ad figures ──────────────────────────────────────────────────────────────


async def test_ad_rows_round_trip_and_roll_up_to_the_asin(db):
    """Two shapes from one store: per (asin, sku) for the split, per ASIN for the row."""
    rows = [
        {"child_asin": "B0AAA00001", "seller_sku": "2kg kc",
         "cost": 1444.0, "attributed_sales": 0.0, "purchases": 0, "clicks": 12, "impressions": 900},
        {"child_asin": "B0AAA00001", "seller_sku": "2kg kc FBA",
         "cost": 5176.0, "attributed_sales": 14254.0, "purchases": 8, "clicks": 40,
         "impressions": 3000},
    ]
    stored = await repository.save_ads_snapshot(db, WINDOW[0], WINDOW[1], rows)
    assert stored == 2

    # latest_window() reads the ECONOMICS table, so the window has to exist there too.
    await _seed_snapshot(db)
    by_asin, by_sku = await repository.load_ads_snapshot(db, WINDOW)
    assert by_sku[("B0AAA00001", "2kg kc")]["cost"] == 1444.0
    assert by_asin["B0AAA00001"]["cost"] == pytest.approx(6620.0)
    assert by_asin["B0AAA00001"]["attributed_sales"] == pytest.approx(14254.0)
    assert by_asin["B0AAA00001"]["clicks"] == 52


async def test_ad_rows_upsert_rather_than_doubling_the_spend(db):
    """Pressing Refresh twice must correct the ad figures, not double them."""
    from sqlalchemy import func, select

    from app.models import AdsSnapshot

    rows = [{"child_asin": "B0AAA00001", "seller_sku": "s", "cost": 100.0,
             "attributed_sales": 200.0, "purchases": 1, "clicks": 2, "impressions": 3}]
    for _ in range(3):
        await repository.save_ads_snapshot(db, WINDOW[0], WINDOW[1], rows)

    count = (await db.execute(select(func.count()).select_from(AdsSnapshot))).scalar()
    assert count == 1, f"{count} rows for one (asin, sku) — the upsert is inserting"


async def test_no_decimal_reaches_json_from_the_ad_rows(db):
    """`Numeric` returns Decimal, which JSONResponse cannot serialise."""
    await repository.save_ads_snapshot(db, WINDOW[0], WINDOW[1], [
        {"child_asin": "B0AAA00001", "seller_sku": "s", "cost": 12.34,
         "attributed_sales": 56.78, "purchases": 1, "clicks": 2, "impressions": 3},
    ])
    await _seed_snapshot(db)
    by_asin, by_sku = await repository.load_ads_snapshot(db, WINDOW)
    json.dumps({"by_asin": by_asin, "by_sku": list(by_sku.values())})


# ─── Thresholds ──────────────────────────────────────────────────────────────


async def test_thresholds_round_trip_and_reset_by_deleting_the_row(db):
    """An empty save RESETS, so "reset" and "never edited" are one state rather than two."""
    from sqlalchemy import func, select

    from app.models import PortfolioSettings
    from app.portfolio import logic

    saved = await repository.save_settings(db, {"kill_tacos": 0.40}, updated_by="owner")
    assert saved["kill_tacos"] == 0.40
    assert saved["good_net"] == logic.DEFAULT_THRESHOLDS["good_net"], "unedited keys should default"

    reset = await repository.save_settings(db, {})
    assert reset == logic.thresholds_or_default({})
    count = (await db.execute(select(func.count()).select_from(PortfolioSettings))).scalar()
    assert count == 0, "resetting left a row behind, so 'reset' and 'default' are two states"


async def test_the_settings_route_reports_the_rules_in_words(auth_client, db):
    """The panel needs the explanation AND the defaults, so Reset needs no hardcoded numbers."""
    response = await auth_client.get("/portfolio/settings")
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["thresholds"] and body["defaults"]
    assert set(body["thresholds"]) == set(body["defaults"])
    for verdict in body["verdict_order"]:
        assert body["help"].get(verdict), f"{verdict} has no explanation"
    # The help text substitutes live values rather than naming a constant.
    assert "50%" in body["help"]["KILL"]


async def test_saving_a_threshold_changes_the_help_text_with_it(auth_client, db):
    """Help naming 50% while the rule fires at 40% would be worse than no help."""
    response = await auth_client.post(
        "/portfolio/settings", json={"thresholds": {"kill_tacos": 0.40}}
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["thresholds"]["kill_tacos"] == 0.40
    assert "40%" in body["help"]["KILL"], body["help"]["KILL"]


async def test_an_unknown_threshold_is_refused_rather_than_ignored(auth_client, db):
    """**A typo that appeared to save and changed nothing would be worse than an error.**

    The owner would believe a rule had moved when it had not, and act on verdicts computed the old
    way. So the route 400s and names the valid keys.
    """
    response = await auth_client.post(
        "/portfolio/settings", json={"thresholds": {"kil_tacos": 0.40}}
    )
    assert response.status_code == 400, response.text
    assert "kil_tacos" in response.json()["error"]
    assert "kill_tacos" in response.json()["error"], "the error should name the valid keys"


async def test_a_non_numeric_threshold_is_refused(auth_client, db):
    response = await auth_client.post(
        "/portfolio/settings", json={"thresholds": {"kill_tacos": "half"}}
    )
    assert response.status_code == 400, response.text


async def test_edited_thresholds_reach_the_dashboard(auth_client, db):
    """The whole point: change a rule, and the verdicts recompute from the stored rows.

    No Amazon call is involved, so this is instant — which is what makes the setting worth having.
    """
    await _seed_snapshot(db)
    before = (await auth_client.get("/portfolio")).json()
    await auth_client.post("/portfolio/settings", json={"thresholds": {"dead_units": 500}})
    after = (await auth_client.get("/portfolio")).json()

    assert after["thresholds"]["dead_units"] == 500
    # With every product below the volume floor, they all become DEAD.
    assert after["totals"]["verdicts"]["DEAD"] > before["totals"]["verdicts"]["DEAD"], (
        "raising the volume floor did not change any verdict, so the saved thresholds are "
        "not reaching logic.portfolio"
    )


# ─── Window selection ────────────────────────────────────────────────────────


async def test_a_window_longer_than_ninety_days_is_refused(auth_client, db):
    response = await auth_client.get("/portfolio?start=2026-01-01&end=2026-08-26")
    assert response.status_code == 400, response.text
    assert "90" in response.json()["error"]


async def test_a_window_including_today_is_refused(auth_client, db):
    """Today's figures are still settling: an ad charge lands hours after its sale.

    A range including today would show a punishing ACOS every morning that settled by evening — a
    number that moves on its own invites a decision the data cannot support.
    """
    from datetime import date as _date

    today = _date.today().isoformat()
    response = await auth_client.get(f"/portfolio?start=2026-08-01&end={today}")
    assert response.status_code == 400, response.text
    assert "still settling" in response.json()["error"]


async def test_a_reversed_window_is_refused(auth_client, db):
    response = await auth_client.get("/portfolio?start=2026-08-26&end=2026-08-01")
    assert response.status_code == 400, response.text


async def test_an_uncached_window_returns_empty_rather_than_fetching(auth_client, db):
    """**A GET must never block on a twelve-minute ad report.**

    So an unfetched window renders empty and the screen offers a Fetch button. Returning 200 with
    no rows rather than 404, because "we hold nothing for that range" is a valid answer about a
    valid range.
    """
    await _seed_snapshot(db)
    response = await auth_client.get("/portfolio?start=2026-06-01&end=2026-06-30")
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["parents"] == []
    assert body["window"] == ["2026-06-01", "2026-06-30"]


async def test_the_available_windows_are_reported_so_the_picker_can_mark_them(auth_client, db):
    """The cost of a click should be visible before clicking."""
    await _seed_snapshot(db)
    body = (await auth_client.get("/portfolio")).json()
    windows = body["windows_available"]
    assert windows, "no cached windows reported, so every range would look uncached"
    assert windows[0]["start"] == WINDOW[0] and windows[0]["end"] == WINDOW[1]
    assert windows[0]["rows"] > 0


async def test_the_payload_carries_the_acos_and_view_data_the_screen_needs(auth_client, db):
    """One request feeds the whole page; a second round trip per control would be slower and
    could disagree with the first."""
    await _seed_snapshot(db)
    body = (await auth_client.get("/portfolio")).json()
    for key in ("skus", "thresholds", "verdict_help", "verdict_order", "phase_labels",
                "windows_available", "acos_available", "max_window_days"):
        assert key in body, f"the payload is missing {key}"
    assert body["totals"]["sku_verdicts"], "no per-SKU verdict counts for the SKU view's chips"
    # Every phase the refresh can report must have a label, or the bar shows a raw key.
    from app.portfolio import refresh as portfolio_refresh
    assert set(portfolio_refresh.PHASE_BOUNDS) == set(body["phase_labels"])


# ─── The template ────────────────────────────────────────────────────────────


def _template() -> str:
    return (Path(__file__).parent.parent / "templates" / "portfolio.html").read_text(
        encoding="utf-8"
    )


def test_the_template_has_every_control_the_five_features_need():
    """Asserted on the markup because these ids are contracts with the JavaScript.

    `tests/test_template_render_targets.py` catches a getElementById with no element; this catches
    the other direction — an element the design calls for that was never added.
    """
    source = _template()
    for element in ("window-bar", "view-toggle", "filter-btn", "filter-area",
                    "rules-btn", "rules-panel", "search", "table-area"):
        assert f'id="{element}"' in source, f"{element} is missing"


def test_the_template_uses_delegated_listeners_not_inline_handlers():
    """**Product names come from an uploaded sheet and SKUs from Amazon.**

    Building an `onclick` out of either is an injection waiting to happen. The Orders tab already
    had to be fixed for exactly this, so the rule is asserted rather than remembered.
    """
    source = _template()
    for forbidden in ("onclick=\"save", "onclick=\"tick", "onclick='", "onchange=\"save"):
        assert forbidden not in source, f"inline handler {forbidden} builds a handler from data"
    # And the controls ARE wired up.
    for wired in ('$("table-area").addEventListener', '$("window-bar").addEventListener',
                  '$("view-toggle").addEventListener', '$("filter-area").addEventListener',
                  '$("rules-panel").addEventListener'):
        assert wired in source, f"{wired} is missing, so that control does nothing"


def test_the_template_shows_tacos_and_acos_as_separate_columns():
    """Both, always. They answer different questions and neither replaces the other.

    Measured: TACOS 33.1% against ACOS 89.9% on the same spend — collapsing them into one number
    would lose the fact that the advertising barely breaks even.
    """
    source = _template()
    assert '"tacos"' in source and '"acos"' in source
    assert "ad spend over TOTAL sales" in source, "the TACOS column has no explanation"
    assert "ad spend over ad-ATTRIBUTED sales" in source, "the ACOS column has no explanation"


def test_the_template_never_renders_a_zero_acos_for_an_unadvertised_product():
    """A dash, and a distinct label for spend-with-no-sales.

    Three states, three renderings: no ads at all, ads with no attributed sales, and a real ratio.
    Rendering the first as 0% would make it the most efficient product in the portfolio.
    """
    source = _template()
    assert "function acosCell(" in source
    start = source.index("function acosCell(")
    body = source[start:start + 600]
    assert "acos_infinite" in body, "the spend-with-no-sales case is not distinguished"
    assert "—" in body, "an unadvertised product does not render a dash"


def test_the_window_picker_cannot_offer_today():
    """Today's data is partial, so the date inputs are capped at yesterday.

    Enforced server-side too (a 400), but capping the input stops the owner picking something that
    will be refused — a control that offers an invalid choice is a bad control.
    """
    source = _template()
    assert "function maxDate(" in source
    start = source.index("function maxDate(")
    assert "getDate() - 1" in source[start:start + 300], (
        "maxDate does not subtract a day, so the picker would offer today"
    )
    assert 'max="${esc(maxDate())}"' in source, "the date inputs are not capped"


def test_the_view_and_sort_survive_a_reload():
    """A saved decision re-renders the table; losing the sort would throw the owner to the top.

    sessionStorage rather than localStorage: a filter left over from last week is a confusing way
    to open a dashboard, but losing one mid-session is worse.
    """
    source = _template()
    assert "sessionStorage" in source
    assert 'remembered("view"' in source and 'remembered("sort"' in source
    assert 'remembered("filters"' in source and 'remembered("window"' in source
    assert "localStorage" not in source, (
        "localStorage would carry a filter across sessions, so the tab would open filtered"
    )


# ─── Regression: ISSUE-001 — an empty filter box hid 45 of 90 products ────────
#
# Found by /qa on 2026-08-28.
# Report: .gstack/qa-reports/qa-report-portfolio-2026-08-28.md


def test_an_empty_filter_value_is_not_treated_as_a_zero_threshold():
    """**`Number("")` is 0, not NaN — which made a blank filter box a live `> 0` comparison.**

    A new filter row is SEEDED at `TACOS > 50` (a useful starting point, and it filters straight
    away — verified live: 13 of 90). The bug bit on the next gesture: the owner clears that 50 to
    type his own number, and for those keystrokes the value is "". The old `filterThreshold` ran
    `Number("")` → 0, passed `isFinite`, and returned 0 — so an empty box became `TACOS > 0` and
    silently dropped every product with no TACOS at all. Measured on real data: 45 of 90 products
    vanished mid-edit, with the count label calling it a match.

    Now clearing the box shows all 90 while the owner types, which is the only honest reading of
    "no threshold entered".

    Asserted on the template because the guard is one line of JavaScript and the symptom is
    invisible from Python: the payload was correct, the rendering was correct, and the filter
    was quietly doing exactly what it was told.
    """
    source = _template()
    start = source.index("function filterThreshold(")
    body = source[start:start + 700]
    assert 'if(text === "") return null;' in body, (
        "filterThreshold does not short-circuit on an empty value, so a blank filter box "
        "becomes a `> 0` comparison and silently hides every row with no value for that field"
    )
    # The trim matters too: a box containing only spaces is just as empty.
    assert ".trim()" in body, "an all-whitespace filter value would still coerce to 0"


# ─── Regression: ISSUE-003 — the active verdict chip vanished on a grain switch ─
#
# Found by /qa on 2026-08-28.


def test_the_active_verdict_chip_survives_a_grain_with_no_rows():
    """**Switching Products → SKUs while SURGICAL was selected left an empty table and no chip.**

    The chips are counted for the grain on screen, which is right — "9 KILL products" and "24 KILL
    SKUs" are both true. But the loop skipped any verdict with a zero count, and SURGICAL is
    *structurally* zero for SKUs: it compares a parent against its own sizes, so no single size
    can carry it. The filter stayed applied with its chip gone, so the screen read as broken
    rather than as filtered, and there was no control left to click to undo it.
    """
    source = _template()
    start = source.index("function renderVerdicts(")
    body = source[start:start + 1400]
    assert "filter !== v" in body, (
        "renderVerdicts drops a zero-count verdict unconditionally, so selecting a parent-only "
        "verdict and switching to SKUs removes the only control that could undo the filter"
    )


def test_an_empty_table_names_the_control_that_emptied_it():
    """Three controls can empty this grid, and "Nothing matches that filter" points at none.

    The verdict chip, the custom filters and the search box are three separate ways to reach zero
    rows, and the owner has to be able to tell which one did it. The SURGICAL-under-SKUs case gets
    a sentence of its own, because there the honest answer is not "no matches" but "this verdict
    cannot describe a single size".
    """
    source = _template()
    start = source.index("function renderTable(")
    body = source[start:start + 1800]
    assert "custom filter(s)" in body, "the empty note does not mention the custom filters"
    assert "search" in body, "the empty note does not mention the search box"
    assert 'filter === "SURGICAL" && isSkus' in body, (
        "the one structurally-impossible combination is not explained, so it reads as a dead end"
    )


# ─── Regression: ISSUE-004 — the table dragged the whole page sideways ─────────
#
# Found by /qa on 2026-08-28. Measured: at a 350px viewport the document scrolled 744px.


def test_the_table_scrolls_inside_a_wrapper_rather_than_moving_the_page():
    """**Every cell here is `white-space:nowrap`, so the table is legitimately wider than a phone.**

    The question is where the overflow goes. With no wrapper it went to the PAGE: measured at a
    350px viewport the document scrolled 744px sideways, which takes the nav, the window bar and
    the verdict chips off-screen and leaves no anchored column to scroll back to.

    `min-width` on the table is half the fix and easy to omit: without it the table obeys
    `width:100%`, squeezes to the container, and the nowrap cells overflow their own gridlines
    instead — the same defect as the PDF cells that printed SKUs over product names.
    """
    source = _template()
    assert ".table-wrap{" in source, "no scroll container is declared"
    assert "overflow-x:auto" in source
    assert '<div class="table-wrap">' in source, (
        "the wrapper is styled but the table is not inside it"
    )
    table_rule = source[source.index("table{width:100%"):]
    table_rule = table_rule[:table_rule.index("}")]
    assert "min-width" in table_rule, (
        "without a min-width the table shrinks to the wrapper and the nowrap cells overflow "
        "their own gridlines, so the wrapper never scrolls and the fix does nothing"
    )


# ─── Regression: ISSUE-005 — the sortable headers were mouse-only ─────────────
#
# Found by /qa on 2026-08-28.


def test_every_sortable_header_is_reachable_from_the_keyboard():
    """A `th` with a click listener and no tabindex is a control only a mouse can use.

    `role="button"` promises Enter and Space, so both are handled, and Space is
    preventDefault'ed — its default action scrolls the page, so without that a keyboard user
    would sort the grid and jump a screenful away from it in one keystroke.
    """
    source = _template()
    start = source.index("function headerHtml(")
    body = source[start:start + 1200]
    assert 'tabindex="0"' in body, "a sortable header is not in the tab order"
    assert 'role="button"' in body, "a sortable header does not announce itself as pressable"
    assert "aria-sort=" in body, (
        "nothing tells a screen reader which column is ordering the grid — the arrow glyph "
        "carries that visually and is aria-hidden"
    )
    assert 'scope="col"' in body, "the header cells are not associated with their columns"

    assert '$("table-area").addEventListener("keydown"' in source, "no key handler is bound"
    handler = source[source.index('$("table-area").addEventListener("keydown"'):]
    handler = handler[:600]
    assert '"Enter"' in handler and '" "' in handler, "Enter and Space are not both handled"
    assert "preventDefault" in handler, (
        "Space is not preventDefault'ed, so sorting also scrolls the page away from the table"
    )


def test_the_sort_rule_is_one_function_shared_by_mouse_and_keyboard():
    """Two copies of the toggle rule is how a keyboard user gets different ordering from a mouse.

    Also: `renderTable()` rebuilds the thead, so the focused element is destroyed on every sort.
    Without restoring focus the caret returns to the top of the document and tabbing to the next
    column starts over, which makes the keyboard path present but unusable.
    """
    source = _template()
    assert "function applySort(" in source
    body = source[source.index("function applySort("):]
    body = body[:900]
    assert "remember(" in body and "renderTable()" in body
    assert ".focus()" in body, (
        "focus is not restored after the thead is rebuilt, so every keyboard sort throws the "
        "caret back to the top of the page"
    )
    # And both entry points go through it rather than reimplementing the toggle.
    assert source.count("dir: -sort.dir") == 1, (
        "the sort toggle is written more than once, so the mouse and keyboard paths can drift"
    )


def test_the_window_controls_carry_accessible_names():
    """Two bare date inputs and a coloured dot that means "instant" versus "20 minutes".

    The dot is not decoration: a cached window loads immediately and an uncached one starts a
    ~20-minute Amazon report, so that distinction goes into the accessible name in words.
    """
    source = _template()
    start = source.index('id="win-from"')
    assert "aria-label=" in source[start:start + 200], "the start date input is unlabelled"
    start = source.index('id="win-to"')
    assert "aria-label=" in source[start:start + 220], "the end date input is unlabelled"
    presets = source[source.index("const buttons = [7, 30, 60, 90]"):]
    presets = presets[:900]
    assert "already fetched" in presets and "not fetched yet" in presets, (
        "the cached/uncached distinction is carried only by a coloured dot, which says nothing "
        "to a screen reader — and it is the difference between instant and ~20 minutes"
    )
    assert 'aria-hidden="true"' in presets, "the dot glyph is announced as well as its words"
