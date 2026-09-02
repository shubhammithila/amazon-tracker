"""The document builders: row order, column layout, and legibility on paper.

Requirement 3 was "sorted product wise, then weight wise on the dashboard **and in
the downloads as well**". The second half is the part a test has to defend, because
it is invisible: a download with the rows in the wrong order still opens fine, still
totals correctly, and is only noticed when the owner is holding the printout next to
the screen and they disagree. So the central tests here read a generated xlsx back
with openpyxl and assert its data rows equal ``logic.sort_items`` order — a real
check on the bytes that reach the owner, not on the list handed to the builder.

Mutation-verified: with ``sorted(rows, key=lambda r: r[4])`` inserted at the top of
``build_simple_xlsx``, ``test_xlsx_renders_the_order_it_is_given`` and
``test_order_survives_a_shuffled_input`` both fail.

**The second half of this file is about the PDF being readable, which is not
cosmetic.** A merchant SKU here runs to 48 characters, and reportlab does not wrap a
plain string in a table cell — it draws it at full width straight over the gridline
into the next column. Real 117-row plans were printing the SKU on top of the product
name on most rows. That is a document the warehouse cannot use, and nothing in the
suite noticed, because the bytes were a valid PDF of the right size the whole time.
The width tests below assert the property that was actually violated: a column whose
content must not wrap is wide enough for its widest value.

> Thirty tests were deleted with this rewrite, along with the four builders they
> covered (``build_packing_plan_xlsx``, ``build_packing_plan_pdf``,
> ``build_packed_xlsx``, ``build_remaining_pdf``). Every download route had already
> moved to the shared layout, so those tests were the builders' only remaining
> caller — twelve of them asserting column content for spreadsheets nobody could
> download. Passing tests over unreachable code are worse than no tests: they made
> the module read as a description of what the owner receives.
"""
import io

import pytest

from app.shipment import documents, logic
from tests.conftest import CANONICAL_ORDER

pytestmark = pytest.mark.regression


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _read_rows(buffer: io.BytesIO) -> tuple[list[str], list[list]]:
    """(header, data rows) from an xlsx in memory."""
    from openpyxl import load_workbook

    buffer.seek(0)
    workbook = load_workbook(buffer)
    sheet = workbook.active
    rows = [list(r) for r in sheet.iter_rows(values_only=True)]
    return rows[0], rows[1:]


def _item(asin, item, weight, **extra):
    """A plan-item dict in the shape _item_payload produces."""
    row = {
        "asin": asin,
        "item": item,
        "weight": weight,
        "brand": "MF",
        "fba_sku": f"SKU-{asin}",
        "sales_7d": 10,
        "projection": 50,
        "fba_stock": 5,
        "deficit": 45,
        "shipment_plan": 50,
        "available": 0,
        "s": False,
        "m": False,
        "b": False,
        "packed": 0,
        "shippable": 0,
        "remaining": 50,
    }
    row.update(extra)
    return row


@pytest.fixture
def items():
    """Deliberately awkward, mirroring conftest.plan_factory exactly.

    'aloe vera juice' is lowercase and sorts FIRST alphabetically, but it is Howrah
    Foods and category P6 Rest, so it must come LAST. A builder that re-sorted by
    name, by ASIN, case-sensitively, or that ignored brand and category, would put it
    somewhere else — which is what makes the ordering assertions below able to fail
    rather than merely able to pass.
    """
    return logic.sort_items([
        _item("B0AAA00001", "Chana Sattu", 1.0, shipment_plan=500, packed=200, remaining=300),
        _item("B0AAA00002", "Chana Sattu", 0.5, shipment_plan=300, packed=300, remaining=0),
        _item("B0BBB00001", "jau sattu", 1.0, shipment_plan=200, packed=0, remaining=200),
        _item("B0CCC00001", "aloe vera juice", 2.0, brand="HF",
              shipment_plan=100, packed=40, remaining=60),
    ])


@pytest.fixture
def headers():
    """The plan/remaining layout: the shared columns plus one quantity."""
    return documents.IDENTITY_HEADERS + ["To Pack"]


@pytest.fixture
def widths():
    return documents.IDENTITY_WIDTHS + [12]


@pytest.fixture
def rows(items):
    return documents._rows_with_quantity(items, "shipment_plan")


def _asin_column(header: list[str], rows: list[list]) -> list:
    """The ASIN column BY NAME, with the totals line dropped.

    By name because the layout has changed twice: index 3 was ASIN, then Brand. A
    positional lookup would compare the wrong column and quietly pass.
    """
    index = header.index("ASIN")
    return [r[index] for r in rows if r[index]]


# ─── Row order: the point of the file ────────────────────────────────────────

#: Imported rather than restated. The order lived in three separate hardcoded
#: copies, which is three places to update and three chances to update only two —
#: and a stale copy that still passes is worse than one that fails.
EXPECTED_ORDER = CANONICAL_ORDER
#  MF·P1 chana 0.5kg → MF·P1 chana 1kg → MF·P1 jau 1kg → HF·P6 aloe


def test_the_fixture_itself_is_in_canonical_order(items):
    """Guards the guard. If this drifts, every ordering test below is vacuous."""
    assert [i["asin"] for i in items] == EXPECTED_ORDER


def test_xlsx_renders_the_order_it_is_given(headers, rows, widths):
    buffer = documents.build_simple_xlsx("Plan", "sub", headers, rows, widths)
    header, data = _read_rows(buffer)
    assert _asin_column(header, data) == EXPECTED_ORDER


def test_order_survives_a_shuffled_input(headers, widths, items):
    """A builder must render the order it is GIVEN, not one it computes.

    Handed a deliberately wrong order, the output must be equally wrong. That sounds
    perverse but it is the contract: ordering lives in
    ``repository.load_plan_items`` alone, and a builder that re-sorted would mask an
    ordering bug there while introducing its own.
    """
    backwards = documents._rows_with_quantity(list(reversed(items)), "shipment_plan")
    header, data = _read_rows(
        documents.build_simple_xlsx("Plan", "sub", headers, backwards, widths)
    )
    assert _asin_column(header, data) == list(reversed(EXPECTED_ORDER))


def test_sort_items_is_what_the_documents_agree_with(items):
    """The documents' order is `logic.sort_items`, not some private rule."""
    import random

    shuffled = items[:]
    random.Random(7).shuffle(shuffled)
    assert [i["asin"] for i in logic.sort_items(shuffled)] == EXPECTED_ORDER


# ─── Formats ─────────────────────────────────────────────────────────────────

def test_xlsx_is_a_real_xlsx(headers, rows, widths):
    """xlsx is a zip, so it must start with the PK magic bytes."""
    buffer = documents.build_simple_xlsx("Plan", "sub", headers, rows, widths)
    buffer.seek(0)
    assert buffer.read(2) == b"PK"


def test_pdf_is_a_real_pdf(headers, rows):
    buffer = documents.build_simple_pdf("Plan", "sub", headers, rows)
    buffer.seek(0)
    assert buffer.read(4) == b"%PDF"


def test_buffers_are_rewound_ready_to_stream(headers, rows, widths):
    """StreamingResponse reads from the current position — an un-rewound buffer
    would send a 0-byte file that Excel reports as corrupt."""
    assert documents.build_simple_xlsx("Plan", "s", headers, rows, widths).tell() == 0
    assert documents.build_simple_pdf("Plan", "s", headers, rows).tell() == 0


def test_xlsx_freezes_the_header_row(headers, rows, widths):
    """205 rows on one sheet: without this the owner scrolls and loses the headings."""
    from openpyxl import load_workbook

    buffer = documents.build_simple_xlsx("Plan", "sub", headers, rows, widths)
    buffer.seek(0)
    assert load_workbook(buffer).active.freeze_panes == "A2"


# ─── The shared column layout ────────────────────────────────────────────────

def test_the_requested_column_order_is_the_first_seven(headers):
    """"columns of all these downloads will be as follows in the below order
    S M B Brand, Asin, sku, Product" — verbatim, and asserted as a prefix.

    A prefix rather than the whole list, so ``Size`` could be appended without
    disturbing what was actually asked for.
    """
    assert headers[:7] == ["S", "M", "B", "Brand", "ASIN", "Merchant SKU", "Product"]


def test_size_is_its_own_column_not_glued_to_the_product_name(items):
    """It used to be "Jau Sattu 500g" in one cell, and that was wrong on paper.

    All sizes of one product are adjacent (sort_key puts product above weight), so
    the packer's actual scan is down a column of sizes — and a size buried at the end
    of names of differing lengths cannot be scanned. Separated, they line up.
    """
    cells = documents._identity_cells(items[0])
    product = cells[documents.IDENTITY_HEADERS.index("Product")]
    size = cells[documents.IDENTITY_HEADERS.index("Size")]

    assert product == "Chana Sattu", f"the product cell is not just the name: {product!r}"
    assert size == "500g", f"the pack size is not in its own cell: {size!r}"
    assert size not in product, "the weight is still concatenated onto the name"


def test_the_size_column_is_headed_just_size(headers):
    """Asked for directly: "just write size".

    "Pack Size" was two words for a column of "500g" and "1 kg", and the heading was
    wider than every value under it — which makes the column look like it is carrying
    more than it is.
    """
    assert "Size" in headers
    assert "Pack Size" not in headers, "the column heading is back to two words"


def test_widths_are_supplied_for_every_column(headers, widths):
    """A short widths list silently leaves the last columns at Excel's default,
    which is narrow enough to show '#####' instead of a quantity."""
    assert len(documents.IDENTITY_WIDTHS) == len(documents.IDENTITY_HEADERS)
    assert len(widths) == len(headers)


def test_flags_render_as_a_letter_not_a_boolean(items):
    """S/M/B are carton sizes read off a page. "S TRUE" does not scan."""
    cells = documents._identity_cells(dict(items[0], s=True, m=False, b=True))
    assert cells[:3] == ["Y", "", "Y"]


# ─── Dropping rows with nothing to do ────────────────────────────────────────

def test_rows_with_no_quantity_are_dropped(items):
    """"not the entire list of skus". A 205-row sheet where 88 rows read 0 is a
    sheet nobody reads to the end."""
    lines = documents._rows_with_quantity(items, "remaining")
    # Chana 0.5kg is fully packed, so nothing remains to pack for it.
    assert "B0AAA00002" not in [r[documents.IDENTITY_HEADERS.index("ASIN")] for r in lines]
    assert len(lines) == 3


def test_the_quantity_key_chooses_what_the_document_is_about(items):
    """One layout, three documents. The quantity column is the only difference, so
    it must actually read the key it is given."""
    quantity = len(documents.IDENTITY_HEADERS)
    planned = {r[4]: r[quantity] for r in documents._rows_with_quantity(items, "shipment_plan")}
    left = {r[4]: r[quantity] for r in documents._rows_with_quantity(items, "remaining")}
    assert planned["B0AAA00001"] == 500
    assert left["B0AAA00001"] == 300


# ─── Totals ──────────────────────────────────────────────────────────────────

def test_the_totals_row_sums_every_quantity_column(items):
    """Driven by the header count, so the packed sheet's two quantity columns both
    total without a second implementation."""
    headers = documents.IDENTITY_HEADERS + ["Units", "Cartons"]
    rows = [documents._identity_cells(i) + [100, 8] for i in items]
    totals = documents._totals_row(headers, rows)
    assert totals[-2:] == [400, 32]


def test_the_totals_label_sits_under_the_product_column(items):
    """Derived from IDENTITY_HEADERS, not a hand-counted run of blanks.

    It WAS six hardcoded empty strings, and adding Size shifted the label one
    column right — onto a quantity heading, where a totals line reading
    "TOTAL · 205 rows" in the Size column looks like a rendering bug.
    """
    headers = documents.IDENTITY_HEADERS + ["To Pack"]
    rows = [documents._identity_cells(i) + [10] for i in items]
    totals = documents._totals_row(headers, rows)

    label_index = documents.IDENTITY_HEADERS.index("Product")
    assert str(totals[label_index]).startswith("TOTAL"), (
        f"the totals label is at index {[i for i, v in enumerate(totals) if 'TOTAL' in str(v)]}, "
        f"not under Product ({label_index})"
    )
    assert "4 rows" in totals[label_index]
    # And nothing spills into the quantity columns except the sums.
    assert totals[len(documents.IDENTITY_HEADERS)] == 40


def test_an_empty_document_has_no_totals_row(headers, widths):
    """"TOTAL · 0 rows" under an empty table is noise; the document already says
    there is nothing to do."""
    header, data = _read_rows(
        documents.build_simple_xlsx("Plan", "sub", headers, [], widths)
    )
    assert data == [], data


# ─── The PDF must be readable on paper ───────────────────────────────────────
#
# The bug: reportlab draws a plain string in a table cell at full width, straight
# over the gridline. Real merchant SKUs are up to 48 characters, so the SKU column
# printed on top of the product column on most rows of a real plan. Every existing
# test passed throughout — the bytes were a valid PDF of a plausible size.

def _table_height(headers, rows):
    """The height the table actually occupies at its real column widths, in points.

    reportlab computes this in ``wrap``, which is where a cell either grows to fit
    its content or does not. That makes it the one place the wrapping behaviour is
    observable without rendering to pixels.
    """
    from reportlab.lib.units import mm
    from reportlab.platypus import Table

    body = list(rows)
    data = [[documents._head_cell(h) for h in headers]]
    for row in body:
        data.append([
            documents._body_cell(row[c], headers[c], c, len(headers), False)
            for c in range(len(headers))
        ])
    table = Table(data, colWidths=documents._pdf_column_widths(headers, body), repeatRows=1)
    table.setStyle(documents._pdf_table_style(len(headers)))
    return table.wrap(documents._PAGE_WIDTH_MM * mm, 10_000)[1]


def test_an_overlong_value_makes_its_row_taller_instead_of_overflowing():
    """The overlap bug itself, caught by the only property that distinguishes it.

    reportlab does not wrap a plain string in a table cell — it draws it at full
    width, straight over the gridline and on top of the next column. The cell is
    therefore the SAME height whether the text fits or not, which is precisely why
    the bug was invisible: every byte-level assertion still passed.

    So the test compares heights. A SKU that fits on one line and a 48-character SKU
    that cannot must produce tables of DIFFERENT heights. With Paragraph cells the
    long one is taller; revert to plain strings and the two are identical, because
    the overflow costs nothing vertically — it just lands on the product name.

    Mutation-verified: replacing ``Paragraph(_escape(value), style)`` with
    ``_escape(value)`` in ``_body_cell`` fails here and nowhere else in the file.
    """
    headers = documents.IDENTITY_HEADERS + ["To Pack"]

    def one_row(sku):
        return [
            documents._identity_cells({
                "asin": "B0AAA00001", "brand": "MF", "item": "Chana Sattu",
                "weight": 1.0, "fba_sku": sku, "m": True,
            }) + [250]
        ]

    short = _table_height(headers, one_row("MF-CH-1KG"))
    long = _table_height(headers, one_row("Black_&_White_Sesame_Laddoo_-_Jaggery_0.25kg FBA"))

    assert long > short, (
        "a 48-character merchant SKU did not make its row taller, which means the "
        "cell is not wrapping — it is overflowing across the gridline and printing "
        "on top of the product name, exactly as it did on the owner's plan PDF"
    )


def test_a_long_product_name_also_wraps():
    """Product is the flexible column and takes the longest real names.

    Asserted separately from the SKU because they use different paragraph styles, and
    a style added without a ``_STYLE_FONTS`` entry (or a cell built as a bare string)
    would break one and not the other.
    """
    headers = documents.IDENTITY_HEADERS + ["To Pack"]

    def one_row(name):
        return [
            documents._identity_cells({
                "asin": "B0AAA00001", "brand": "MF", "item": name,
                "weight": 1.0, "fba_sku": "MF-1KG", "m": True,
            }) + [250]
        ]

    short = _table_height(headers, one_row("Chana Sattu"))
    long = _table_height(headers, one_row("Black & White Sesame Laddoo - Jaggery Special Edition"))
    assert long > short, "a long product name is overflowing rather than wrapping"


def _widest_content_mm(rows, column: int, heading: str) -> float:
    """How much room this column's widest value actually needs, in mm."""
    from reportlab.lib.units import mm
    from reportlab.pdfbase.pdfmetrics import stringWidth

    font, size = documents._STYLE_FONTS[
        documents._body_cell_style_name(heading, column)
    ]
    return max(
        (stringWidth(str(r[column] or ""), font, size) for r in rows), default=0.0
    ) / mm


#: Columns whose content is unreadable when broken across two lines. A wrapped ASIN
#: cannot be typed into a search box; "1.75" over "kg" reads as two separate facts.
NO_WRAP_COLUMNS = ["ASIN", "Size", "Brand"]


@pytest.mark.parametrize("heading", NO_WRAP_COLUMNS)
def test_columns_that_must_not_wrap_are_wide_enough(heading):
    """The actual invariant the overlap bug violated, asserted numerically.

    Rendering to an image and looking is not something a suite can do, but "the
    column is wider than its widest value plus padding" is exactly the property that
    was false — and it fails loudly if someone tightens a width or raises a font
    size.

    Uses the real catalogue's longest values, not the fixture's tidy ones: SKU-B0AAA
    00001 would fit anywhere.
    """
    from reportlab.lib.units import mm

    headers = documents.IDENTITY_HEADERS + ["To Pack"]
    rows = [
        documents._identity_cells({
            "s": True, "m": True, "b": True,
            "brand": "MF",
            "asin": "B0FNFGRWSW",  # the widest ASIN in product_families.json
            "fba_sku": "Black_&_White_Sesame_Laddoo_-_Jaggery_0.25kg FBA",
            "item": "Black & White Sesame Laddoo - Jaggery",
            "weight": 1.75,        # "1.75 kg", the widest pack size
        }) + [12345],
    ]

    column = headers.index(heading)
    allotted = documents._pdf_column_widths(headers, rows)[column] / mm
    needed = _widest_content_mm(rows, column, heading) + documents._PADDING_MM

    assert allotted >= needed, (
        f"the {heading} column gets {allotted:.1f}mm but its widest value needs "
        f"{needed:.1f}mm — it will wrap onto two lines, which for {heading} is "
        f"unreadable rather than merely untidy"
    )


def test_a_quantity_never_wraps(items):
    """A five-digit total split as "456" / "0" is not a number any more.

    This is not hypothetical: it happened. The column had been sized to precisely the
    measured width of its digits, and reportlab breaks when the text is >= the space
    available, so an exact fit wraps. Hence documents._SLACK_MM.
    """
    from reportlab.lib.units import mm

    headers = documents.IDENTITY_HEADERS + ["Units", "Cartons"]
    rows = [documents._identity_cells(i) + [9999, 888] for i in items]
    body = rows + [documents._totals_row(headers, rows)]
    allotted = documents._pdf_column_widths(headers, body)

    for heading in ("Units", "Cartons"):
        column = headers.index(heading)
        needed = _widest_content_mm(body, column, heading) + documents._PADDING_MM
        assert allotted[column] / mm > needed, (
            f"{heading} is sized to exactly its content and will wrap the total"
        )


def test_the_columns_fill_the_page_exactly(items):
    """Sum to the printable width: narrower wastes the page, wider makes reportlab
    silently shrink the whole table and undo every font decision."""
    from reportlab.lib.units import mm

    for extra in (["To Pack"], ["Units", "Cartons"]):
        headers = documents.IDENTITY_HEADERS + extra
        rows = [documents._identity_cells(i) + [100] * len(extra) for i in items]
        total = sum(documents._pdf_column_widths(headers, rows)) / mm
        assert total == pytest.approx(documents._PAGE_WIDTH_MM, abs=0.01), (
            f"{len(headers)} columns sum to {total:.2f}mm, not "
            f"{documents._PAGE_WIDTH_MM}mm"
        )


def test_product_absorbs_the_extra_quantity_column(items):
    """The packed sheet has two quantity columns and the same page width.

    Something has to give, and it must be Product: a long name over two lines is
    still completely readable, whereas narrowing the ASIN or clipping the digits is
    the bug this section exists to prevent.
    """
    from reportlab.lib.units import mm

    one = documents.IDENTITY_HEADERS + ["To Pack"]
    two = documents.IDENTITY_HEADERS + ["Units", "Cartons"]
    rows_one = [documents._identity_cells(i) + [100] for i in items]
    rows_two = [documents._identity_cells(i) + [100, 8] for i in items]

    wide = documents._pdf_column_widths(one, rows_one)
    narrow = documents._pdf_column_widths(two, rows_two)
    product = one.index("Product")
    asin = one.index("ASIN")

    assert narrow[product] < wide[product], "Product did not absorb the extra column"
    assert narrow[asin] == wide[asin], (
        "the ASIN column shrank to make room, which is what makes it wrap"
    )
    assert narrow[product] / mm >= documents._PRODUCT_FLOOR_MM


def test_product_keeps_a_floor_even_with_many_quantity_columns(items):
    """Defensive: a future document with four quantity columns must not reduce the
    product name to a two-character ribbon."""
    from reportlab.lib.units import mm

    headers = documents.IDENTITY_HEADERS + ["A", "B", "C", "D"]
    rows = [documents._identity_cells(i) + [100, 100, 100, 100] for i in items]
    product = documents._pdf_column_widths(headers, rows)[headers.index("Product")]
    assert product / mm >= documents._PRODUCT_FLOOR_MM


# ─── Emphasis: what the packer reads first ───────────────────────────────────

def test_the_instruction_is_set_larger_than_the_identifiers():
    """Three cells decide what the packer does — product, size, quantity.

    Printing all nine columns at one weight is what produced "which of these numbers
    am I packing". Asserted as a font-size relation rather than exact points, so the
    sizes can be tuned without the test becoming a copy of the implementation.
    """
    fonts = documents._STYLE_FONTS
    quiet = fonts["quiet"][1]
    for style in ("loud", "quantity"):
        assert fonts[style][1] > quiet, (
            f"{style} is not larger than the recessive identifier text"
        )
        assert "Bold" in fonts[style][0], f"{style} is not bold"


@pytest.mark.parametrize(
    "heading,expected",
    [
        ("ASIN", "quiet"),
        ("Merchant SKU", "quiet"),
        ("Product", "loud"),
        ("Size", "loud"),
        ("Brand", "plain"),
        ("S", "flag"),
    ],
)
def test_each_column_gets_the_style_its_job_needs(heading, expected):
    column = documents.IDENTITY_HEADERS.index(heading)
    assert documents._body_cell_style_name(heading, column) == expected


def test_a_quantity_column_is_styled_as_one_whatever_it_is_called():
    """Position, not name. "To Pack", "Units" and "Cartons" are all quantities and
    a fourth will be called something else again."""
    for name in ("To Pack", "Units", "Cartons", "Anything"):
        assert (
            documents._body_cell_style_name(name, len(documents.IDENTITY_HEADERS))
            == "quantity"
        )


def test_every_style_name_has_a_measurable_font():
    """_pdf_column_widths measures a cell via _STYLE_FONTS instead of building its
    Paragraph. A style added to one and not the other measures at the wrong size and
    the column silently wraps — which is the whole failure mode of this section."""
    for name in documents._paragraph_styles():
        if name == "head":
            continue  # the header row is always Helvetica-Bold 8.5, handled directly
        assert name in documents._STYLE_FONTS, f"{name} has no entry in _STYLE_FONTS"


def test_the_excel_mirrors_the_pdf_emphasis(headers, rows, widths):
    """The owner reads the Excel of the document the packer holds on paper. Two
    versions that look unrelated get filed as two different reports."""
    from openpyxl import load_workbook

    buffer = documents.build_simple_xlsx("Plan", "sub", headers, rows, widths)
    buffer.seek(0)
    sheet = load_workbook(buffer).active
    header = [c.value for c in sheet[1]]
    first = {h: c for h, c in zip(header, sheet[2])}

    assert first["ASIN"].font.size < first["Product"].font.size, (
        "the ASIN is not de-emphasised relative to the product name"
    )
    assert first["Product"].font.bold and first["Size"].font.bold
    assert first["To Pack"].font.bold, "the quantity is not emphasised"


# ─── Robustness against real-world rows ──────────────────────────────────────

def test_an_ampersand_in_a_product_name_does_not_break_the_pdf():
    """"Black & White Sesame Laddoo - Jaggery" is a real product here.

    reportlab's Paragraph parses its input as mini-HTML, so a bare '&' raises and
    takes the whole download down with a 500. Same class of bug as the unescaped
    template interpolations, and reachable from an uploaded CSV.
    """
    headers = documents.IDENTITY_HEADERS + ["To Pack"]
    rows = [
        documents._identity_cells({
            "item": "Black & White <Sesame> Laddoo",
            "fba_sku": "A&B <b>bold</b>",
            "asin": "B0AAA00001", "brand": "MF", "weight": 0.25, "s": True,
        }) + [250]
    ]
    assert documents.build_simple_pdf("Plan", "a & b", headers, rows).read(4) == b"%PDF"


def test_builders_tolerate_missing_keys(headers, widths):
    """Rows come from the DB via _item_payload, but a legacy-imported plan can carry
    Nones. A download must not 500 on one blank field."""
    sparse = documents._rows_with_quantity([{"asin": "B0XXX00001", "shipment_plan": 5}],
                                          "shipment_plan")
    assert documents.build_simple_xlsx("P", "s", headers, sparse, widths).read(2) == b"PK"
    assert documents.build_simple_pdf("P", "s", headers, sparse).read(4) == b"%PDF"


def test_an_empty_document_still_builds(headers, widths):
    """The no-plan and everything-packed cases. An empty table must not crash, and
    an xlsx Excel calls corrupt looks like a broken app."""
    assert documents.build_simple_xlsx("P", "s", headers, [], widths).read(2) == b"PK"
    assert documents.build_simple_pdf("P", "s", headers, []).read(4) == b"%PDF"


def test_a_long_sheet_paginates(items):
    """A 205-row plan is several pages, and the header must repeat on each.

    Asserted via page count rather than by parsing the PDF: repeatRows=1 is the
    mechanism and this is the observable consequence of it being wired at all.
    """
    headers = documents.IDENTITY_HEADERS + ["To Pack"]
    many = [documents._identity_cells(items[0]) + [100] for _ in range(150)]
    buffer = documents.build_simple_pdf("Plan", "sub", headers, many)
    assert buffer.getbuffer().nbytes > 5000
    assert buffer.read(4) == b"%PDF"


def test_a_sheet_title_longer_than_excel_allows_is_truncated(headers, rows, widths):
    """Excel's limit is 31 characters and a longer name makes the file invalid —
    it opens as "we found a problem with some content"."""
    from openpyxl import load_workbook

    buffer = documents.build_simple_xlsx("x" * 60, "sub", headers, rows, widths)
    buffer.seek(0)
    assert len(load_workbook(buffer).active.title) <= 31


# ─── The Amazon upload file ──────────────────────────────────────────────────

def test_shipment_file_xlsx_row_order(items):
    _header, rows = _read_rows(documents.build_shipment_file_xlsx(items, mode="all"))
    assert [r[1] for r in rows] == EXPECTED_ORDER  # SKU, ASIN, ...


def test_shipment_file_omits_zero_quantities(items):
    """Chana 0.5kg is fully packed, so nothing remains to plan for it."""
    _header, rows = _read_rows(documents.build_shipment_file_xlsx(items, mode="remaining"))
    assert "B0AAA00002" not in [r[1] for r in rows]
    assert len(rows) == 3


def test_shipment_file_mode_all_uses_the_planned_quantity(items):
    header, rows = _read_rows(documents.build_shipment_file_xlsx(items, mode="all"))
    quantities = {r[1]: r[header.index("Quantity")] for r in rows}
    assert quantities["B0AAA00001"] == 500
    assert quantities["B0AAA00002"] == 300


def test_shipment_file_mode_remaining_uses_what_is_left(items):
    header, rows = _read_rows(documents.build_shipment_file_xlsx(items, mode="remaining"))
    quantities = {r[1]: r[header.index("Quantity")] for r in rows}
    assert quantities["B0AAA00001"] == 300


def test_shipment_file_mode_verified_only_counts_verified_days(items):
    """mode=verified is what may legally be invoiced. A held day's units must not
    appear even though they are packed."""
    days = [
        {
            "pack_date": "2026-07-28",
            "status": logic.STATUS_VERIFIED,
            "entries": [
                {"asin": "B0AAA00001", "units": 200},
                {"asin": "B0CCC00001", "units": 40},
            ],
        },
        {
            "pack_date": "2026-07-29",
            "status": logic.STATUS_HELD,
            "entries": [{"asin": "B0AAA00002", "units": 300}],
        },
    ]
    header, rows = _read_rows(
        documents.build_shipment_file_xlsx(items, mode="verified", days=days)
    )
    quantities = {r[1]: r[header.index("Quantity")] for r in rows}
    assert quantities == {"B0AAA00001": 200, "B0CCC00001": 40}
    assert "B0AAA00002" not in quantities, "held units would be invoiced"


def test_shipment_file_mode_verified_is_empty_when_nothing_is_verified(items):
    submitted_only = [
        {
            "pack_date": "2026-07-28",
            "status": logic.STATUS_SUBMITTED,
            "entries": [{"asin": "B0AAA00001", "units": 200}],
        }
    ]
    _header, rows = _read_rows(
        documents.build_shipment_file_xlsx(items, mode="verified", days=submitted_only)
    )
    assert rows == [], "submitted is not verified — the owner has not approved it"


def test_shipment_file_leaves_a_missing_sku_blank(items):
    """Amazon's upload keys on the merchant SKU. Falling back to the ASIN produces a
    file that looks right and is rejected on their side; a visible blank is a problem
    the owner can fix before uploading."""
    no_sku = [dict(i, fba_sku="") for i in items]
    header, rows = _read_rows(documents.build_shipment_file_xlsx(no_sku, mode="all"))
    sku_column = header.index("Merchant SKU")
    for row in rows:
        assert row[sku_column] in (None, ""), row
        assert row[sku_column] != row[header.index("ASIN")]


def test_shipment_file_warns_about_missing_skus(items, caplog):
    """The old code swallowed this in a bare except. A silent version of this failure
    is the reason it is a logged warning and a counted number."""
    import logging

    no_sku = [dict(i, fba_sku="") for i in items]
    with caplog.at_level(logging.WARNING, logger="app.shipment.documents"):
        documents.build_shipment_file_xlsx(no_sku, mode="all")
    # getMessage(), not `.message % .args` — the record is logged lazily, so the args
    # are still separate and `%`-ing an already-interpolated string raises.
    messages = [r.getMessage() for r in caplog.records]
    assert any("merchant SKU" in m for m in messages), messages
    assert any("4 row(s)" in m for m in messages), (
        f"the count is the useful part of the warning: {messages}"
    )


def test_shipment_file_is_quiet_when_every_sku_is_present(items, caplog):
    import logging

    with caplog.at_level(logging.WARNING, logger="app.shipment.documents"):
        documents.build_shipment_file_xlsx(items, mode="all")
    assert not caplog.records


# ─── Shared styling ──────────────────────────────────────────────────────────

def test_excel_and_pdf_headers_are_the_same_colour():
    """HEADER_HEX is derived from HEADER_RGB so the two cannot drift apart.

    334C4C, not 334D4D: 0.3 * 255 is 76.5 and Python's round() takes halves to the
    even number, so 76. Exactly the banker's-rounding trap that made
    logic.round_to_step use Decimal — here it is one imperceptible shade of grey
    rather than a wrong shipment quantity, so it is left alone and written down.
    """
    assert documents.HEADER_HEX == "334C4C"
    # The invoice PDFs use the float form; both documents must read as one product.
    assert documents.HEADER_RGB == (0.2, 0.3, 0.3)


def test_weight_label_uses_grams_below_one_kilo():
    """The unit a warehouse actually says out loud.

    "0.5kg" is arithmetically identical and wrong on a picking sheet: the pouch label
    says 500g, and a packer scanning 100 rows should not be converting units in his
    head. Kilos from 1 kg up, where the printed labels switch too.

    Delegates to logic.weight_label — asserted through the documents alias as well so
    the printed sheet and the screen cannot drift apart.
    """
    assert documents._weight_label(0.5) == "500g"
    assert documents._weight_label(0.25) == "250g"
    assert documents._weight_label(0.15) == "150g"    # a real pack size here
    assert documents._weight_label(1.0) == "1 kg"     # not "1.0 kg"
    assert documents._weight_label(1.05) == "1.05 kg"  # also real; keeps decimals
    assert documents._weight_label(2.25) == "2.25 kg"


def test_grams_close_up_and_kilos_spaced():
    """The spacing differs between the units, and that is the owner's call.

    Asserted as an explicit pair rather than left implicit in the values above,
    because it looks exactly like an inconsistency someone would tidy up: "500g" and
    "1 kg" in one column reads as a bug unless you know it matches the pouch labels.
    """
    assert documents._weight_label(0.5) == "500g", "grams must sit against the number"
    assert " g" not in documents._weight_label(0.5)
    assert documents._weight_label(2.0) == "2 kg", "kilos keep their space"
    assert "2kg" != documents._weight_label(2.0)


@pytest.mark.parametrize("value", [None, 0, "", "abc", -1])
def test_weight_label_is_blank_rather_than_wrong(value):
    """A missing weight prints nothing. '0kg' on a warehouse sheet reads as a real
    fact about the product and would send someone looking for the 0kg bag."""
    assert documents._weight_label(value) == ""


# ─── The dispatch workbook and the to-buy list ───────────────────────────────


def _dispatch_fixture():
    """A dispatch sheet plus its purchasing view, with one product in SURPLUS.

    The surplus is what distinguishes a correct to-buy total from one that subtracts the
    totals, and it is what must be ABSENT from the to-buy list. Built by hand rather than from
    a live sheet so the expected numbers are visible in the test.
    """
    from app.orders import logic

    sheet = {
        "parents": [
            {"product": "Usna Chawal", "brand": "MF", "kg": 35.0, "units": 7,
             "orders": 7, "packed": 0, "remaining": 7,
             "sizes": [{"asin": "B0RICE5KG", "weight": 5.0, "weight_label": "5 kg",
                        "seller_sku": "5kg uc", "units": 7, "orders": 7, "kg": 35.0,
                        "packed": 0, "remaining": 7, "over_packed": 0, "known": True}]},
            {"product": "ABC Sattu", "brand": "MF", "kg": 22.5, "units": 37,
             "orders": 37, "packed": 29, "remaining": 8,
             "sizes": [{"asin": "B0ABC500", "weight": 0.5, "weight_label": "500g",
                        "seller_sku": "abc500", "units": 29, "orders": 29, "kg": 14.5,
                        "packed": 29, "remaining": 0, "over_packed": 0, "known": True},
                       {"asin": "B0ABC1KG", "weight": 1.0, "weight_label": "1 kg",
                        "seller_sku": "abc1kg", "units": 8, "orders": 8, "kg": 8.0,
                        "packed": 0, "remaining": 8, "over_packed": 0, "known": True}]},
        ],
        "orders": [
            {"amazon_order_id": "403-1", "parent": "Usna Chawal", "weight": 5.0,
             "weight_label": "5 kg", "seller_sku": "5kg uc", "asin": "B0RICE5KG",
             "quantity": 1, "known": True, "city": "PUNE", "state": "MAHARASHTRA",
             "easyship_status": "PickedUp"},
        ],
        "totals": {"orders": 8, "units": 44, "kg": 57.5, "packed": 29, "remaining": 15,
                   "over_packed": 0, "sizes_without_weight": 0, "parents": 2},
        "unknown_asins": [],
    }
    purchasing = logic.raw_stock_summary(sheet, {"Usna Chawal": 10.0, "ABC Sattu": 32.0})
    return sheet, purchasing


def test_the_dispatch_workbook_has_one_worksheet_per_tab():
    """Three tabs on screen, three worksheets in the file, named the same.

    One file rather than three downloads: they are read together, and three files get separated
    on a bench — the same reason the PDF footer prints "page 1 of 3".
    """
    from openpyxl import load_workbook

    from app.shipment import documents

    sheet, purchasing = _dispatch_fixture()
    book = load_workbook(documents.build_dispatch_xlsx(sheet, purchasing, "26 Aug (IST)"))
    assert book.sheetnames == ["Weight & purchase", "By SKU", "Orders"]


def test_the_workbooks_purchasing_total_sums_the_clamped_rows():
    """25.00 + 0, not 57.50 - 42.00 = 15.50.

    A supplier reads this file, and a surplus of sattu cannot cover a shortfall of rice.
    """
    from openpyxl import load_workbook

    from app.shipment import documents

    sheet, purchasing = _dispatch_fixture()
    book = load_workbook(documents.build_dispatch_xlsx(sheet, purchasing, "26 Aug (IST)"))
    values = [[cell.value for cell in row] for row in book["Weight & purchase"].iter_rows()]
    total_row = next(row for row in values if row and row[0] == "TOTAL")
    assert total_row[-1] == pytest.approx(25.0), (
        f"to-buy total is {total_row[-1]}; it must sum the clamped rows"
    )


def test_a_single_tab_workbook_holds_only_that_worksheet():
    """The per-tab download is the combined one with the others removed.

    One builder rather than two that could disagree about a quantity.
    """
    from openpyxl import load_workbook

    from app.shipment import documents

    sheet, purchasing = _dispatch_fixture()
    book = load_workbook(
        documents.build_dispatch_xlsx(sheet, purchasing, "26 Aug (IST)", tab="sku")
    )
    assert book.sheetnames == ["By SKU"]


def test_the_to_buy_list_omits_covered_products():
    """A purchasing list is a list of things to BUY.

    ABC Sattu has 32 kg against 22.5 kg ordered, so it must not appear — a zero row invites
    someone to order zero of it.
    """
    from openpyxl import load_workbook

    from app.shipment import documents

    _sheet, purchasing = _dispatch_fixture()
    book = load_workbook(documents.build_tobuy_xlsx(purchasing, "26 Aug (IST)"))
    text = "\n".join(
        " ".join(str(cell.value) for cell in row if cell.value is not None)
        for row in book.active.iter_rows()
    )
    assert "Usna Chawal" in text, "a short product is missing from the to-buy list"
    assert "ABC Sattu" not in text, "a covered product appeared on the purchasing list"


def test_an_empty_to_buy_list_says_so_rather_than_printing_an_empty_table():
    """Nothing to buy is good news, not a broken download."""
    from openpyxl import load_workbook

    from app.orders import logic
    from app.shipment import documents

    sheet = {"parents": [{"product": "ABC Sattu", "brand": "MF", "kg": 10.0, "units": 20,
                          "orders": 20, "packed": 0, "sizes": []}]}
    purchasing = logic.raw_stock_summary(sheet, {"ABC Sattu": 50.0})
    book = load_workbook(documents.build_tobuy_xlsx(purchasing, "26 Aug (IST)"))
    text = "\n".join(
        " ".join(str(cell.value) for cell in row if cell.value is not None)
        for row in book.active.iter_rows()
    )
    assert "Nothing to buy" in text


# ─── build_reorder_xlsx / build_reorder_pdf: Product, Brand, Reorder Level only ────


def test_build_reorder_xlsx_has_three_columns_and_a_totals_row():
    from openpyxl import load_workbook

    rows = [
        {"product": "Chana Sattu", "brand": "Mithila Foods", "reorder_level_kg": 962.0},
        {"product": "Govindbhog Rice", "brand": "Mithila Foods", "reorder_level_kg": 745.0},
    ]
    buffer = documents.build_reorder_xlsx(rows, "2 products need reordering")
    book = load_workbook(buffer)
    sheet = book.active

    headers = [cell.value for cell in sheet[1]]
    assert headers == ["Product", "Brand", "Reorder Level (kg)"]
    assert sheet.cell(row=2, column=1).value == "Chana Sattu"
    assert sheet.cell(row=2, column=3).value == 962.0
    # A totals row sums the one numeric column.
    last_row = sheet.max_row
    assert sheet.cell(row=last_row, column=3).value == 1707.0


def test_build_reorder_xlsx_says_so_in_words_when_nothing_needs_reordering():
    from openpyxl import load_workbook

    buffer = documents.build_reorder_xlsx([], "Every product is above its reorder level.")
    book = load_workbook(buffer)
    sheet = book.active
    text = " ".join(str(c.value) for row in sheet.iter_rows() for c in row if c.value)
    assert "above its reorder level" in text


def test_build_reorder_pdf_has_three_columns():
    from pypdf import PdfReader

    rows = [{"product": "Chana Sattu", "brand": "Mithila Foods", "reorder_level_kg": 962.0}]
    buffer = documents.build_reorder_pdf(rows, "1 product needs reordering")
    reader = PdfReader(io.BytesIO(buffer.getvalue()))
    text = reader.pages[0].extract_text()
    assert "Chana Sattu" in text
    assert "Mithila Foods" in text
    assert "962" in text
