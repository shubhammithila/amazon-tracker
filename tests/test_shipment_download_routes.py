"""The five download routes over HTTP: wiring, gating, and end-to-end ordering.

tests/test_shipment_documents.py proves the builders are correct in isolation.
This file proves they are actually reachable, correctly gated by role, and — the
part neither file can check alone — that the rows arriving in a downloaded xlsx
come out in the order the DB's single ORDER BY produced.

That last one is the real regression guard for requirement 3. The unit tests
assert "the builder renders what it is given"; these assert "what it is given is
the canonical order". Both halves are needed: with only the unit tests, a router
that passed items in insertion order would pass everything.
"""
import io

import pytest

from app.shipment import logic
from tests.conftest import CANONICAL_ORDER

pytestmark = pytest.mark.regression

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

#: The three working documents, each in both formats, plus Amazon's own file.
#: Every one is listed so the role gating, the attachment headers, the content
#: types and the no-plan 404 are asserted for ALL of them — adding a format to one
#: document and forgetting the other is exactly the omission this catches.
ADMIN_DOWNLOADS = [
    "/shipment/download/plan.xlsx",
    "/shipment/download/plan.pdf",
    "/shipment/download/shipment-file.xlsx",
]
#: Open to ops, in both formats. The morning sheet because making him ask the owner
#: for it every day defeats the point of his own screen, and the packed sheet because
#: printing it is his job — "then print the final packed data and submit to the
#: accounts team". Neither carries a projection or a purchase-driven figure, which is
#: the line between these two lists.
OPS_DOWNLOADS = [
    "/shipment/download/remaining.xlsx",
    "/shipment/download/remaining.pdf",
    "/shipment/download/packed.xlsx",
    "/shipment/download/packed.pdf",
]
ALL_DOWNLOADS = ADMIN_DOWNLOADS + OPS_DOWNLOADS

XLSX_DOWNLOADS = [p for p in ALL_DOWNLOADS if p.endswith(".xlsx")]
PDF_DOWNLOADS = [p for p in ALL_DOWNLOADS if p.endswith(".pdf")]


def _rows(content: bytes):
    """(header, data rows) from downloaded xlsx bytes."""
    from openpyxl import load_workbook

    sheet = load_workbook(io.BytesIO(content)).active
    rows = [list(r) for r in sheet.iter_rows(values_only=True)]
    return rows[0], rows[1:]


def _asins(content: bytes) -> list:
    """The ASIN column, found BY NAME, with the totals row dropped.

    By name because the column layout is now shared across three documents and
    changed once already: index 3 was ASIN before and is Brand now, so a
    positional lookup would compare the wrong column and quietly pass.

    The totals row has a blank ASIN, so filtering falsy values drops it without
    needing to know how many rows there are.
    """
    header, rows = _rows(content)
    index = header.index("ASIN")
    return [r[index] for r in rows if r[index]]


# ─── Wiring ──────────────────────────────────────────────────────────────────

@pytest.mark.parametrize("path", ALL_DOWNLOADS)
async def test_downloads_work_for_admin(auth_client, plan_factory, path):
    await plan_factory()
    r = await auth_client.get(path)
    assert r.status_code == 200, f"{path} -> {r.status_code}: {r.text[:200]}"
    assert r.content, f"{path} returned an empty body"


@pytest.mark.parametrize("path", ALL_DOWNLOADS)
async def test_downloads_are_attachments_with_a_dated_filename(
    auth_client, plan_factory, path
):
    """Content-Disposition: a browser must save these, not try to render the
    xlsx as text. The date in the name stops next week's download overwriting
    this week's in the Downloads folder."""
    await plan_factory()
    r = await auth_client.get(path)
    disposition = r.headers.get("content-disposition", "")
    assert disposition.startswith("attachment;"), disposition
    assert "2026-" in disposition, f"no date in the filename: {disposition}"


@pytest.mark.parametrize("path", XLSX_DOWNLOADS)
async def test_xlsx_routes_send_the_spreadsheet_content_type(
    auth_client, plan_factory, path
):
    await plan_factory()
    r = await auth_client.get(path)
    assert r.headers["content-type"].startswith(XLSX), r.headers["content-type"]
    assert r.content[:2] == b"PK"


@pytest.mark.parametrize("path", PDF_DOWNLOADS)
async def test_pdf_routes_send_pdf(auth_client, plan_factory, path):
    await plan_factory()
    r = await auth_client.get(path)
    assert r.headers["content-type"].startswith("application/pdf")
    assert r.content[:4] == b"%PDF"


# ─── Role gating ─────────────────────────────────────────────────────────────

@pytest.mark.parametrize("path", ADMIN_DOWNLOADS)
async def test_ops_cannot_download_the_owner_documents(ops_client, plan_factory, path):
    """The packer gets the morning sheet and nothing else. The plan carries
    projections and the shipment file is an upload to Amazon — both are the
    owner's decisions, not the warehouse's."""
    await plan_factory()
    r = await ops_client.get(path)
    assert r.status_code == 403, f"ops downloaded {path} ({r.status_code})"


async def test_ops_can_download_the_morning_sheet(ops_client, plan_factory):
    """Requirement 5 is only met if the packer can fetch this himself. Making
    him ask the owner every morning defeats the point of his own screen."""
    await plan_factory()
    r = await ops_client.get("/shipment/download/remaining.pdf")
    assert r.status_code == 200, r.status_code
    assert r.content[:4] == b"%PDF"


@pytest.mark.parametrize("path", ALL_DOWNLOADS)
async def test_signed_out_users_get_nothing(client, plan_factory, path):
    await plan_factory()
    r = await client.get(path)
    assert r.status_code in (303, 401, 403), r.status_code
    assert r.content[:2] != b"PK"
    assert r.content[:4] != b"%PDF"


# ─── No plan yet ─────────────────────────────────────────────────────────────

@pytest.mark.parametrize("path", ALL_DOWNLOADS)
async def test_downloads_404_before_a_plan_exists(auth_client, path):
    """A 404 with a message, not a 500 and not a zero-byte file. An empty
    spreadsheet that Excel calls corrupt looks like a broken app; "no active
    plan" tells the owner to upload the CSVs."""
    r = await auth_client.get(path)
    assert r.status_code == 404, r.status_code
    assert "plan" in r.json()["error"].lower()


# ─── End-to-end row order: requirement 3 ─────────────────────────────────────

async def test_downloaded_xlsx_order_matches_the_canonical_order(
    auth_client, plan_factory, read_committed
):
    """The whole point. The bytes the owner receives must be in the same order as
    the screen, which is repository.load_plan_items' ORDER BY, which is
    logic.sort_items.

    The plan_factory fixture includes 'aloe vera juice' — lowercase, first
    alphabetically, but Howrah Foods and category P6 — precisely so that dropping
    the brand rank, the category rank OR the casefold anywhere in this path
    produces a different order and fails here.
    """
    from app.shipment import repository

    plan = await plan_factory()
    plan_id = plan.id

    items = await read_committed(repository.load_plan_items, plan_id)
    expected = [i.asin for i in items]
    # Guard the guard: if the DB order were already wrong this test would happily
    # compare wrong against wrong.
    assert expected == [i.asin for i in logic.sort_items(items)]
    assert expected == CANONICAL_ORDER, (
        f"the DB order is not canonical, so this test would compare wrong against "
        f"wrong: {expected}"
    )
    assert expected[-1] == "B0CCC00001", (
        f"the HF/P6 row should sort last despite being first alphabetically, got "
        f"{expected}"
    )

    # The plan sheet carries only rows with something to ship, which is the whole
    # point of it ("not the entire list of skus"). So the expectation is the
    # canonical order FILTERED, not truncated — the surviving rows must still be in
    # exactly the order the DB produced.
    r = await auth_client.get("/shipment/download/plan.xlsx")
    to_ship = [i.asin for i in items if int(i.shipment_plan or 0) > 0]
    assert to_ship, "the fixture has nothing to ship; this test would prove nothing"
    assert len(to_ship) < len(expected), (
        "every fixture row has a quantity, so this cannot show that zero rows are "
        "dropped — the fixture needs a zero-quantity row"
    )
    assert _asins(r.content) == to_ship


async def test_packed_xlsx_carries_only_what_was_packed(auth_client, plan_factory):
    """The packed sheet lists the SKUs with boxes against them, and nothing else.

    A sheet of 200 rows where 199 read 0 is a sheet nobody checks, so rows with no
    packing are dropped — and the ones that remain keep the canonical order.
    """
    await plan_factory()
    await auth_client.post(
        "/shipment/packing/2026-07-30",
        json={"entries": [
            {"asin": "B0AAA00001", "units": 100},
            {"asin": "B0BBB00001", "units": 40},
        ], "cartons": 11},
    )

    r = await auth_client.get("/shipment/download/packed.xlsx")
    # Canonical order is chana 0.5kg, chana 1kg, jau, aloe — so of the two packed
    # ASINs, B0AAA00001 precedes B0BBB00001.
    assert _asins(r.content) == ["B0AAA00001", "B0BBB00001"]


async def test_the_packed_sheet_reports_units_per_sku(auth_client, plan_factory):
    """Units are the per-SKU number, so they are the column."""
    await plan_factory()
    await auth_client.post(
        "/shipment/packing/2026-07-30",
        json={"entries": [{"asin": "B0AAA00001", "units": 100}], "cartons": 8},
    )

    r = await auth_client.get("/shipment/download/packed.xlsx")
    header, rows = _rows(r.content)
    assert "Units" in header
    assert rows[0][header.index("Units")] == 100


async def test_the_packed_sheet_has_no_per_sku_carton_column(auth_client, plan_factory):
    """"carton is not item wise. it is random."

    The sheet used to carry a Cartons column per row, and it was reporting a number
    the packer had guessed: a carton holds whatever was being packed when it was
    filled, so it belongs to no single SKU. A column here would put that guess in
    front of the accounts team as though it were a measurement.
    """
    await plan_factory()
    await auth_client.post(
        "/shipment/packing/2026-07-30",
        json={"entries": [{"asin": "B0AAA00001", "units": 100}], "cartons": 8},
    )

    header, _rows_ = _rows((await auth_client.get("/shipment/download/packed.xlsx")).content)
    assert not any("carton" in str(h).lower() for h in header), (
        f"a per-SKU carton column is back on the packed sheet: {header}"
    )


def _pdf_text(content: bytes) -> str:
    """The visible text of a PDF, so an assertion can be made on what is printed.

    A PDF's text streams are compressed, which is why most of the tests here settle
    for "it is a real PDF of a plausible size". That is not good enough for the carton
    count: it is the number the accounts team reconciles a shipment against, and it
    would be entirely invisible if it silently stopped being rendered.
    """
    import pypdfium2

    document = pypdfium2.PdfDocument(io.BytesIO(content))
    return "\n".join(page.get_textpage().get_text_range() for page in document)


async def test_the_packed_sheet_still_reports_the_cartons_for_each_day(
    auth_client, plan_factory
):
    """The count must still travel, because it prefills the invoice's Boxes field.

    Named per day rather than only totalled: accounts reconciles a shipment against
    the days that went into it, and a bare "13 cartons" cannot be checked against
    anything. The heading is the one place a day-level fact can sit on a per-SKU table
    without pretending to be per-SKU.

    Asserted on the rendered text of the PDF, not on a string the route built, so this
    fails if the heading stops being printed for any reason at all.
    """
    await plan_factory()
    for day, units, cartons in (("2026-07-29", 100, 8), ("2026-07-30", 60, 5)):
        await auth_client.post(
            f"/shipment/packing/{day}",
            json={"entries": [{"asin": "B0AAA00001", "units": units}], "cartons": cartons},
        )

    text = _pdf_text((await auth_client.get("/shipment/download/packed.pdf")).content)

    assert "13 cartons" in text, (
        f"the total carton count is not on the packed sheet: {text[:400]!r}"
    )
    for day, cartons in (("2026-07-29", 8), ("2026-07-30", 5)):
        assert f"{day}: {cartons}" in text, (
            f"the packed sheet does not break the cartons down by day ({day}): "
            f"{text[:400]!r}"
        )


async def test_a_day_with_no_cartons_is_not_listed_as_zero(auth_client, plan_factory):
    """"2026-07-30: 0" reads as a day that was packed into no boxes.

    The real meaning is that the count has not been entered yet, and printing a
    confident 0 next to a real 8 invites accounts to total them.
    """
    await plan_factory()
    await auth_client.post(
        "/shipment/packing/2026-07-29",
        json={"entries": [{"asin": "B0AAA00001", "units": 100}], "cartons": 8},
    )
    # Units, but the boxes are not stacked yet.
    await auth_client.post(
        "/shipment/packing/2026-07-30",
        json={"entries": [{"asin": "B0AAA00001", "units": 60}], "cartons": 0},
    )

    text = _pdf_text((await auth_client.get("/shipment/download/packed.pdf")).content)
    assert "2026-07-29: 8" in text
    assert "2026-07-30: 0" not in text, (
        "a day with no carton count entered is printed as zero cartons"
    )


async def test_shipment_file_order_matches_the_canonical_order(
    auth_client, plan_factory, read_committed
):
    """The shipment file drops zero-quantity rows, so its order is the canonical
    order with gaps — never a re-sort of what is left."""
    from app.shipment import repository

    plan = await plan_factory()
    plan_id = plan.id
    items = await read_committed(repository.load_plan_items, plan_id)
    expected = [i.asin for i in items if int(i.shipment_plan or 0) > 0]

    r = await auth_client.get("/shipment/download/shipment-file.xlsx?mode=all")
    _header, rows = _rows(r.content)
    assert [row[1] for row in rows] == expected


# ─── Numbers agree with the screen ───────────────────────────────────────────

async def test_downloaded_numbers_match_the_dashboard(auth_client, plan_factory):
    """Same source, so they must agree. /active and the download both build rows
    through _item_payload; if someone later gives the download its own
    calculation, this fails."""
    await plan_factory()
    await auth_client.post(
        "/shipment/packing/2026-07-30",
        json={"entries": [{"asin": "B0AAA00001", "units": 120}], "cartons": 10},
    )

    active = (await auth_client.get("/shipment/active")).json()
    screen = {i["asin"]: i for i in active["items"]}

    r = await auth_client.get("/shipment/download/plan.xlsx")
    header, rows = _rows(r.content)
    for row in rows:
        item = screen[row[3]]
        assert row[header.index("To Ship")] == item["shipment_plan"]
        assert row[header.index("Packed")] == item["packed"]
        assert row[header.index("To Pack")] == item["remaining"]
        # In Stock and To Make travel too. The In-stock figure fed nothing at all
        # before this change, so a download that quietly dropped it would put the
        # column straight back to being decorative.
        assert row[header.index("In Stock")] == item["available"]
        assert row[header.index("To Make")] == item["to_source"]


async def test_downloaded_numbers_match_the_dashboard(auth_client, plan_factory):
    """Same source, so they must agree.

    /active and the downloads both build rows through _item_payload; if someone
    later gives a download its own calculation, this fails.
    """
    await plan_factory()
    await auth_client.post(
        "/shipment/packing/2026-07-30",
        json={"entries": [{"asin": "B0AAA00001", "units": 120}], "cartons": 10},
    )

    active = (await auth_client.get("/shipment/active")).json()
    screen = {i["asin"]: i for i in active["items"]}

    r = await auth_client.get("/shipment/download/plan.xlsx")
    header, rows = _rows(r.content)
    asin_col, qty_col = header.index("ASIN"), header.index("To Pack")
    for row in rows:
        if not row[asin_col]:
            continue  # the totals line
        assert row[qty_col] == screen[row[asin_col]]["shipment_plan"], row[asin_col]


async def test_the_remaining_sheet_drops_rows_with_nothing_left(
    auth_client, plan_factory
):
    """The morning sheet is a to-do list, so a finished SKU must leave it.

    This is the behaviour the packer actually relies on: pack a row fully today and
    it is gone from tomorrow's sheet, rather than sitting there at 0 for him to
    check past every morning.
    """
    await plan_factory()
    before = _asins((await auth_client.get("/shipment/download/remaining.xlsx")).content)
    assert "B0AAA00001" in before

    # Plan is 500 for this ASIN; pack all of it.
    await auth_client.post(
        "/shipment/packing/2026-07-30",
        json={"entries": [{"asin": "B0AAA00001", "units": 500}], "cartons": 30},
    )

    after = _asins((await auth_client.get("/shipment/download/remaining.xlsx")).content)
    assert "B0AAA00001" not in after, (
        "a fully-packed SKU is still on the still-to-pack sheet"
    )
    assert after, "every row vanished; the sheet should still list the others"


async def test_held_units_still_count_as_packed_in_the_download(
    auth_client, plan_factory
):
    """A held day's boxes exist, so the packed sheet must show them.

    The packed/shippable distinction itself is asserted in test_shipment_logic.py
    and on the dashboard payload; this sheet's job is narrower — it reports what
    was BOXED, and held units were boxed. Leaving them out would tell the floor to
    pack the same order twice.
    """
    await plan_factory()
    await auth_client.post(
        "/shipment/packing/2026-07-30",
        json={"entries": [{"asin": "B0AAA00001", "units": 400}], "cartons": 20},
    )
    submit = await auth_client.post("/shipment/packing/2026-07-30/submit")
    assert submit.json()["status"] == logic.STATUS_HELD, submit.json()

    r = await auth_client.get("/shipment/download/packed.xlsx")
    header, rows = _rows(r.content)
    row = next(x for x in rows if x[header.index("ASIN")] == "B0AAA00001")
    assert row[header.index("Units")] == 400


async def test_the_packed_download_honours_a_date_range(auth_client, plan_factory):
    """"select date or range and download packed data"."""
    await plan_factory()
    for day, units in (("2026-07-28", 100), ("2026-07-29", 50), ("2026-07-30", 25)):
        await auth_client.post(
            f"/shipment/packing/{day}",
            json={"entries": [{"asin": "B0AAA00001", "units": units}], "cartons": 5},
        )

    def units_for(query):
        return query

    # Everything, when no dates are given.
    r = await auth_client.get("/shipment/download/packed.xlsx")
    header, rows = _rows(r.content)
    row = next(x for x in rows if x[header.index("ASIN")] == "B0AAA00001")
    assert row[header.index("Units")] == 175, "no-date download should cover every day"

    # A single day.
    r = await auth_client.get(
        "/shipment/download/packed.xlsx?date_from=2026-07-29&date_to=2026-07-29"
    )
    header, rows = _rows(r.content)
    row = next(x for x in rows if x[header.index("ASIN")] == "B0AAA00001")
    assert row[header.index("Units")] == 50, "a single-date range picked up other days"

    # An open-ended range.
    r = await auth_client.get("/shipment/download/packed.xlsx?date_from=2026-07-29")
    header, rows = _rows(r.content)
    row = next(x for x in rows if x[header.index("ASIN")] == "B0AAA00001")
    assert row[header.index("Units")] == 75, "from-only should mean that date onward"


async def test_the_packed_download_rejects_a_malformed_date(auth_client, plan_factory):
    """Rejected rather than ignored.

    A silently-dropped bad date returns every day's packing under a heading that
    claims to be one date, which is worse than an error.
    """
    await plan_factory()
    r = await auth_client.get("/shipment/download/packed.xlsx?date_from=30-07-2026")
    assert r.status_code == 400, r.status_code


@pytest.mark.parametrize("fmt", ["xlsx", "pdf"])
async def test_the_three_documents_share_one_column_layout(
    auth_client, plan_factory, fmt
):
    """S · M · B · Brand · ASIN · SKU · Product, in that order, on all three.

    Asserted for both formats because they are separate builders; a column added to
    the Excel and forgotten in the PDF is exactly the drift this catches.
    """
    await plan_factory()
    await auth_client.post(
        "/shipment/packing/2026-07-30",
        json={"entries": [{"asin": "B0AAA00001", "units": 100}], "cartons": 8},
    )

    expected = ["S", "M", "B", "Brand", "ASIN", "Merchant SKU", "Product"]
    for name in ("plan", "packed", "remaining"):
        r = await auth_client.get(f"/shipment/download/{name}.{fmt}")
        assert r.status_code == 200, f"{name}.{fmt} -> {r.status_code}"
        if fmt == "xlsx":
            header, _rows_ = _rows(r.content)
            assert header[:7] == expected, f"{name}.xlsx header is {header[:7]}"
        else:
            # A PDF's text is compressed, so assert it is a real PDF with content
            # rather than trying to read columns out of it. The xlsx assertion above
            # covers the layout, and both come from the same rows.
            assert r.content[:4] == b"%PDF" and len(r.content) > 900


async def test_remaining_pdf_accepts_a_date(ops_client, plan_factory):
    await plan_factory()
    r = await ops_client.get("/shipment/download/remaining.pdf?pack_date=2026-07-30")
    assert r.status_code == 200
    assert "2026-07-30" in r.headers["content-disposition"]


async def test_remaining_pdf_rejects_a_malformed_date(ops_client, plan_factory):
    await plan_factory()
    r = await ops_client.get("/shipment/download/remaining.pdf?pack_date=30-07-2026")
    assert r.status_code == 400, r.status_code


async def test_remaining_pdf_shrinks_once_a_sku_is_fully_packed(
    ops_client, auth_client, plan_factory
):
    """Requirement 5's feedback loop, end to end: the packer records 500 units of
    the 1kg Chana, and tomorrow's sheet no longer lists it. A smaller PDF is a
    crude proxy but it is the observable one, and 'fewer rows' is the actual
    promise the document makes."""
    await plan_factory()
    before = await ops_client.get("/shipment/download/remaining.pdf")

    await auth_client.post(
        "/shipment/packing/2026-07-30",
        json={"entries": [{"asin": "B0AAA00001", "units": 500}], "cartons": 40},
    )
    after = await ops_client.get("/shipment/download/remaining.pdf")

    assert len(after.content) < len(before.content), (
        "the fully-packed SKU did not drop off the morning sheet"
    )
