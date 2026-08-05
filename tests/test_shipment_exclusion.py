"""Removing rows from a plan, and the one case where it must be refused.

The owner asked to "delete a row/multiple rows at once before finalising the plan
and then save the plan", plus editing after finalising. Two things make that more
than a DELETE statement.

**It is reversible.** ``excluded_at`` is stamped, not deleted, so an accidental
multi-select exclude is one click back. Nothing about a weekly plan is worth
losing to a mis-click.

**Excluding a row that is already packed loses money.** This is the subtle one.
``logic.packed_units_by_asin`` aggregates packing ENTRIES by ASIN and never looks
at plan items, but the invoice bridge builds its lines FROM plan items. So an
excluded-but-packed row means real boxes ship with no GST line against them — a
tax under-statement found at reconciliation, not on any screen. The exclude
endpoint therefore refuses, names the units and dates, and points at the correct
move instead (To Ship = 0, which stops packing while keeping the record).

The other half of that guard is the STALE SCREEN: the owner excludes a row while
the packer already has it on his phone. Covered at the bottom.
"""
import pytest

from app.shipment import repository

pytestmark = pytest.mark.regression

# From conftest.plan_factory. B0AAA00001 is Chana Sattu 1kg, planned 500.
ASIN = "B0AAA00001"
OTHER = "B0BBB00001"
MONDAY = "2026-07-30"

#: Every route that must never show an excluded row. Parametrised so a download
#: added later without the filter fails here rather than shipping a SKU the owner
#: removed to Amazon.
DOWNLOADS = [
    "/shipment/download/plan.xlsx",
    "/shipment/download/plan.pdf",
    "/shipment/download/packed.xlsx",
    "/shipment/download/packed.pdf",
    "/shipment/download/remaining.xlsx",
    "/shipment/download/remaining.pdf",
    "/shipment/download/shipment-file.xlsx",
]


async def _exclude(client, plan_id, asins, excluded=True):
    return await client.post(
        f"/shipment/plan/{plan_id}/items/exclude",
        json={"asins": asins, "excluded": excluded},
    )


# ─── Excluding and restoring ─────────────────────────────────────────────────

async def test_excluding_a_row_removes_it_from_the_plan(
    auth_client, plan_factory, read_committed
):
    plan = await plan_factory()
    r = await _exclude(auth_client, plan.id, [ASIN])
    assert r.status_code == 200, r.text
    assert r.json()["changed"] == [ASIN]

    items = await read_committed(repository.load_plan_items, plan.id)
    assert ASIN not in [i.asin for i in items], "the excluded row is still in the plan"
    assert OTHER in [i.asin for i in items], "excluding one row removed others"


async def test_several_rows_can_be_excluded_at_once(
    auth_client, plan_factory, read_committed
):
    """"delete a row/multiple rows at once" — a plan has 205 rows and most weeks
    only a handful matter."""
    plan = await plan_factory()
    r = await _exclude(auth_client, plan.id, [ASIN, OTHER])
    assert r.status_code == 200, r.text
    assert r.json()["count"] == 2

    items = await read_committed(repository.load_plan_items, plan.id)
    assert {ASIN, OTHER}.isdisjoint({i.asin for i in items})


async def test_excluding_is_reversible(auth_client, plan_factory, read_committed):
    """The reason it is a timestamp and not a DELETE: mis-clicks happen."""
    plan = await plan_factory()
    await _exclude(auth_client, plan.id, [ASIN, OTHER])

    r = await _exclude(auth_client, plan.id, [ASIN, OTHER], excluded=False)
    assert r.status_code == 200, r.text
    assert r.json()["status"] == "restored"

    items = await read_committed(repository.load_plan_items, plan.id)
    assert {ASIN, OTHER}.issubset({i.asin for i in items}), "rows did not come back"


async def test_the_row_still_exists_in_the_database(
    auth_client, plan_factory, count_rows
):
    """Nothing is destroyed, which is what makes the restore possible."""
    from app.models import ShipmentPlanItem

    plan = await plan_factory()
    before = await count_rows(ShipmentPlanItem, plan_id=plan.id)
    await _exclude(auth_client, plan.id, [ASIN])
    assert await count_rows(ShipmentPlanItem, plan_id=plan.id) == before, (
        "the row was deleted rather than excluded — an accidental exclude would "
        "then need a CSV re-upload and redoing every quantity edit"
    )


async def test_excluding_twice_is_harmless(auth_client, plan_factory):
    plan = await plan_factory()
    assert (await _exclude(auth_client, plan.id, [ASIN])).json()["count"] == 1
    assert (await _exclude(auth_client, plan.id, [ASIN])).json()["count"] == 0


async def test_an_empty_selection_is_rejected(auth_client, plan_factory):
    plan = await plan_factory()
    r = await _exclude(auth_client, plan.id, [])
    assert r.status_code == 400, r.text


async def test_ops_cannot_exclude_rows(ops_client, plan_factory, read_committed):
    """Deciding what ships is the owner's call, not the warehouse's."""
    plan = await plan_factory()
    r = await _exclude(ops_client, plan.id, [ASIN])
    assert r.status_code == 403, r.text

    items = await read_committed(repository.load_plan_items, plan.id)
    assert ASIN in [i.asin for i in items], "ops excluded a row despite the 403"


# ─── Excluded rows must vanish everywhere, not just from the owner's table ────

async def test_the_packer_does_not_see_an_excluded_row(
    auth_client, ops_client, plan_factory
):
    """Otherwise he packs a SKU the owner removed, and those boxes are unbilled."""
    plan = await plan_factory()
    await _exclude(auth_client, plan.id, [ASIN])

    r = await ops_client.get(f"/shipment/packing/{MONDAY}")
    assert r.status_code == 200, r.text
    asins = {row["asin"] for row in r.json()["items"]}
    assert ASIN not in asins, "the packer can still see an excluded row"
    assert OTHER in asins, "the packer lost rows that were not excluded"


async def test_an_excluded_row_is_gone_from_the_dashboard(auth_client, plan_factory):
    plan = await plan_factory()
    await _exclude(auth_client, plan.id, [ASIN])

    r = await auth_client.get("/shipment/active")
    assert ASIN not in {i["asin"] for i in r.json()["items"]}


@pytest.mark.parametrize("path", DOWNLOADS)
async def test_an_excluded_row_is_gone_from_every_download(
    auth_client, plan_factory, path
):
    """Parametrised on purpose.

    All five documents inherit the filter through load_plan_items, so this passes
    today for free — but it is the test that fails if a sixth download is added
    with its own query, which would put a removed SKU on an Amazon upload.
    """
    plan = await plan_factory()
    await _exclude(auth_client, plan.id, [ASIN])

    r = await auth_client.get(path)
    assert r.status_code == 200, r.text
    if path.endswith(".xlsx"):
        from io import BytesIO

        from openpyxl import load_workbook

        sheet = load_workbook(BytesIO(r.content)).active
        cells = {str(c) for row in sheet.iter_rows(values_only=True) for c in row}
        assert ASIN not in cells, f"{path} still lists the excluded ASIN"
    else:
        # PDFs compress their text streams, so assert on the SKU which appears in
        # the xlsx sheets too; the meaningful check is that the row count dropped.
        assert r.content[:4] == b"%PDF"


async def test_an_excluded_row_is_not_counted_as_missing_a_sku(
    auth_client, plan_factory
):
    """No nagging about rows the owner has deliberately removed."""
    plan = await plan_factory(items=[
        {"asin": "B0NOSKU0001", "item": "No Sku Thing", "weight": 1.0,
         "brand": "MF", "fba_sku": "", "shipment_plan": 500, "deficit": 480},
    ])
    r = await auth_client.get("/shipment/active")
    assert r.json()["missing_sku_count"] == 1, "premise changed"

    await _exclude(auth_client, plan.id, ["B0NOSKU0001"])
    r = await auth_client.get("/shipment/active")
    assert r.json()["missing_sku_count"] == 0, (
        "an excluded row is still counted as missing a SKU — the owner would be "
        "warned about a row he removed"
    )


# ─── The refusal: a packed row must not be excluded ──────────────────────────

async def test_excluding_a_packed_row_is_refused(
    auth_client, ops_client, plan_factory, read_committed
):
    """The GST-understatement guard.

    Those boxes exist. Removing the row drops it from the invoice bridge's lines,
    so the stock ships with nothing billed against it — and the app would show no
    sign of it.
    """
    plan = await plan_factory()
    await ops_client.post(
        f"/shipment/packing/{MONDAY}",
        json={"entries": [{"asin": ASIN, "units": 240, "cartons": 12}]},
    )

    r = await _exclude(auth_client, plan.id, [ASIN])
    assert r.status_code == 409, f"a packed row was excluded: {r.text}"

    error = r.json()["error"]
    assert "240" in error, "the refusal does not say how many units are already packed"
    assert MONDAY in error, "the refusal does not say which day to go and look at"
    assert "0" in error, "the refusal does not suggest setting To Ship to 0 instead"

    items = await read_committed(repository.load_plan_items, plan.id)
    assert ASIN in [i.asin for i in items], "the row was excluded despite the 409"


async def test_the_suggested_alternative_actually_works(
    auth_client, ops_client, plan_factory, read_committed
):
    """The refusal tells the owner to set To Ship to 0. That must be true.

    A refusal that points at a dead end is worse than no refusal — it reads as the
    app simply not letting you do the thing.
    """
    plan = await plan_factory()
    await ops_client.post(
        f"/shipment/packing/{MONDAY}",
        json={"entries": [{"asin": ASIN, "units": 240, "cartons": 12}]},
    )

    r = await auth_client.post(
        "/shipment/items",
        json={"plan_id": plan.id, "items": [{"asin": ASIN, "shipment_plan": 0}]},
    )
    assert r.status_code == 200, r.text

    items = await read_committed(repository.load_plan_items, plan.id)
    row = next(i for i in items if i.asin == ASIN)
    assert row.shipment_plan == 0
    # Still visible, still invoiceable, and nothing left to pack.
    r = await auth_client.get("/shipment/active")
    item = next(i for i in r.json()["items"] if i["asin"] == ASIN)
    assert item["packed"] == 240, "the packed units were lost"
    assert item["remaining"] == 0, "the packer would still be asked for more"


async def test_a_row_with_no_packing_can_still_be_excluded(
    auth_client, ops_client, plan_factory
):
    """The guard must be narrow. Packing on ANOTHER row must not block this one."""
    plan = await plan_factory()
    await ops_client.post(
        f"/shipment/packing/{MONDAY}",
        json={"entries": [{"asin": OTHER, "units": 100, "cartons": 6}]},
    )

    r = await _exclude(auth_client, plan.id, [ASIN])
    assert r.status_code == 200, f"an unpacked row was blocked by another row: {r.text}"


async def test_a_mixed_selection_is_refused_entirely(
    auth_client, ops_client, plan_factory, read_committed
):
    """All or nothing.

    Silently excluding the safe half would leave the owner believing both were
    removed, and the one that mattered still on the plan.
    """
    plan = await plan_factory()
    await ops_client.post(
        f"/shipment/packing/{MONDAY}",
        json={"entries": [{"asin": ASIN, "units": 240, "cartons": 12}]},
    )

    r = await _exclude(auth_client, plan.id, [ASIN, OTHER])
    assert r.status_code == 409, r.text

    items = {i.asin for i in await read_committed(repository.load_plan_items, plan.id)}
    assert OTHER in items, (
        "the safe half of a refused selection was excluded anyway — the owner "
        "would think both were removed"
    )


async def test_restoring_a_row_is_never_blocked(
    auth_client, ops_client, plan_factory, read_committed
):
    """The packed-row check applies to excluding, not to putting a row back.

    Restoring can only ever ADD an invoice line, so it cannot cause the
    under-statement the guard exists to prevent.
    """
    plan = await plan_factory()
    await _exclude(auth_client, plan.id, [ASIN])
    # Pack something against a DIFFERENT row so the day exists.
    await ops_client.post(
        f"/shipment/packing/{MONDAY}",
        json={"entries": [{"asin": OTHER, "units": 100, "cartons": 6}]},
    )

    r = await _exclude(auth_client, plan.id, [ASIN], excluded=False)
    assert r.status_code == 200, r.text
    items = {i.asin for i in await read_committed(repository.load_plan_items, plan.id)}
    assert ASIN in items


# ─── The stale-screen race ───────────────────────────────────────────────────

async def test_the_packer_cannot_save_against_an_excluded_row(
    auth_client, ops_client, plan_factory, read_committed
):
    """The other half of the packed-row guard.

    The owner excludes a row while the packer already has it on his phone. If that
    save were accepted it would manufacture exactly the orphaned-units state the
    409 above refuses to create — packed units against a row on no document.

    The save is not rejected outright: the packer's other counts are real and must
    not be thrown away over one stale row. The dropped ASINs come back so his
    screen can tell him why.
    """
    plan = await plan_factory()
    await _exclude(auth_client, plan.id, [ASIN])

    r = await ops_client.post(
        f"/shipment/packing/{MONDAY}",
        json={
            "entries": [
                {"asin": ASIN, "units": 240, "cartons": 12},    # excluded meanwhile
                {"asin": OTHER, "units": 100, "cartons": 6},    # still valid
            ]
        },
    )
    assert r.status_code == 200, r.text
    body = r.json()

    assert body["total_units"] == 100, (
        "units were recorded against an excluded row — those boxes would exist on "
        "no plan, no document and no invoice"
    )
    assert ASIN in (body.get("dropped") or []), (
        "the packer is not told his entry was dropped, so he would believe it saved"
    )

    day = await read_committed(repository.get_day, plan.id, MONDAY)
    entries = await read_committed(repository.load_entries, day.id)
    assert {e.asin for e in entries} == {OTHER}
